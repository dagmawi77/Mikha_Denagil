from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, make_response
import cx_Oracle
import tempfile
from flask import request, redirect, url_for
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
import csv
import io
from fpdf import FPDF
from reportlab.lib.pagesizes import letter
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from openpyxl import Workbook  # To handle Excel export
import pandas as pd
from io import BytesIO
from math import ceil
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, jsonify
from werkzeug.security import generate_password_hash
from werkzeug.security import check_password_hash
from datetime import datetime, timezone
from datetime import datetime, timedelta
import uuid
import io
import csv
from datetime import datetime
from werkzeug.utils import secure_filename
from flask import make_response, flash, redirect, url_for

app = Flask(__name__)
#app.run(host='0.0.0.0',port=5001,debug=True)
app.secret_key = 'your_secret_key'  # Change this to a secure key
app.permanent_session_lifetime = timedelta(minutes=3)  # Set session timeout to 3 minutes
def get_db_connection():
    # conn = cx_Oracle.connect('aawsa_api/aawsa_api@10.12.110.76:1521/apiaawsa')
    conn = cx_Oracle.connect('aawsa_api/aawsa_api@10.12.110.76:1521/apiaawsa')
    return conn

def get_db_connection_billing():
    conn_bill = cx_Oracle.connect('apps/ndf36fgn@10.12.110.14:1522/basis2db.aawsa.local')
    return conn_bill
# Middleware to track session timeout
def track_session_timeout():
    if 'last_activity' in session:
        now = datetime.now()
        last_activity = session['last_activity']
        if (now - last_activity).total_seconds() > 180:  # 3 minutes
            session.clear()  # Clear session if inactive for more than 3 minutes
            flash("Session timed out due to inactivity.", "warning")
            return redirect(url_for('login'))
    session['last_activity'] = datetime.now()  # Update last activity time

@app.before_request
def make_session_permanent():
    session.permanent = True 

# Decorator to protect routes
def session_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'payroll_number' not in session:  # Check if user is logged in
            flash("Please log in to access this page.", "danger")
            return redirect(url_for('login'))
        track_session_timeout()  # Track session timeout on each request
        return f(*args, **kwargs)
    return decorated_function

def role_required(*allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'payroll_number' not in session:
                flash('Please log in to access this page', 'danger')
                return redirect(url_for('login'))

            conn = get_db_connection()
            cursor = conn.cursor()
            try:
                # Get user's role
                cursor.execute("""
                    SELECT r.role_name 
                    FROM aawsa_user u
                    JOIN roles r ON u.role_id = r.role_id
                    WHERE u.payroll_number = :payroll_number
                """, {'payroll_number': session['payroll_number']})
                user_role = cursor.fetchone()
                
                if not user_role or user_role[0] not in allowed_roles:
                    flash('You do not have permission to access this page', 'danger')
                    return redirect(url_for('navigation'))
                
                return f(*args, **kwargs)
            finally:
                cursor.close()
                conn.close()
        return decorated_function
    return decorator

# Example usage:
@app.route('/admin_dashboard')
@role_required('Super Admin')
def admin_dashboard():
    return render_template('admin_dashboard.html')

@app.context_processor
def inject_navigation():
    def get_authorized_routes():
        if 'payroll_number' not in session:
            return []
            
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT r.endpoint, r.route_name
                FROM routes r
                JOIN role_routes rr ON r.route_id = rr.route_id
                JOIN roles ro ON rr.role_id = ro.role_id
                JOIN aawsa_user u ON ro.role_id = u.role_id
                WHERE u.payroll_number = :payroll_number
                ORDER BY r.route_name
            """, {'payroll_number': session['payroll_number']})
            return cursor.fetchall()
        finally:
            cursor.close()
            conn.close()
    
    return dict(authorized_routes=get_authorized_routes())

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        payroll_number = request.form['payroll_number']
        password = request.form['password']

        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get user with hashed password
        cursor.execute("""
            SELECT u.*, r.role_name 
            FROM aawsa_user u
            JOIN roles r ON u.role_id = r.role_id
            WHERE u.payroll_number = :payroll_number
        """, {"payroll_number": payroll_number})
        
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if user and check_password_hash(user[2], password):  # Or better: user['password']
            session['payroll_number'] = payroll_number
            session['role'] = user[-1]  # Or: user['role_name']
            session['last_activity'] = datetime.now().isoformat()
            flash("Login successful!", "success")
            return redirect(url_for('navigation'))
        else:
            flash("Invalid credentials", "danger")

    return render_template('login.html')




app.secret_key = 'your_very_strong_secret_key_here'  # Change this to a secure random key
app.permanent_session_lifetime = timedelta(minutes=30)  # Session expires after 30 minutes

@app.before_request
def check_session():
    # Skip session check for login page and static files
    if request.endpoint in ['login', 'static']:
        return
    
    # Check if user is logged in
    if 'payroll_number' not in session:
        flash('Please log in to access this page', 'danger')
        return redirect(url_for('login'))
    
    # Check session timeout
    last_activity = session.get('last_activity')
    if last_activity:
        # Convert ISO string back to datetime object
        last_activity_dt = datetime.fromisoformat(last_activity)
        if (datetime.now() - last_activity_dt).total_seconds() > 1800:
            session.clear()
            flash('Session timed out due to inactivity', 'warning')
            return redirect(url_for('login'))
    
    # Update last activity time
    session['last_activity'] = datetime.now().isoformat()


# Middleware to protect routes (login required)
def login_required(func):
    def wrapper(*args, **kwargs):
        if 'payroll_number' not in session:
            # Redirect to the login page if not logged in
            return redirect(url_for('login'))
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__  # Ensure the function name is correct for routing
    return wrapper

def initialize_rbac_tables():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if tables exist first
        cursor.execute("""
        DECLARE
            v_count NUMBER;
        BEGIN
            SELECT COUNT(*) INTO v_count FROM user_tables WHERE table_name = 'ROLES';
            IF v_count = 0 THEN
                EXECUTE IMMEDIATE 'CREATE TABLE roles (
                    role_id NUMBER GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
                    role_name VARCHAR2(100) NOT NULL UNIQUE,
                    description VARCHAR2(255)
                )';
            END IF;
            
            SELECT COUNT(*) INTO v_count FROM user_tables WHERE table_name = 'ROUTES';
            IF v_count = 0 THEN
                EXECUTE IMMEDIATE 'CREATE TABLE routes (
                    route_id NUMBER GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
                    route_name VARCHAR2(100) NOT NULL UNIQUE,
                    endpoint VARCHAR2(255) NOT NULL UNIQUE,
                    description VARCHAR2(255)
                )';
            END IF;
            
            SELECT COUNT(*) INTO v_count FROM user_tables WHERE table_name = 'ROLE_ROUTES';
            IF v_count = 0 THEN
                EXECUTE IMMEDIATE 'CREATE TABLE role_routes (
                    role_id NUMBER NOT NULL,
                    route_id NUMBER NOT NULL,
                    PRIMARY KEY (role_id, route_id),
                    FOREIGN KEY (role_id) REFERENCES roles(role_id),
                    FOREIGN KEY (route_id) REFERENCES routes(route_id)
                )';
            END IF;
            
            -- Check if column exists before adding
            SELECT COUNT(*) INTO v_count FROM user_tab_columns 
            WHERE table_name = 'AAWSA_USER' AND column_name = 'ROLE_ID';
            IF v_count = 0 THEN
                EXECUTE IMMEDIATE 'ALTER TABLE aawsa_user ADD (role_id NUMBER)';
                EXECUTE IMMEDIATE 'ALTER TABLE aawsa_user ADD CONSTRAINT fk_user_role 
                                  FOREIGN KEY (role_id) REFERENCES roles(role_id)';
            END IF;
        END;
        """)
        conn.commit()
        print("RBAC tables initialized/verified")
    except Exception as e:
        conn.rollback()
        print(f"Error initializing RBAC tables: {str(e)}")
    finally:
        cursor.close()
        conn.close()

# Call this at application startup
initialize_rbac_tables()
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'payroll_number' not in session:
            flash('Please log in to access this page', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function
@app.context_processor
def inject_navigation():
    def get_authorized_routes():
        if 'payroll_number' not in session:
            return []
            
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            # First check if RBAC tables exist
            cursor.execute("""
            SELECT COUNT(*) FROM user_tables WHERE table_name = 'ROLES'
            """)
            if cursor.fetchone()[0] == 0:
                return []
                
            # Get user's authorized routes
            cursor.execute("""
                SELECT r.endpoint, r.route_name
                FROM routes r
                JOIN role_routes rr ON r.route_id = rr.route_id
                JOIN roles ro ON rr.role_id = ro.role_id
                JOIN aawsa_user u ON ro.role_id = u.role_id
                WHERE u.payroll_number = :payroll_number
                ORDER BY r.route_name
            """, {'payroll_number': session['payroll_number']})
            return cursor.fetchall()
        except cx_Oracle.DatabaseError as e:
            print(f"Error getting authorized routes: {str(e)}")
            return []
        finally:
            cursor.close()
            conn.close()
    
    return dict(authorized_routes=get_authorized_routes())

@app.context_processor
def utility_processor():
    def get_assigned_routes(role_id):
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT r.route_name 
            FROM routes r
            JOIN role_routes rr ON r.route_id = rr.route_id
            WHERE rr.role_id = :role_id
        """, {'role_id': role_id})
        routes = [row[0] for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return routes
    return dict(get_assigned_routes=get_assigned_routes)

@app.route('/manage_roles')
@login_required
# @login_required
@role_required('Super Admin')  # Only Super Admin can access
def manage_roles():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all roles
    cursor.execute("SELECT role_id, role_name, description FROM roles ORDER BY role_name")
    roles = cursor.fetchall()
    
    # Get all routes for the role assignment form
    cursor.execute("SELECT route_id, route_name FROM routes ORDER BY route_name")
    all_routes = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('manage_roles.html', 
                         roles=roles,
                         all_routes=all_routes)

@app.route('/add_role', methods=['POST'])
@login_required
# @login_required
@role_required('Super Admin') 
# @login_required
# @role_required('Super Admin')
def add_role():
    role_name = request.form['role_name']
    description = request.form.get('description', '')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO roles (role_name, description)
            VALUES (:role_name, :description)
        """, {'role_name': role_name, 'description': description})
        conn.commit()
        flash('Role created successfully!', 'success')
    except cx_Oracle.DatabaseError as e:
        conn.rollback()
        flash(f'Error creating role: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('manage_roles'))

@app.route('/update_role_routes/<int:role_id>', methods=['POST'])
@login_required
# @login_required
@role_required('Super Admin') 
# @login_required
# @role_required('Super Admin')
def update_role_routes(role_id):
    selected_routes = request.form.getlist('routes')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Clear existing routes for this role
        cursor.execute("DELETE FROM role_routes WHERE role_id = :role_id", 
                      {'role_id': role_id})
        
        # Add the new selected routes
        for route_id in selected_routes:
            cursor.execute("""
                INSERT INTO role_routes (role_id, route_id)
                VALUES (:role_id, :route_id)
            """, {'role_id': role_id, 'route_id': route_id})
        
        conn.commit()
        flash('Route permissions updated successfully!', 'success')
    except cx_Oracle.DatabaseError as e:
        conn.rollback()
        flash(f'Error updating route permissions: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('manage_roles'))

@app.route('/delete_role/<int:role_id>', methods=['POST'])
@login_required
# @login_required
@role_required('Super Admin') 
# @login_required
# @role_required('Super Admin')
def delete_role(role_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # First check if any users have this role
        cursor.execute("SELECT COUNT(*) FROM aawsa_user WHERE role_id = :role_id", 
                      {'role_id': role_id})
        user_count = cursor.fetchone()[0]
        
        if user_count > 0:
            flash('Cannot delete role - there are users assigned to it!', 'danger')
        else:
            # Delete role routes first
            cursor.execute("DELETE FROM role_routes WHERE role_id = :role_id", 
                          {'role_id': role_id})
            # Then delete the role
            cursor.execute("DELETE FROM roles WHERE role_id = :role_id", 
                          {'role_id': role_id})
            conn.commit()
            flash('Role deleted successfully!', 'success')
    except cx_Oracle.DatabaseError as e:
        conn.rollback()
        flash(f'Error deleting role: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('manage_roles'))

def initialize_default_roles_and_routes():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Create default roles
        default_roles = [
            ('Super Admin', 'Full access to all system features'),
            ('Finance', 'Access to billing, payments, and financial reports'),
            ('Customer Service', 'Access to customer data and support functions'),
            ('Branch Manager', 'Access to branch-specific operations')
        ]
        
        for role_name, description in default_roles:
            cursor.execute("""
                INSERT INTO roles (role_name, description)
                VALUES (:role_name, :description)
            """, {'role_name': role_name, 'description': description})
        
        # Create default routes
        default_routes = [
            ('Dashboard', 'navigation', 'Main dashboard'),
            ('List Paid Bills', 'list_paid_bills', 'View paid bills'),
            ('List Sent Bills', 'list_sent_bills', 'View sent bills'),
            ('List Unsettled Bills', 'list_unsettled_bills', 'View unsettled bills'),
            ('Call Center Unsettled', 'call_center_unsettled', 'Call center unresolved issues'),
            ('Reports', 'report_page', 'View reports'),
            ('Export Data', 'export_form', 'Export data'),
            ('User Management', 'create_user', 'Manage users'),
            ('Role Management', 'manage_roles', 'Manage roles'),
            ('Route Management', 'manage_routes', 'Manage routes'),
            ('Upload Bills', 'upload_bill', 'Upload bill data'),
            ('Upload VAT', 'upload_vat', 'Upload VAT data'),
            ('Move Data', 'move_data', 'Archive old data')
        ]
        
        for route_name, endpoint, description in default_routes:
            cursor.execute("""
                INSERT INTO routes (route_name, endpoint, description)
                VALUES (:route_name, :endpoint, :description)
            """, {
                'route_name': route_name,
                'endpoint': endpoint,
                'description': description
            })
        
        # Assign all routes to Super Admin
        cursor.execute("SELECT role_id FROM roles WHERE role_name = 'Super Admin'")
        super_admin_id = cursor.fetchone()[0]
        
        cursor.execute("SELECT route_id FROM routes")
        all_route_ids = [row[0] for row in cursor.fetchall()]
        
        for route_id in all_route_ids:
            cursor.execute("""
                INSERT INTO role_routes (role_id, route_id)
                VALUES (:role_id, :route_id)
            """, {'role_id': super_admin_id, 'route_id': route_id})
        
        conn.commit()
        print("Default roles and routes initialized successfully")
    except Exception as e:
        conn.rollback()
        print(f"Error initializing default roles and routes: {str(e)}")
    finally:
        cursor.close()
        conn.close()

# Call this function once when setting up your application
initialize_default_roles_and_routes()

@app.route('/roles/<int:role_id>/routes', methods=['GET', 'POST'])
@login_required
# @login_required
@role_required('Super Admin') 
def manage_role_routes(role_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        # Update the routes assigned to this role
        selected_routes = request.form.getlist('routes')
        
        try:
            # First clear existing routes for this role
            cursor.execute("DELETE FROM role_routes WHERE role_id = :role_id", {'role_id': role_id})
            
            # Add the new selected routes
            for route_id in selected_routes:
                cursor.execute("""
                    INSERT INTO role_routes (role_id, route_id)
                    VALUES (:role_id, :route_id)
                """, {'role_id': role_id, 'route_id': route_id})
            
            conn.commit()
            flash('Route permissions updated successfully!', 'success')
        except cx_Oracle.DatabaseError as e:
            conn.rollback()
            flash(f'Error updating route permissions: {str(e)}', 'danger')
        
        return redirect(url_for('manage_role_routes', role_id=role_id))
    
    # GET request - show current routes and available routes
    # Get role info
    cursor.execute("SELECT role_name FROM roles WHERE role_id = :role_id", {'role_id': role_id})
    role = cursor.fetchone()
    
    if not role:
        flash('Role not found', 'danger')
        return redirect(url_for('manage_roles'))
    
    # Get all available routes
    cursor.execute("SELECT route_id, route_name, endpoint FROM routes ORDER BY route_name")
    all_routes = cursor.fetchall()
    
    # Get routes currently assigned to this role
    cursor.execute("""
        SELECT r.route_id, r.route_name 
        FROM routes r
        JOIN role_routes rr ON r.route_id = rr.route_id
        WHERE rr.role_id = :role_id
    """, {'role_id': role_id})
    assigned_routes = {row[0] for row in cursor.fetchall()}
    
    cursor.close()
    conn.close()
    
    return render_template('manage_role_routes.html', 
                         role={'role_id': role_id, 'role_name': role[0]},
                         all_routes=all_routes,
                         assigned_routes=assigned_routes)

@app.route('/routes', methods=['GET', 'POST'])
@login_required
# @login_required
@role_required('Super Admin') 
# @login_required
def manage_routes():
    if request.method == 'POST':
        route_name = request.form['route_name']
        endpoint = request.form['endpoint']
        description = request.form.get('description', '')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO routes (route_name, endpoint, description)
                VALUES (:route_name, :endpoint, :description)
            """, {
                'route_name': route_name,
                'endpoint': endpoint,
                'description': description
            })
            conn.commit()
            flash('Route added successfully!', 'success')
        except cx_Oracle.DatabaseError as e:
            conn.rollback()
            flash(f'Error adding route: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
        
        return redirect(url_for('manage_routes'))
    
    # GET request - show all routes
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT route_id, route_name, endpoint, description FROM routes")
    routes = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return render_template('manage_routes.html', routes=routes)


# Middleware to protect routes (login required)
def login_required(func):
    def wrapper(*args, **kwargs):
        if 'payroll_number' not in session:
            # Redirect to the login page if not logged in
            return redirect(url_for('login'))
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__  # Ensure the function name is correct for routing
    return wrapper

@app.route('/navigation')
@login_required
def navigation():
    connection = get_db_connection()
    
    # Initialize all variables with default values
    recent_sent_bills = []
    total_bills = [0, 'N/A', 0, 0]  # Default values for count, reason, total, current month
    payment_total = [0, 'N/A', 0, 0]
    comparison = [0, 0, 0, 0]  # paid, sent, number_sent, number_paid
    branch_stats = []
    payment_by_channel = []
    current_settled = []
    current_delivery = []
    current_unsettled = []
    
    try:
        cursor = connection.cursor()
        
        # Query for recent sent bills
        cursor.execute("""
            SELECT b.BILLKEY, b.CUSTOMERNAME, b.TOTALBILLAMOUNT, p.AMOUNT, p.PAYMENTCOMPLETEDAT
            FROM BILL b
            JOIN PAYMENT p ON b.BILLKEY = p.BILLKEY
            
            ORDER BY p.PAYMENTCOMPLETEDAT DESC
            FETCH FIRST 10 ROWS ONLY
        """)
        recent_sent_bills = cursor.fetchall() or []  # Default to empty list if None

        cursor.execute("""
                    
    select CUSTOMERBRANCH,REASON,count(*) number_bill,TO_CHAR(sum(CONS),'FM999,999,999'),TO_CHAR(sum(THISMONTHBILLAMT),'FM999,999,999') current_bill 
    ,TO_CHAR(sum(OUTSTANDINGAMT),'FM999,999,999') outstanding_bill ,TO_CHAR(sum(TOTALBILLAMOUNT),'FM999,999,999') total_bill_amount from bill 
    where billkey in (select billkey from payment) and  REASON!='March-2025' group by CUSTOMERBRANCH,REASON order by 
    TOTAL_BILL_AMOUNT desc
        """)
        current_settled = cursor.fetchall() or []
        cursor.execute("""
                    
    select CUSTOMERBRANCH,REASON,count(*) number_bill,TO_CHAR(sum(CONS),'FM999,999,999'),TO_CHAR(sum(THISMONTHBILLAMT),'FM999,999,999') current_bill 
    ,TO_CHAR(sum(OUTSTANDINGAMT),'FM999,999,999') outstanding_bill ,TO_CHAR(sum(TOTALBILLAMOUNT),'FM999,999,999') total_bill_amount from bill 
    where   REASON!='March-2025' group by CUSTOMERBRANCH,REASON order by 
    TOTAL_BILL_AMOUNT desc
        """)
        current_delivery = cursor.fetchall() or []
        cursor.execute("""
                    
    select CUSTOMERBRANCH,REASON,count(*) number_bill,TO_CHAR(sum(CONS),'FM999,999,999'),TO_CHAR(sum(THISMONTHBILLAMT),'FM999,999,999') current_bill 
    ,TO_CHAR(sum(OUTSTANDINGAMT),'FM999,999,999') outstanding_bill ,TO_CHAR(sum(TOTALBILLAMOUNT),'FM999,999,999') total_bill_amount from bill 
    where  billkey not in (select billkey from payment) and REASON!='March-2025' group by CUSTOMERBRANCH,REASON order by 
    TOTAL_BILL_AMOUNT desc
        """)
        current_unsettled = cursor.fetchall() or []
        # Query for bill statistics
        cursor.execute(""" 
            select TO_CHAR(Count(*),'FM999,999,999'),REASON,
                   TO_CHAR(sum(TOTALBILLAMOUNT),'999,999,999.00') as TOTALBILLAMOUNT,
                   TO_CHAR(sum(THISMONTHBILLAMT),'999,999,999.00') as THISMONTHBILLAMT  
            from BILL 
            group by REASON
        """)
        total_bills_result = cursor.fetchone()
        if total_bills_result:
            total_bills = list(total_bills_result)

        # Query for payment statistics
        cursor.execute(""" 
            select TO_CHAR(count(*),'FM999,999,999') no_bills,REASON,
                   TO_CHAR(sum(TOTALBILLAMOUNT),'999,999,999.00') as TOTALBILLAMOUNT,
                   TO_CHAR(sum(THISMONTHBILLAMT),'999,999,999.00') THISMONTHBILLAMT 
            from payment a, bill b 
            where a.billkey = b.billkey 
            group by REASON  
        """)
        payment_result = cursor.fetchone()
        if payment_result:
            payment_total = list(payment_result)
           # payment_total[0] = int(payment_total[0].replace(',', '')) if payment_total[0] else 0
           # payment_total[2] = float(payment_total[2].replace(',', '')) if payment_total[2] else 0
            #payment_total[3] = float(payment_total[3].replace(',', '')) if payment_total[3] else 0

        # Query for sent vs paid comparison
        cursor.execute("""
            SELECT 
                COALESCE(SUM(CASE WHEN p.PAYMENTDATE IS NOT NULL THEN p.AMOUNT ELSE 0 END), 0) AS total_paid_amount,
                COALESCE(SUM(b.TOTALBILLAMOUNT), 0) AS total_sent_amount,
                COALESCE((COUNT(b.BILLKEY) - COUNT(DISTINCT p.BILLKEY)), 0) AS number_sent,
                COALESCE(COUNT(DISTINCT p.BILLKEY), 0) AS number_paid
            FROM BILL b
            LEFT JOIN PAYMENT p ON b.BILLKEY = p.BILLKEY
        """)
        comparison_result = cursor.fetchone()
        if comparison_result:
            comparison = list(comparison_result)

        # Query for branch-based sent data
        cursor.execute("""
            SELECT CUSTOMERBRANCH, 
                   COUNT(b.BILLKEY) AS number_sent, 
                   SUM(b.TOTALBILLAMOUNT) AS total_amount_sent,
                   SUM(OUTSTANDINGAMT) as outstanding_amount,
                   SUM(THISMONTHBILLAMT) as THISMONTHBILLAMT
            FROM BILL b
            GROUP BY CUSTOMERBRANCH
            ORDER BY CUSTOMERBRANCH
        """)
        branch_stats = cursor.fetchall() or []

        # Query for payment channels
        cursor.execute("""
             SELECT case when p.PAYMENTCHANNEL = 'QueryBill' then 'CBEBIRR'
    else PAYMENTCHANNEL end as PAYMENTCHANNEL
    , COALESCE(SUM(p.AMOUNT), 0) AS total_amount
            FROM PAYMENT p
            GROUP BY p.PAYMENTCHANNEL

        """)
        payment_by_channel = cursor.fetchall() or []
        
    except Exception as e:
        app.logger.error(f"Error in navigation route: {str(e)}", exc_info=True)
        # Continue with default values if there's an error
        
    finally:
        connection.close()

    # Extract values with null checks
    total_paid_amount = comparison[0] if len(comparison) > 0 else 0
    total_sent_amount = comparison[1] if len(comparison) > 1 else 0
    number_sent = comparison[2] if len(comparison) > 2 else 0
    number_paid = comparison[3] if len(comparison) > 3 else 0

    
    current_bill_count = total_bills[0] if len(total_bills) > 0 else 0
    bill_period = total_bills[1] if len(total_bills) > 1 else 'N/A'
    current_month_total_bill_amnt = total_bills[2] if len(total_bills) > 2 else 0
    current_month_bill_amnt = total_bills[3] if len(total_bills) > 3 else 0
    
    payment_count = payment_total[0] if len(payment_total) > 0 else 0
    payment_month = payment_total[1] if len(payment_total) > 1 else 'N/A'
    payment_total_bill_amnt = payment_total[2] if len(payment_total) > 2 else 0
    payment_this_month_bill_amnt = payment_total[3] if len(payment_total) > 3 else 0

    # Prepare data for charts with null checks
    branch_labels = [row[0] for row in branch_stats] if branch_stats else []
    number_sent_data = [row[1] for row in branch_stats] if branch_stats else []
    total_amount_sent_data = [row[2] for row in branch_stats] if branch_stats else []
    outstanding_amount = [row[3] for row in branch_stats] if branch_stats else []
    this_month_bill_amnt = [row[4] for row in branch_stats] if branch_stats else []

    channel_labels = [row[0] for row in payment_by_channel] if payment_by_channel else []
    channel_amounts = [row[1] for row in payment_by_channel] if payment_by_channel else []

    return render_template(
        'navigation.html',
        recent_sent_bills=recent_sent_bills,
        current_settled=current_settled,
        current_delivery=current_delivery,
        current_unsettled=current_unsettled,
        total_paid_amount=total_paid_amount,
        total_sent_amount=total_sent_amount,
        total_bills=total_bills,
        number_sent=number_sent,
        current_bill_count=current_bill_count,
        bill_period=bill_period,
        current_month_total_bill_amnt=current_month_total_bill_amnt,
        current_month_bill_amnt=current_month_bill_amnt,
        payment_count=payment_count,
        payment_month=payment_month,
        payment_total_bill_amnt=payment_total_bill_amnt,
        payment_this_month_bill_amnt=payment_this_month_bill_amnt,
        branch_labels=branch_labels,
        number_sent_data=number_sent_data,
        total_amount_sent_data=total_amount_sent_data,
        outstanding_amount=outstanding_amount,
        this_month_bill_amnt=this_month_bill_amnt,
        channel_labels=channel_labels,
        channel_amounts=channel_amounts
    )
@app.route('/dashboard-data')
@login_required
def dashboard_data():
    connection = get_db_connection()
    
    data = {
        # Card metrics
        'current_bill_count': "0",
        'bill_period': 'N/A',
        'current_month_total_bill_amnt': "0",
        'current_month_bill_amnt': "0",
        'payment_count': "0",
        'payment_month': 'N/A',
        'payment_total_bill_amnt': "0",
        'payment_this_month_bill_amnt': "0",
        
        # Branch status tables
        'current_settled': [],
        'current_delivery': [],
        'current_unsettled': [],
        
        # Charts data
        'branch_labels': [],
        'number_sent_data': [],
        'total_amount_sent_data': [],
        'outstanding_amount': [],
        'this_month_bill_amnt': [],
        'channel_labels': [],
        'channel_amounts': []
    }
    
    try:
        cursor = connection.cursor()
        
        # Get bill statistics
        cursor.execute(""" 
            SELECT TO_CHAR(Count(*),'FM999,999,999'), REASON,
                   TO_CHAR(SUM(TOTALBILLAMOUNT),'999,999,999.00'),
                   TO_CHAR(SUM(THISMONTHBILLAMT),'999,999,999.00')
            FROM BILL where REASON!='March-2025'
            GROUP BY REASON
        """)
        total_bills = cursor.fetchone()
        if total_bills:
            data.update({
                'current_bill_count': total_bills[0],
                'bill_period': total_bills[1] or 'N/A',
                'current_month_total_bill_amnt': total_bills[2] or "0",
                'current_month_bill_amnt': total_bills[3] or "0"
            })

        # Get payment statistics
        cursor.execute(""" 
            SELECT TO_CHAR(COUNT(*),'FM999,999,999'), REASON,
                   TO_CHAR(SUM(TOTALBILLAMOUNT),'999,999,999.00'),
                   TO_CHAR(SUM(THISMONTHBILLAMT),'999,999,999.00')
            FROM payment a, bill b 
            WHERE a.billkey = b.billkey  and REASON!='March-2025'
            GROUP BY REASON  
        """)
        payment = cursor.fetchone()
        if payment:
            data.update({
                'payment_count': payment[0],
                'payment_month': payment[1] or 'N/A',
                'payment_total_bill_amnt': payment[2] or "0",
                'payment_this_month_bill_amnt': payment[3] or "0"
            })

        # Get branch-based data for charts
        cursor.execute("""
            SELECT CUSTOMERBRANCH, 
                   COUNT(b.BILLKEY),
                   SUM(b.TOTALBILLAMOUNT),
                   SUM(OUTSTANDINGAMT),
                   SUM(THISMONTHBILLAMT)
            FROM BILL b where REASON!='March-2025'
            GROUP BY CUSTOMERBRANCH
            ORDER BY CUSTOMERBRANCH
        """)
        branch_stats = cursor.fetchall()
        if branch_stats:
            data['branch_labels'] = [row[0] for row in branch_stats]
            data['number_sent_data'] = [row[1] for row in branch_stats]
            data['total_amount_sent_data'] = [row[2] for row in branch_stats]
            data['outstanding_amount'] = [row[3] for row in branch_stats]
            data['this_month_bill_amnt'] = [row[4] for row in branch_stats]

        # Get payment channels data
        cursor.execute("""
            SELECT CASE WHEN p.PAYMENTCHANNEL = 'QueryBill' THEN 'CBEBIRR'
                   ELSE PAYMENTCHANNEL END,
                   COALESCE(SUM(p.AMOUNT), 0)
            FROM PAYMENT p
            GROUP BY p.PAYMENTCHANNEL
        """)
        payment_by_channel = cursor.fetchall()
        if payment_by_channel:
            data['channel_labels'] = [row[0] for row in payment_by_channel]
            data['channel_amounts'] = [float(row[1]) for row in payment_by_channel]

        # Get settled bills
        cursor.execute("""
            SELECT CUSTOMERBRANCH, REASON, 
                   TO_CHAR(COUNT(*),'FM999,999,999'),
                   TO_CHAR(SUM(CONS),'FM999,999,999'),
                   TO_CHAR(SUM(THISMONTHBILLAMT),'FM999,999,999'),
                   TO_CHAR(SUM(OUTSTANDINGAMT),'FM999,999,999'),
                   TO_CHAR(SUM(TOTALBILLAMOUNT),'FM999,999,999')
            FROM bill 
            WHERE billkey IN (SELECT billkey FROM payment) and REASON!='March-2025'
            GROUP BY CUSTOMERBRANCH, REASON
        """)
        data['current_settled'] = [list(row) for row in cursor.fetchall()] or []

        # Get delivery bills
        cursor.execute("""
            SELECT CUSTOMERBRANCH, REASON, 
                   TO_CHAR(COUNT(*),'FM999,999,999'),
                   TO_CHAR(SUM(CONS),'FM999,999,999'),
                   TO_CHAR(SUM(THISMONTHBILLAMT),'FM999,999,999'),
                   TO_CHAR(SUM(OUTSTANDINGAMT),'FM999,999,999'),
                   TO_CHAR(SUM(TOTALBILLAMOUNT),'FM999,999,999')
            FROM bill where REASON!='March-2025'
            GROUP BY CUSTOMERBRANCH, REASON
        """)
        data['current_delivery'] = [list(row) for row in cursor.fetchall()] or []

        # Get unsettled bills
        cursor.execute("""
            SELECT CUSTOMERBRANCH, REASON, 
                   TO_CHAR(COUNT(*),'FM999,999,999'),
                   TO_CHAR(SUM(CONS),'FM999,999,999'),
                   TO_CHAR(SUM(THISMONTHBILLAMT),'FM999,999,999'),
                   TO_CHAR(SUM(OUTSTANDINGAMT),'FM999,999,999'),
                   TO_CHAR(SUM(TOTALBILLAMOUNT),'FM999,999,999')
            FROM bill 
            WHERE billkey NOT IN (SELECT billkey FROM payment) and REASON!='March-2025'
            GROUP BY CUSTOMERBRANCH, REASON
        """)
        data['current_unsettled'] = [list(row) for row in cursor.fetchall()] or []
        
    except Exception as e:
        app.logger.error(f"Error in dashboard-data route: {str(e)}", exc_info=True)
        data['error'] = True
        data['message'] = str(e)
        
    finally:
        connection.close()

    return jsonify(data)
@app.route('/start_payment')
@login_required
# @login_required
@role_required('Super Admin') 
def start_payment():
    return render_template('start_payment.html')

@app.route('/end_payment')
@login_required
# @login_required
@role_required('Super Admin') 
def end_payment():
    return render_template('end_payment.html')

@app.route('/export', methods=['GET'])
@login_required
# @login_required
@role_required('Super Admin') 
def export_form():
    # Fetch available branches and periods from the database
    branches = []
    periods = []
    
    with get_db_connection_billing() as conn:
        cursor = conn.cursor()
        
        # Fetch distinct branches
        cursor.execute("SELECT DISTINCT branch FROM AAWSA_List_of_branch")
        branches = cursor.fetchall()
        
        # Fetch distinct periods
        cursor.execute("SELECT DISTINCT period FROM AAWSA_List_of_period")
        periods = cursor.fetchall()
    
    return render_template('old_delivery.html', branches=branches, periods=periods)

@app.route('/export', methods=['POST'])
@login_required
def export_data():
    # Access form data safely
    branch = request.form.get('branch')
    period = request.form.get('period')
    file_format = request.form.get('file_format')

    # Check if all required fields are provided
    if not branch or not period or not file_format:
        return "All fields are required.", 400

    # Query the database for the selected branch and period
    query = """
    SELECT * FROM aawsa_2021all_delivery
    WHERE branch = :branch AND period = :period
    """
    
    with get_db_connection_billing() as conn:
        data = pd.read_sql(query, conn, params={'branch': branch, 'period': period})

    # Determine the file format and export accordingly
    if file_format == 'csv':
        output_file = f"{branch}_bills_{period}.csv"
        data.to_csv(output_file, index=False)
        return send_file(output_file, as_attachment=True)
    
    elif file_format == 'excel':
        output_file = f"{branch}_bills_{period}.xlsx"
        data.to_excel(output_file, index=False)
        return send_file(output_file, as_attachment=True)

    return "Invalid file format", 400


@app.route('/export_settled', methods=['GET'])
@login_required
def export_form_settled():
    # Fetch available branches and periods from the database
    branches = []
    periods = []
    
    with get_db_connection_billing() as conn:
        cursor = conn.cursor()
        
        # Fetch distinct branches
        cursor.execute("SELECT DISTINCT branch FROM AAWSA_List_of_branch")
        branches = cursor.fetchall()
        
        # Fetch distinct periods
        cursor.execute("SELECT DISTINCT period FROM AAWSA_List_of_period")
        periods = cursor.fetchall()
    
    return render_template('old_settled.html', branches=branches, periods=periods)

@app.route('/export_settled', methods=['POST'])
@login_required
def export_data_settled():
    # Access form data safely
    branch = request.form.get('branch')
    period = request.form.get('period')
    file_format = request.form.get('file_format')

    # Check if all required fields are provided
    if not branch or not period or not file_format:
        return "All fields are required.", 400

    # Query the database for the selected branch and period
    query = """
    SELECT * FROM aawsa_2021all_delivery
    WHERE branch = :branch AND period = :period and bill_key in (select bill_key from aawsa_2021all_settled)
    """
    
    with get_db_connection_billing() as conn:
        data = pd.read_sql(query, conn, params={'branch': branch, 'period': period})

    # Determine the file format and export accordingly
    if file_format == 'csv':
        output_file = f"{branch}_bills_{period}.csv"
        data.to_csv(output_file, index=False)
        return send_file(output_file, as_attachment=True)
    
    elif file_format == 'excel':
        output_file = f"{branch}_bills_{period}.xlsx"
        data.to_excel(output_file, index=False)
        return send_file(output_file, as_attachment=True)

    return "Invalid file format", 400
@app.route('/export_unsettled', methods=['GET'])
def export_form_unsettled():
    # Fetch available branches and periods from the database
    branches = []
    periods = []
    
    with get_db_connection_billing() as conn:
        cursor = conn.cursor()
        
        # Fetch distinct branches
        cursor.execute("SELECT DISTINCT branch FROM AAWSA_List_of_branch")
        branches = cursor.fetchall()
        
        # Fetch distinct periods
        cursor.execute("SELECT DISTINCT period from AAWSA_List_of_period")
        periods = cursor.fetchall()
    
    return render_template('old_settled.html', branches=branches, periods=periods)

@app.route('/export_unsettled', methods=['POST'])
@login_required
def export_data_unsettled():
    # Access form data safely
    branch = request.form.get('branch')
    period = request.form.get('period')
    file_format = request.form.get('file_format')

    # Check if all required fields are provided
    if not branch or not period or not file_format:
        return "All fields are required.", 400

    # Query the database for the selected branch and period
    query = """
    SELECT * FROM aawsa_2021all_delivery
    WHERE branch = :branch AND period = :period and bill_key not in (select bill_key from aawsa_2021all_settled)
    """
    
    with export_data_unsettled() as conn:
        data = pd.read_sql(query, conn, params={'branch': branch, 'period': period})

    # Determine the file format and export accordingly
    if file_format == 'csv':
        output_file = f"{branch}_bills_{period}.csv"
        data.to_csv(output_file, index=False)
        return send_file(output_file, as_attachment=True)
    
    elif file_format == 'excel':
        output_file = f"{branch}_bills_{period}.xlsx"
        data.to_excel(output_file, index=False)
        return send_file(output_file, as_attachment=True)

    return "Invalid file format", 400
@app.route('/list_sent_bills', methods=['GET', 'POST'])
@login_required
def list_sent_bills():
    search_query = request.args.get('search', '')
    selected_branch = request.args.get('branch', 'All')  # Get the selected branch from query parameters
    export_format = request.args.get('format', None)  # Get export format from query parameters
    page = request.args.get('page', 1, type=int)  # Get the current page number, default to 1
    per_page = 300  # Number of items per page

    if request.method == 'POST':
        search_query = request.form.get('search', '')

    # Database connection
    conn = get_db_connection()
    cursor = conn.cursor()

    # Base query for filtering
    query = """
     Select BILLKEY,
CUSTOMERKEY,
CUSTOMERNAME,
CUSTOMERTIN,
CUSTOMERBRANCH,
REASON,
CURRREAD,
PREVREAD,
CONS,
TOTALBILLAMOUNT,
THISMONTHBILLAMT,
OUTSTANDINGAMT,
PENALTYAMT,
DRACCTNO,
CRACCTNO from bill a
    """

    params = {}

    # Add search filter if search_query is provided
    if search_query:
        query += " WHERE (a.CUSTOMERKEY LIKE :search OR a.BILLKEY LIKE :search)"
        params['search'] = f'%{search_query}%'

    # Add branch filter if a branch is selected
    if selected_branch != 'All':
        query += "Where  a.CUSTOMERBRANCH = :branch"
        params['branch'] = selected_branch

    # Finalize query order

    # Execute the query to fetch filtered records for export and pagination
    cursor.execute(query, params)
    print(query)
    filtered_records = cursor.fetchall()  # Now this contains only filtered results
    # print(filtered_records)

    # Pagination logic
    total = len(filtered_records)  # Total number of filtered records
    pages = ceil(total / per_page)  # Total number of pages
    start = (page - 1) * per_page  # Start index for the current page
    end = start + per_page  # End index for the current page

    # Fetch only the rows for the current page (pagination)
    payments_paginated = filtered_records[start:end]

    # Export functionality: pass the filtered data (filtered_records) to export functions
    if export_format == 'csv':
        return export_csv(filtered_records)  # Export filtered data
    elif export_format == 'pdf':
        return export_pdf(filtered_records)  # Export filtered data
    elif export_format == 'excel':
        return export_excel(filtered_records)  # Export filtered data

    # Close database connection
    cursor.close()
    conn.close()

    # Fetch branches for the filter dropdown using get_branches()
    branches = get_branches()

    # Render the template and pass paginated payments data
    return render_template(
        'list_sent_bills.html',
        delivery=payments_paginated,
        search_query=search_query,
        branches=branches,
        selected_branch=selected_branch,
        # start_date=start_date,
        # end_date=end_date,
        page=page,
        pages=pages
    )


@app.route('/list_unsettled_bills', methods=['GET', 'POST'])
@login_required
def list_unsettled_bills():
    search_query = request.args.get('search', '')
    selected_branch = request.args.get('branch', 'All')  # Get the selected branch from query parameters
    export_format = request.args.get('format', None)  # Get export format from query parameters
    page = request.args.get('page', 1, type=int)  # Get the current page number, default to 1
    per_page = 300  # Number of items per page

    if request.method == 'POST':
        search_query = request.form.get('search', '')

    # Database connection
    conn = get_db_connection()
    cursor = conn.cursor()

    # Base query for filtering
    query = """
     Select BILLKEY,
CUSTOMERKEY,
CUSTOMERNAME,
CUSTOMERTIN,
CUSTOMERBRANCH,
REASON,
CURRREAD,
PREVREAD,
CONS,
TOTALBILLAMOUNT,
THISMONTHBILLAMT,
OUTSTANDINGAMT,
PENALTYAMT,
DRACCTNO,
CRACCTNO,debit_30,debit_30_60,debit_60 from bill a,delivery b where a.billkey=b.bill_key and BILLKEY not in (select BILLKEY from payment)
    """

    params = {}

    # Add search filter if search_query is provided
    if search_query:
        query += " AND (a.CUSTOMERKEY LIKE :search OR a.BILLKEY LIKE :search)"
        params['search'] = f'%{search_query}%'

    # Add branch filter if a branch is selected
    if selected_branch != 'All':
        query += "AND  a.CUSTOMERBRANCH = :branch"
        params['branch'] = selected_branch
    # Finalize query order

    # Execute the query to fetch filtered records for export and pagination
    cursor.execute(query, params)
    print(query)
    filtered_records = cursor.fetchall()  # Now this contains only filtered results
    # print(filtered_records)

    # Pagination logic
    total = len(filtered_records)  # Total number of filtered records
    pages = ceil(total / per_page)  # Total number of pages
    start = (page - 1) * per_page  # Start index for the current page
    end = start + per_page  # End index for the current page

    # Fetch only the rows for the current page (pagination)
    payments_paginated = filtered_records[start:end]

    # Export functionality: pass the filtered data (filtered_records) to export functions
    if export_format == 'csv':
        return export_csv(filtered_records)  # Export filtered data
    elif export_format == 'pdf':
        return export_pdf(filtered_records)  # Export filtered data
    elif export_format == 'excel':
        return export_excel(filtered_records)  # Export filtered data

    # Close database connection
    cursor.close()
    conn.close()

    # Fetch branches for the filter dropdown using get_branches()
    branches = get_branches()

    # Render the template and pass paginated payments data
    return render_template(
        'list_unsettled_bills.html',
        delivery=payments_paginated,
        search_query=search_query,
        branches=branches,
        selected_branch=selected_branch,
        # start_date=start_date,
        # end_date=end_date,
        page=page,
        pages=pages
    )



@app.route('/call_center_unsettled', methods=['GET', 'POST'])
@login_required
def call_center_unsettled():
    search_query = request.args.get('search', '')
    selected_branch = request.args.get('branch', 'All')  # Get the selected branch from query parameters
    export_format = request.args.get('format', None)  # Get export format from query parameters
    page = request.args.get('page', 1, type=int)  # Get the current page number, default to 1
    per_page = 300  # Number of items per page

    if request.method == 'POST':
        search_query = request.form.get('search', '')

    # Database connection
    conn = get_db_connection()
    cursor = conn.cursor()

    # Base query for filtering
    query = """
     Select BILLKEY,
a.CUSTOMERKEY,
SUBSTR(CUSTOMERNAME,1,30) as CUSTOMERNAME,
CUSTOMERTIN,
CUSTOMERBRANCH,
REASON,
CURRREAD,
PREVREAD,
CONS,
TOTALBILLAMOUNT,
THISMONTHBILLAMT,
OUTSTANDINGAMT,
PENALTYAMT,
DRACCTNO,
CRACCTNO,PHONENUMBER from bill a,cust_phone b where a.CUSTOMERKEY=b.CUSTOMERKEY and BILLKEY not in (select BILLKEY from payment)
    """

    params = {}

    # Add search filter if search_query is provided
    if search_query:
        query += " AND (a.CUSTOMERKEY LIKE :search OR a.BILLKEY LIKE :search)"
        params['search'] = f'%{search_query}%'

    # Add branch filter if a branch is selected
    if selected_branch != 'All':
        query += "AND  a.CUSTOMERBRANCH = :branch"
        params['branch'] = selected_branch

    # Finalize query order

    # Execute the query to fetch filtered records for export and pagination
    cursor.execute(query, params)
    filtered_records = cursor.fetchall()  # Now this contains only filtered results
    # print(filtered_records)

    # Pagination logic
    total = len(filtered_records)  # Total number of filtered records
    pages = ceil(total / per_page)  # Total number of pages
    start = (page - 1) * per_page  # Start index for the current page
    end = start + per_page  # End index for the current page

    # Fetch only the rows for the current page (pagination)
    payments_paginated = filtered_records[start:end]

    # Export functionality: pass the filtered data (filtered_records) to export functions
    if export_format == 'csv':
        return export_csv(filtered_records)  # Export filtered data
    elif export_format == 'pdf':
        return export_pdf(filtered_records)  # Export filtered data
    elif export_format == 'excel':
        return export_excel(filtered_records)  # Export filtered data

    # Close database connection
    cursor.close()
    conn.close()

    # Fetch branches for the filter dropdown using get_branches()
    branches = get_branches()

    # Render the template and pass paginated payments data
    return render_template(
        'call_center_unsettled.html',
        call_center_unsettled=payments_paginated,
        search_query=search_query,
        branches=branches,
        selected_branch=selected_branch,
        # start_date=start_date,
        # end_date=end_date,
        page=page,
        pages=pages
    )


@app.route('/bill_table_ubs_export', methods=['GET', 'POST'])
@login_required
def bill_table_ubs_export():
    search_query = request.args.get('search', '')
    selected_branch = request.args.get('branch', 'All')  # Get the selected branch from query parameters
    selected_period = request.args.get('period', 'All')  # Get the selected period from query parameters
    export_format = request.args.get('format', None)  # Get export format from query parameters
    page = request.args.get('page', 1, type=int)  # Get the current page number, default to 1
    per_page = 10  # Number of items per page

    if request.method == 'POST':
        search_query = request.form.get('search', '')

    # Database connection
    conn = get_db_connection()
    cursor = conn.cursor()

    # Base query for filtering
    query = """
        SELECT NAME,
               CUST_KEY,
               INST_KEY,
               CONTRACT_NUMBER,
               CHARGE_GROUP,
               ROUTE_KEY,
               READING_DATE,
               BILL_KEY,
               WALK_ORDER,
               BRANCH,
               PERIOD,
               ADDRESS,
               WATER_CONS,
               MAINTAINANCE,
               SEWERAGE,
               SANITATION,
               METER_RENT,
               METER_KEY,
               OLD_METER_CONS,
               LAST_RDG_OLD_MTR,
               PREV_RDG,
               CURR_RDG,
               CURR_CONS,
               TOT_CONS,
               CATEGORY_TYPE,
               BLOCK1_CNS,
               BLOCK2_CNS,
               BLOCK3_CNS,
               BLOCK4_CNS,
               BLOCK5_CNS,
               BLOCK6_CNS,
               BLOCK7_CNS,
               BLOCK1_TRF,
               BLOCK2_TRF,
               BLOCK3_TRF,
               BLOCK4_TRF,
               BLOCK5_TRF,
               BLOCK6_TRF,
               BLOCK7_TRF,
               BLOCK1_AMNT,
               BLOCK2_AMNT,
               BLOCK3_AMNT,
               BLOCK4_AMNT,
               BLOCK5_AMNT,
               BLOCK6_AMNT,
               BLOCK7_AMNT,
               METER_DIAMETER,
               BILL_AMOUNT,
               DEBIT_30,
               DEBIT_30_60,
               DEBIT_60,
               TOT_AMNT,
               CONS_PERIOD,
               DUE_DATE,
               IS_FETCHED,
               SEND_COUNT,
               BILL_ID,
               INST_ID,
               CUST_ID,
               BILL_DATE,
               BILL_FILE_TRAN_ID
        FROM bill_table_ubs
        WHERE 1=1
    """

    params = {}

    # Add search filter if search_query is provided
    if search_query:
        query += " AND (CUST_KEY LIKE :search OR NAME LIKE :search OR BILL_KEY LIKE :search)"
        params['search'] = f'%{search_query}%'

    # Add branch filter if a branch is selected
    if selected_branch != 'All':
        query += " AND BRANCH = :branch"
        params['branch'] = selected_branch

    # Add period filter if a period is selected
    if selected_period != 'All':
        query += " AND PERIOD = :period"
        params['period'] = selected_period

    # Execute the query to fetch filtered records for export and pagination
    cursor.execute(query, params)
    filtered_records = cursor.fetchall()

    # Pagination logic
    total = len(filtered_records)
    pages = ceil(total / per_page)  # Total number of pages
    start = (page - 1) * per_page  # Start index for the current page
    end = start + per_page  # End index for the current page

    # Fetch only the rows for the current page (pagination)
    paginated_records = filtered_records[start:end]

    # Export functionality
    if export_format == 'csv':
        return export_csv(filtered_records)  # Export all filtered data
    elif export_format == 'pdf':
        return export_pdf(filtered_records)  # Export all filtered data
    elif export_format == 'excel':
        return export_excel(filtered_records)  # Export all filtered data

    # Close database connection
    cursor.close()
    conn.close()

    # Fetch branches for the filter dropdown using get_branches() and periods with get_periods()
    branches = get_branches_ubs()
    periods = get_periods()

    # Render the template and pass paginated data and filter options
    return render_template(
        'bill_table_ubs.html',
        records=paginated_records,
        search_query=search_query,
        branches=branches,
        periods=periods,
        selected_branch=selected_branch,
        selected_period=selected_period,
        page=page,
        pages=pages
    )

def get_periods():
    # Establish database connection
    conn = get_db_connection()
    cursor = conn.cursor()

    # Query to get distinct periods from the bill_table_ubs
    cursor.execute("SELECT DISTINCT PERIOD FROM bill_table_ubs ORDER BY PERIOD DESC")
    
    # Fetch all the periods
    periods = cursor.fetchall()

    # Close connection
    cursor.close()
    conn.close()

    # Return the list of periods
    return [row[0] for row in periods]  # Assuming the first column contains the PERIOD
@app.route('/list_paid_bills', methods=['GET', 'POST'])
@login_required
def list_paid_bills():
    search_query = request.args.get('search', '')
    selected_branch = request.args.get('branch', 'All')  # Get the selected branch from query parameters
    export_format = request.args.get('format', None)  # Get export format from query parameters
    page = request.args.get('page', 1, type=int)  # Get the current page number, default to 1
    per_page = 50  # Number of items per page
    start_date = request.args.get('start_date', '')  # Get start date for date filtering
    end_date = request.args.get('end_date', '')  # Get end date for date filtering

    if request.method == 'POST':
        search_query = request.form.get('search', '')
        start_date = request.form.get('start_date', '')
        end_date = request.form.get('end_date', '')

    # Database connection
    conn = get_db_connection()
    cursor = conn.cursor()

    # Base query for filtering
    query = """
     SELECT PAYMENTKEY,a.BILLKEY,a.CUSTOMERKEY,a.AMOUNT,BANKTXNREF,PAYMENTCHANNEL,
    PAYMENTDATE,REQUESTID,CHANNEL,CUSTOMERNAME,CUSTOMERTIN,
    CUSTOMERBRANCH,REASON,CURRREAD,PREVREAD,CONS,TOTALBILLAMOUNT,THISMONTHBILLAMT,
    OUTSTANDINGAMT,PENALTYAMT,DRACCTNO,CRACCTNO,PAYMENTCOMPLETEDAT
    FROM payment a, bill b 
    WHERE a.BILLKEY = b.BILLKEY
    """

    params = {}

    # Add search filter if search_query is provided
    if search_query:
        query += " AND (a.CUSTOMERKEY LIKE :search OR a.BILLKEY LIKE :search)"
        params['search'] = f'%{search_query}%'

    # Add branch filter if a branch is selected
    if selected_branch != 'All':
        query += " AND b.CUSTOMERBRANCH = :branch"
        params['branch'] = selected_branch

    if start_date and end_date:
        try:
            # Convert from HTML datetime-local format (with 'T' and possibly missing seconds)
            if 'T' in start_date:
                start_date = start_date.replace('T', ' ') + ':00'
            if 'T' in end_date:
                end_date = end_date.replace('T', ' ') + ':00'

            dt_format = '%Y-%m-%d %H:%M:%S'
            start_date_obj = datetime.strptime(start_date, dt_format)
            end_date_obj = datetime.strptime(end_date, dt_format)

            query += """
            AND a.PAYMENTCOMPLETEDAT 
            BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD HH24:MI:SS') 
            AND TO_DATE(:end_date, 'YYYY-MM-DD HH24:MI:SS')
            """
            params['start_date'] = start_date_obj.strftime(dt_format)
            params['end_date'] = end_date_obj.strftime(dt_format)

        except ValueError:
            return "Invalid datetime format. Please use YYYY-MM-DDTHH:MM from input."


    # Finalize query order
    query += " ORDER BY a.BILLKEY"

    # Execute the query to fetch filtered records for export and pagination
    cursor.execute(query, params)
    filtered_records = cursor.fetchall()  # Now this contains only filtered results

    # Pagination logic
    total = len(filtered_records)  # Total number of filtered records
    pages = ceil(total / per_page)  # Total number of pages
    start = (page - 1) * per_page  # Start index for the current page
    end = start + per_page  # End index for the current page

    # Fetch only the rows for the current page (pagination)
    payments_paginated = filtered_records[start:end]

    # Export functionality: pass the filtered data (filtered_records) to export functions
    if export_format == 'csv':
        return export__paid_csv(filtered_records)  # Export filtered data
    elif export_format == 'pdf':
        return export_pdf(filtered_records)  # Export filtered data
    elif export_format == 'excel':
        return export_excel(filtered_records)  # Export filtered data

    # Close database connection
    cursor.close()
    conn.close()

    # Fetch branches for the filter dropdown using get_branches()
    branches = get_branches()

    # Render the template and pass paginated payments data
    return render_template(
        'list_paid_bills.html',
        payments=payments_paginated,
        search_query=search_query,
        branches=branches,
        selected_branch=selected_branch,
        start_date=start_date,
        end_date=end_date,
        page=page,
        pages=pages
    )

def export__paid_csv(payments):
    # Create an in-memory file
    si = io.StringIO()
    cw = csv.writer(si)

    # Define headers
    headers = [
'PAYMENTKEY','BILLKEY','CUSTOMERKEY','AMOUNT','BANKTXNREF','PAYMENTCHANNEL',
    'PAYMENTDATE','REQUESTID','CHANNEL','CUSTOMERNAME','CUSTOMERTIN',
    'CUSTOMERBRANCH','REASON','CURRREAD','PREVREAD','CONS','TOTALBILLAMOUNT','THISMONTHBILLAMT',
    'OUTSTANDINGAMT','PENALTYAMT','DRACCTNO','CRACCTNO','PAYMENTCOMPLETEDAT'    ]
    
    # Write the header row
    cw.writerow(headers)
    
    # Write the data rows
    for payment in payments:
        cw.writerow(payment)

    # Move the cursor back to the start of the StringIO object
    si.seek(0)

    # Return the file as a CSV response
    return send_file(io.BytesIO(si.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name='paid_bills.csv')
# CSV export function
def export_csv(payments):
    # Create an in-memory file
    si = io.StringIO()
    cw = csv.writer(si)

    # Define headers
    headers = [
'PAYMENTKEY','BILLKEY','CUSTOMERKEY','AMOUNT','BANKTXNREF','PAYMENTCHANNEL',
    'PAYMENTDATE','REQUESTID','CHANNEL','CUSTOMERNAME','CUSTOMERTIN',
    'CUSTOMERBRANCH','REASON','CURRREAD','PREVREAD','CONS','TOTALBILLAMOUNT','THISMONTHBILLAMT',
    'OUTSTANDINGAMT','PENALTYAMT','DRACCTNO','CRACCTNO','PAYMENTCOMPLETEDAT'    ]
    
    # Write the header row
    cw.writerow(headers)
    
    # Write the data rows
    for payment in payments:
        cw.writerow(payment)

    # Move the cursor back to the start of the StringIO object
    si.seek(0)

    # Return the file as a CSV response
    return send_file(io.BytesIO(si.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name='paid_bills.csv')

# PDF export function
def export_pdf(payments):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    data = []
    headers = ['PAYMENTKEY','BILLKEY','CUSTOMERKEY','AMOUNT','BANKTXNREF','PAYMENTCHANNEL',
    'PAYMENTDATE','REQUESTID','CHANNEL','CUSTOMERNAME','CUSTOMERTIN',
    'CUSTOMERBRANCH','REASON','CURRREAD','PREVREAD','CONS','TOTALBILLAMOUNT','THISMONTHBILLAMT',
    'OUTSTANDINGAMT','PENALTYAMT','DRACCTNO','CRACCTNO','PAYMENTCOMPLETEDAT'
]
    data.append(headers)
    
    for payment in payments:
        data.append(payment)

    # Define column widths to ensure data fits
    col_widths = [1.5 * inch] * len(headers)  # Adjust the width as needed
    
    table = Table(data, colWidths=col_widths)
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDSPACE', (0, 1), (-1, -1), 10)  # Adjust word spacing if needed
    ]))

    doc.build([table])
    buffer.seek(0)

    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='paid_bills.pdf')
# Excel export function
def export_excel(deliverys):
    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Bills"

    # Define headers
    headers = ['PAYMENTKEY','BILLKEY','CUSTOMERKEY','AMOUNT','BANKTXNREF','PAYMENTCHANNEL',
    'PAYMENTDATE','REQUESTID','CHANNEL','CUSTOMERNAME','CUSTOMERTIN',
    'CUSTOMERBRANCH','REASON','CURRREAD','PREVREAD','CONS','TOTALBILLAMOUNT','THISMONTHBILLAMT',
    'OUTSTANDINGAMT','PENALTYAMT','DRACCTNO','CRACCTNO','PAYMENTCOMPLETEDAT']
    ws.append(headers)

    # Write the data rows
    for delivery in deliverys:
        ws.append(delivery)

    # Save the workbook to an in-memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Return the file as an Excel response
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='paid_bills.xlsx')

def export_csv(deliverys):
    # Create an in-memory file
    si = io.StringIO()
    cw = csv.writer(si)

    # Define headers
    headers = [
'PAYMENTKEY','BILLKEY','CUSTOMERKEY','AMOUNT','BANKTXNREF','PAYMENTCHANNEL',
    'PAYMENTDATE','REQUESTID','CHANNEL','CUSTOMERNAME','CUSTOMERTIN',
    'CUSTOMERBRANCH','REASON','CURRREAD','PREVREAD','CONS','TOTALBILLAMOUNT','THISMONTHBILLAMT',
    'OUTSTANDINGAMT','PENALTYAMT','DRACCTNO','CRACCTNO','PAYMENTCOMPLETEDAT'
    ]
    
    # Write the header row
    cw.writerow(headers)
    
    # Write the data rows
    for delivery in deliverys:
        cw.writerow(delivery)

    # Move the cursor back to the start of the StringIO object
    si.seek(0)

    # Return the file as a CSV response
    return send_file(io.BytesIO(si.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name='delivery_bills_new.csv')

# PDF export function
def export_pdf(deliverys):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    data = []
    headers = ['PAYMENTKEY','BILLKEY','CUSTOMERKEY','AMOUNT','BANKTXNREF','PAYMENTCHANNEL',
    'PAYMENTDATE','REQUESTID','CHANNEL','CUSTOMERNAME','CUSTOMERTIN',
    'CUSTOMERBRANCH','REASON','CURRREAD','PREVREAD','CONS','TOTALBILLAMOUNT','THISMONTHBILLAMT',
    'OUTSTANDINGAMT','PENALTYAMT','DRACCTNO','CRACCTNO','PAYMENTCOMPLETEDAT'

]
    data.append(headers)
    
    for delivery in deliverys:
        data.append(delivery)

    # Define column widths to ensure data fits
    col_widths = [1.5 * inch] * len(headers)  # Adjust the width as needed
    
    table = Table(data, colWidths=col_widths)
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDSPACE', (0, 1), (-1, -1), 10)  # Adjust word spacing if needed
    ]))

    doc.build([table])
    buffer.seek(0)

    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='delivery_bills.pdf')
# Excel export function
def export_excel(deliverys):
    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Bills"

    # Define headers
    headers = ['PAYMENTKEY',
'PAYMENTKEY','BILLKEY','CUSTOMERKEY','AMOUNT','BANKTXNREF','PAYMENTCHANNEL',
    'PAYMENTDATE','REQUESTID','CHANNEL','CUSTOMERNAME','CUSTOMERTIN',
    'CUSTOMERBRANCH','REASON','CURRREAD','PREVREAD','CONS','TOTALBILLAMOUNT','THISMONTHBILLAMT',
    'OUTSTANDINGAMT','PENALTYAMT','DRACCTNO','CRACCTNO','PAYMENTCOMPLETEDAT']
    ws.append(headers)

    # Write the data rows
    for delivery in deliverys:
        ws.append(delivery)

    # Save the workbook to an in-memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Return the file as an Excel response
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='delivery_bills.xlsx')

def export_excel(call_center_unsettleds):
    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Bills"

    # Define headers
    headers = [
        'PAYMENTKEY','BILLKEY','CUSTOMERKEY','AMOUNT','BANKTXNREF','PAYMENTCHANNEL',
    'PAYMENTDATE','REQUESTID','CHANNEL','CUSTOMERNAME','CUSTOMERTIN',
    'CUSTOMERBRANCH','REASON','CURRREAD','PREVREAD','CONS','TOTALBILLAMOUNT','THISMONTHBILLAMT',
    'OUTSTANDINGAMT','PENALTYAMT','DRACCTNO','CRACCTNO','PAYMENTCOMPLETEDAT'

]
    ws.append(headers)

    # Write the data rows
    for call_center_unsettled in call_center_unsettleds:
        ws.append(call_center_unsettled)

    # Save the workbook to an in-memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Return the file as an Excel response
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Unpaid Bills With Phone Number.xlsx')

def export_csv(call_center_unsettleds):
    # Create an in-memory file
    si = io.StringIO()
    cw = csv.writer(si)

    # Define headers
    headers = [
'PAYMENTKEY','BILLKEY','CUSTOMERKEY','AMOUNT','BANKTXNREF','PAYMENTCHANNEL',
    'PAYMENTDATE','REQUESTID','CHANNEL','CUSTOMERNAME','CUSTOMERTIN',
    'CUSTOMERBRANCH','REASON','CURRREAD','PREVREAD','CONS','TOTALBILLAMOUNT','THISMONTHBILLAMT',
    'OUTSTANDINGAMT','PENALTYAMT','DRACCTNO','CRACCTNO','PAYMENTCOMPLETEDAT'
    
    ]
    
    # Write the header row
    cw.writerow(headers)
    
    # Write the data rows
    for call_center_unsettled in call_center_unsettleds:
        cw.writerow(call_center_unsettled)

    # Move the cursor back to the start of the StringIO object
    si.seek(0)

    # Return the file as a CSV response
    return send_file(io.BytesIO(si.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name='delivery_bills.csv')

# PDF export function
def export_pdf(call_center_unsettleds):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A3), rightMargin=10, leftMargin=10, topMargin=10, bottomMargin=18)
    
    data = []
    headers = [
        'BILLKEY',
'CUSTOMERKEY',
'CUSTOMERNAME',
'CUSTOMERTIN',
'CUSTOMERBRANCH',
'REASON',
'CURRREAD',
'PREVREAD',
'CONS',
'TOTALBILLAMOUNT',
'THISMONTHBILLAMT',
'OUTSTANDINGAMT',
'PENALTYAMT',
'DRACCTNO',
'CRACCTNO',
'PHONENUMBER'
]
    data.append(headers)
    
    for call_center_unsettled in call_center_unsettleds:
        data.append(call_center_unsettled)

    # Define column widths to ensure data fits
    col_widths = [1.5 * inch] * len(headers)  # Adjust the width as needed
    
    table = Table(data, colWidths=col_widths)
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGRO UND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDSPACE', (0, 1), (-1, -1), 10)  # Adjust word spacing if needed
    ]))

    doc.build([table])
    buffer.seek(0)

    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='Unsettled With Phone Number.pdf')



# Excel export function
def export_excel(call_center_unsettleds):
    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Unsettled For Call center Bills"

    # Define headers
    headers = ['BILLKEY',
'CUSTOMERKEY',
'CUSTOMERNAME',
'CUSTOMERTIN',
'CUSTOMERBRANCH',
'REASON',
'CURRREAD',
'PREVREAD',
'CONS',
'TOTALBILLAMOUNT',
'THISMONTHBILLAMT',
'OUTSTANDINGAMT',
'PENALTYAMT',
'DRACCTNO',
'CRACCTNO',
'PHONENUMBER']
    ws.append(headers)

    # Write the data rows
    for call_center_unsettled in call_center_unsettleds:
        ws.append(call_center_unsettled)

    # Save the workbook to an in-memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Return the file as an Excel response
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='unsettled with Phone number.xlsx')


@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')


@app.route('/move_data', methods=['GET', 'POST'])
@login_required
# @login_required
@role_required('Super Admin') 
def move_data():
    if request.method == 'POST':
        payroll_number = request.form['payroll_number']
        password = request.form['password']
        action_type = request.form['action_type']  # Can be 'payment' or 'bill'
        
        connection = get_db_connection()
        cursor = connection.cursor()

        try:
            # Verify payroll_number and password from the 'aawsa_user' table
            cursor.execute("""
                SELECT COUNT(*) FROM aawsa_user WHERE payroll_number = :payroll_number AND password = :password
            """, {'payroll_number': payroll_number, 'password': password})
            
            user_exists = cursor.fetchone()[0]

            if user_exists:
                if action_type == 'payment':
                    # Move payment data
                    cursor.execute("""
                        INSERT INTO PAYMENT_OLD (SELECT * FROM PAYMENT)
                    """)
                    cursor.execute("""
                        DELETE FROM PAYMENT
                    """)
                    connection.commit()
                    return render_template('move_data.html', message="Payment data moved and deleted successfully")
                
                elif action_type == 'bill':
                    # Move bill data
                    cursor.execute("""
                        INSERT INTO BILL_OLD (SELECT * FROM BILL)
                    """)
                    cursor.execute("""
                        DELETE FROM BILL
                    """)
                    connection.commit()
                    return render_template('move_data.html', message="Bill data moved and deleted successfully")
                
                else:
                    return render_template('move_data.html', error="Invalid action type")
            else:
                return render_template('move_data.html', error="Invalid Payroll Number or Password")

        except Exception as e:
            connection.rollback()
            return render_template('move_data.html', error=str(e))

        finally:
            connection.close()

    # If the request method is GET, just render the form page
    return render_template('move_data.html')

def execute_query(query):
    """
    Executes the given query and returns the result.
    """
    connection = cx_Oracle.connect("agresso/agressotest@localhost:1522/orcl2")  # Adjust connection string if needed
    cursor = connection.cursor()  # Create a cursor
    cursor.execute(query)  # Execute the query
    result = cursor.fetchall()  # Fetch all the results
    cursor.close()  # Close the cursor
    connection.close()  # Close the connection
    return result


def get_branches():
    """
    Retrieves the list of distinct branches from the BILL table.
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    query = "SELECT DISTINCT CUSTOMERBRANCH FROM BILL ORDER BY CUSTOMERBRANCH"
    cursor.execute(query)
    branches = cursor.fetchall()
    cursor.close()
    conn.close()
    return [branch[0] for branch in branches]  # Extract the branch names from the tuple

def get_branches_ubs():
    """
    Retrieves the list of distinct branches from the BILL table.
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    query = "SELECT DISTINCT BRANCH FROM bill_table_ubs ORDER BY BRANCH"
    cursor.execute(query)
    branches = cursor.fetchall()
    cursor.close()
    conn.close()
    return [branch[0] for branch in branches] 



@app.route('/reports')
@login_required
def report_page():
    branches = get_branches()  # Get list of branches for the filter

    # Fetching delivery report data
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetching delivery report data
    query_delivery = """
    SELECT CUSTOMERBRANCH AS branch, COUNT(*) AS total_bills,
           SUM(TOTALBILLAMOUNT) AS sum_total_bills,
           SUM(CONS) AS sum_consumption
    FROM BILL
    GROUP BY rollup(CUSTOMERBRANCH)
    """
    cursor.execute(query_delivery)
    delivery_report_data = cursor.fetchall()

    # Fetching paid bills summary data
    query_paid_bills_summary = """
  SELECT CUSTOMERBRANCH as branch,COUNT(*) AS total_paid_bills,
           SUM(TOTALBILLAMOUNT) AS sum_total_bills,
           SUM(CONS) AS sum_consumption
    FROM PAYMENT a, BILL b where a.BILLKEY=b.BILLKEY group by rollup(CUSTOMERBRANCH)
    """
    cursor.execute(query_paid_bills_summary)
    paid_bills_summary_data = cursor.fetchall()

    # Fetching unsettled bills summary data
    query_unsettled_summary = """
    SELECT  CUSTOMERBRANCH as branch,COUNT(*) AS total_unsettled_bills,
           SUM(TOTALBILLAMOUNT) AS sum_total_unsettled_amount,
           SUM(CONS) AS sum_consumption
    FROM BILL b
    where BILLKEY not in (select BILLKEY from payment) group by rollup(CUSTOMERBRANCH)
    """
    cursor.execute(query_unsettled_summary)
    unsettled_summary_data = cursor.fetchall()
    
    cursor.close()
    conn.close()

    # Debugging: Print the fetched data
    # print("Delivery Report Data:", delivery_report_data)
    # print("Paid Bills Summary Data:", paid_bills_summary_data)
    # print("Unsettled Summary Data:", unsettled_summary_data)

    # Format data for the template
    formatted_delivery_data = [
        {'branch': row[0], 'total_bills': row[1], 'sum_total_bills': row[2], 'sum_consumption': row[3]}
        for row in delivery_report_data
    ]

    
    
    formatted_paid_bills_summary = [
        {'branch': row[0],
        'total_paid_bills': row[1],
        'sum_total_paid_amount': row[2],
        'sum_consumption': row[3]
        }
        for row in paid_bills_summary_data
    ]

    formatted_unsettled_summary = [
        {'branch': row[0],
        'total_unsettled_bills': row[1],
        'sum_total_unsettled_amount': row[2],
        'sum_consumption': row[3]}
        for row in unsettled_summary_data
    ]

    return render_template('reports.html', branches=branches, delivery_data=formatted_delivery_data,
                           paid_bills_summary=formatted_paid_bills_summary, unsettled_summary=formatted_unsettled_summary)


@app.route('/report/delivery')
@login_required
def delivery_report():
    # Fetch the data for the delivery report
    delivery_data = get_delivery_report_data()  # You need to write this query to fetch the required data
    return render_template('delivery_report.html', data=delivery_data)

@app.route('/report/paid_bills')
@login_required
def paid_bills_report():
    # Fetch the data for the paid bills report
    paid_bills_data = get_paid_bills_report_data()  # Query to fetch paid bills grouped by branch
    return render_template('paid_bills_report.html', data=paid_bills_data)

@app.route('/report/unsettled_bills')
@login_required
def unsettled_bills_report():
    # Fetch the data for the unsettled bills report (bills not in the payment table)
    unsettled_bills_data = get_unsettled_bills_report_data()
    return render_template('unsettled_bills_report.html', data=unsettled_bills_data)

@app.route('/export/<report_type>/<file_format>')
@login_required
def export_report(report_type, file_format):
    
    # Fetch the data for the report type
    data = get_report_data(report_type)
    
    if data is None or len(data) == 0:
        return "No data available for the specified report type.", 404  # Handle no data case
    
    # Convert list to DataFrame
    if isinstance(data, list):
        data = pd.DataFrame(data)

    if file_format == 'csv':
        output = BytesIO()
        data.columns = ['Branch', 'Total Bills', 'Sum of Total Bills', 'Sum of Consumption']
        data.to_csv(output, index=False)
        output.seek(0)
        return send_file(output, mimetype='text/csv', download_name=f"{report_type}.csv", as_attachment=True)
    
    elif file_format == 'excel':
        output = BytesIO()
        # data.columns = ['Branch', 'Total Bills', 'Sum of Total Bills', 'Sum of Consumption']
        data.columns = ['Branch', 'Total Bills', 'Sum of Total Bills', 'Sum of Consumption']
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:  # Use a context manager
            data.to_excel(writer, index=False)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name=f"{report_type}.xlsx", as_attachment=True)
    elif file_format == 'pdf':
    # Create a PDF document
        pdf = FPDF()
        pdf.add_page()

        # Set title
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, 'Report', 0, 1, 'C')

        # Check DataFrame columns
        print(data.columns.tolist())  # Print column names for debugging

        # Add table header
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(40, 10, 'Branch', 1)
        pdf.cell(40, 10, 'Total Bills', 1)
        pdf.cell(40, 10, 'Sum of Total Bills', 1)
        pdf.cell(40, 10, 'Sum of Consumption', 1)
        pdf.ln()

        # Add data to PDF
        pdf.set_font("Arial", '', 10)
        for index, row in data.iterrows():
            branch_name = row.get('Branch', 'N/A')  # Access column safely
            total_bills = row.get('Total Bills', 'N/A')
            sum_of_total_bills = row.get('Sum of Total Bills', 'N/A')
            sum_of_consumption = row.get('Sum of Consumption', 'N/A')

            pdf.cell(40, 10, str(branch_name), 1)
            pdf.cell(40, 10, str(total_bills), 1)
            pdf.cell(40, 10, str(sum_of_total_bills), 1)
            pdf.cell(40, 10, str(sum_of_consumption), 1)
            pdf.ln()

        # Use a temporary file to save the PDF
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            pdf.output(tmp_file.name)  # Save the PDF to the temporary file
            tmp_file.seek(0)  # Move to the beginning of the temporary file
            # Read the PDF into a BytesIO object
            output = BytesIO(tmp_file.read())

        # Return the PDF as a response
        return send_file(output, as_attachment=True, download_name='report.pdf', mimetype='application/pdf')
        # Handle unsupported file formats

def get_report_data(report_type):
    if report_type == 'delivery':
        return get_delivery_report_data()
    elif report_type == 'paid_bills':
        return get_paid_bills_report_data()
    elif report_type == 'unsettled_bills':
        return get_unsettled_bills_report_data()
    else:
        raise ValueError("Unknown report type")


def get_delivery_report_data():
    query = """
    SELECT CUSTOMERBRANCH AS branch, COUNT(*) AS total_bills,
           SUM(TOTALBILLAMOUNT) AS sum_total_bills,
           SUM(CONS) AS sum_consumption
    FROM BILL
    GROUP BY rollup(CUSTOMERBRANCH)
    """
    results = execute_query(query)  # Execute the query
    if results is None:
        print("No results returned from execute_query for delivery report.")  # Debug print
        return []
    return results  # Ensure this returns a list
    # return execute_query(query)  # Assuming you have a DB query execution function

def get_paid_bills_report_data():
    query = """
     SELECT CUSTOMERBRANCH as branch,COUNT(*) AS total_paid_bills,
           SUM(TOTALBILLAMOUNT) AS sum_total_bills,
           SUM(CONS) AS sum_consumption
    FROM PAYMENT a, BILL b where a.BILLKEY=b.BILLKEY group by rollup(CUSTOMERBRANCH)
    """
    return execute_query(query)

def get_unsettled_bills_report_data():
    query = """
     SELECT  CUSTOMERBRANCH as branch,COUNT(*) AS total_unsettled_bills,
           SUM(TOTALBILLAMOUNT) AS sum_total_unsettled_amount,
           SUM(CONS) AS sum_consumption
    FROM BILL b
    where BILLKEY not in (select BILLKEY from payment) group by rollup(CUSTOMERBRANCH)
    """
    return execute_query(query)


def execute_query(query):
    connection = get_db_connection()  # Get the DB connection
    if connection is None:
        print("Failed to connect to the database.")
        return None

    try:
        cursor = connection.cursor()  # Create a cursor object
        cursor.execute(query)          # Execute the SQL query
        result = cursor.fetchall()     # Fetch all results
        return result
    except Exception as e:
        print(f"Error executing query: {e}")
        return None
    finally:
        cursor.close()                # Ensure the cursor is closed
        connection.close()    

from collections import namedtuple

def make_named_tuple(cursor):
    fields = [col[0] for col in cursor.description]
    Row = namedtuple('Row', fields)
    return Row

@app.route('/user_management')
@login_required
# @login_required
@role_required('Super Admin') 
# @login_required
# @role_required('Super Admin')
def user_management():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT u.USER_ID, u.USERNAME, u.PAYROLL_NUMBER, u.BRANCH, 
               r.ROLE_ID, r.ROLE_NAME
        FROM AAWSA_USER u
        JOIN ROLES r ON u.ROLE_ID = r.ROLE_ID
        ORDER BY u.USERNAME
    """)
    
    # Convert to named tuples
    users = [make_named_tuple(cursor)(*row) for row in cursor]
    
    cursor.execute("SELECT ROLE_ID, ROLE_NAME FROM ROLES ORDER BY ROLE_NAME")
    roles = [make_named_tuple(cursor)(*row) for row in cursor]
    
    cursor.close()
    conn.close()
    
    return render_template('user_management.html', users=users, roles=roles)

# Create User
@app.route('/create_user', methods=['POST'])
@login_required
# @login_required
@role_required('Super Admin') 
# @login_required
# @role_required('Super Admin')
def create_user():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    form_data = {
        'username': request.form['username'],
        'password': generate_password_hash(request.form['password']),  # Hashed password
        'payroll_number': request.form['payroll_number'],
        'branch': request.form['branch'],
        'role_id': request.form['role_id']
    }
    
    try:
        # Check if username or payroll number exists
        cursor.execute("""
            SELECT 1 FROM aawsa_user 
            WHERE username = :username OR payroll_number = :payroll_number
        """, {'username': form_data['username'], 'payroll_number': form_data['payroll_number']})
        
        if cursor.fetchone():
            flash('Username or payroll number already exists!', 'danger')
        else:
            # Verify role exists
            cursor.execute("SELECT 1 FROM roles WHERE role_id = :role_id", {'role_id': form_data['role_id']})
            if not cursor.fetchone():
                flash('Invalid role selected!', 'danger')
            else:
                # Insert new user
                cursor.execute("""
                    INSERT INTO aawsa_user (username, password, payroll_number, branch, role_id)
                    VALUES (:username, :password, :payroll_number, :branch, :role_id)
                """, form_data)
                conn.commit()
                flash('User created successfully!', 'success')
                print(request.form)  # In your create_user route
    except Exception as e:
        conn.rollback()
        flash(f'Error creating user: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('user_management'))

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
# @login_required
@role_required('Super Admin') 
# @login_required
# @role_required('Super Admin')
def edit_user(user_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        try:
            # Get form data
            username = request.form['username']
            payroll_number = request.form['payroll_number']
            branch = request.form['branch']
            role_id = request.form['role_id']
            
            # Update user
            cursor.execute("""
                UPDATE aawsa_user 
                SET username = :username,
                    payroll_number = :payroll_number,
                    branch = :branch,
                    role_id = :role_id
                WHERE user_id = :user_id
            """, {
                'username': username,
                'payroll_number': payroll_number,
                'branch': branch,
                'role_id': role_id,
                'user_id': user_id
            })
            conn.commit()
            flash('User updated successfully!', 'success')
            return redirect(url_for('user_management'))
        except Exception as e:
            conn.rollback()
            flash(f'Error updating user: {str(e)}', 'danger')
    
    # GET request - load user data
    try:
        # Get user data (returns as tuple)
        cursor.execute("""
            SELECT user_id, username, payroll_number, branch, role_id
            FROM aawsa_user
            WHERE user_id = :user_id
        """, {'user_id': user_id})
        user_tuple = cursor.fetchone()
        
        if not user_tuple:
            flash('User not found!', 'danger')
            return redirect(url_for('user_management'))
        
        # Convert tuple to dictionary for easier template access
        user = {
            'user_id': user_tuple[0],
            'username': user_tuple[1],
            'payroll_number': user_tuple[2],
            'branch': user_tuple[3],
            'role_id': user_tuple[4]
        }
        
        # Get all roles
        cursor.execute("SELECT role_id, role_name FROM roles ORDER BY role_name")
        roles_tuples = cursor.fetchall()
        
        # Convert roles tuples to dictionaries
        roles = [{'role_id': r[0], 'role_name': r[1]} for r in roles_tuples]
        
        return render_template('edit_user.html', user=user, roles=roles)
    except Exception as e:
        flash(f'Error loading user data: {str(e)}', 'danger')
        return redirect(url_for('user_management'))
    finally:
        cursor.close()
        conn.close()

@app.route('/delete_user/<int:user_id>', methods=['POST'])
@login_required
# @login_required
@role_required('Super Admin') 
# @login_required
# @role_required('Super Admin')
def delete_user(user_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("DELETE FROM aawsa_user WHERE user_id = :user_id", {'user_id': user_id})
        conn.commit()
        flash('User deleted successfully!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting user: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('user_management'))

@app.route('/show_users')
@login_required
# @login_required
@role_required('Super Admin') 
def show_users():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT u.payroll_number, u.payroll_number, b.branch_name, r.role_name
        FROM aawsa_user u
        JOIN branch b ON u.user_id = b.user_id
        JOIN role r ON u.user_id = r.user_id
    """)
    users = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template('show_users.html', users=users)

@app.route('/upload_bill', methods=['GET', 'POST'])
@login_required
# @login_required
@role_required('Super Admin') 
def upload_bill():
    if request.method == 'POST':
        file = request.files['file']
        if not file or not file.filename.endswith('.csv'):
            flash("Please upload a valid CSV file.", "danger")
            return redirect('/upload_bill')

        connection = get_db_connection()
        cursor = connection.cursor()

        duplicate_bill_keys = set()  # Using set for faster lookups
        inserted_count = 0
        csv_bill_keys = set()
        rows = []

        # Define the expected column order
        COLUMNS = [
            'BILLKEY', 'CUSTOMERKEY', 'CUSTOMERNAME', 'CUSTOMERTIN', 'CUSTOMERBRANCH',
            'REASON', 'CURRREAD', 'PREVREAD', 'CONS', 'TOTALBILLAMOUNT',
            'THISMONTHBILLAMT', 'OUTSTANDINGAMT', 'PENALTYAMT', 'DRACCTNO', 'CRACCTNO'
        ]

        try:
            # Process the file in memory for faster access
            file_content = file.stream.read().decode('utf-8').splitlines()
            reader = csv.reader(file_content)
            
            # Skip header if exists
            next(reader, None)

            # Single pass processing with immediate duplicate checking
            for row in reader:
                if not row:
                    continue

                try:
                    row_dict = dict(zip(COLUMNS, row))
                    bill_key = row_dict['BILLKEY']

                    # Check for duplicates in CSV first
                    if bill_key in csv_bill_keys:
                        duplicate_bill_keys.add(bill_key)
                        continue

                    csv_bill_keys.add(bill_key)
                    rows.append(row_dict)
                except (IndexError, KeyError) as e:
                    app.logger.warning(f"Skipping malformed row: {row}. Error: {str(e)}")
                    continue

            # Bulk check for existing keys in database
            if rows:
                # Use batches to avoid too many parameters (Oracle has a limit)
                batch_size = 1000
                existing_keys = set()
                
                for i in range(0, len(csv_bill_keys), batch_size):
                    batch = list(csv_bill_keys)[i:i + batch_size]
                    placeholders = ', '.join([':key'+str(i) for i in range(len(batch))])
                    params = {'key'+str(i): key for i, key in enumerate(batch)}
                    
                    cursor.execute(f"SELECT BILLKEY FROM BILL WHERE BILLKEY IN ({placeholders})", params)
                    existing_keys.update(row[0] for row in cursor.fetchall())

                # Filter and prepare for insert
                unique_rows = []
                for row in rows:
                    if row['BILLKEY'] in existing_keys:
                        duplicate_bill_keys.add(row['BILLKEY'])
                    else:
                        unique_rows.append(row)

                # Bulk insert in batches
                if unique_rows:
                    insert_sql = """
                        INSERT INTO BILL (
                            BILLKEY, CUSTOMERKEY, CUSTOMERNAME, CUSTOMERTIN, CUSTOMERBRANCH, 
                            REASON, CURRREAD, PREVREAD, CONS, TOTALBILLAMOUNT, 
                            THISMONTHBILLAMT, OUTSTANDINGAMT, PENALTYAMT, DRACCTNO, CRACCTNO
                        ) VALUES (
                            :BILLKEY, :CUSTOMERKEY, :CUSTOMERNAME, :CUSTOMERTIN, :CUSTOMERBRANCH, 
                            :REASON, :CURRREAD, :PREVREAD, :CONS, :TOTALBILLAMOUNT, 
                            :THISMONTHBILLAMT, :OUTSTANDINGAMT, :PENALTYAMT, :DRACCTNO, :CRACCTNO
                        )
                    """
                    
                    # Insert in batches to avoid memory issues
                    for i in range(0, len(unique_rows), batch_size):
                        batch = unique_rows[i:i + batch_size]
                        cursor.executemany(insert_sql, batch)
                    
                    inserted_count = len(unique_rows)
                    connection.commit()
                    flash(f"Successfully uploaded {inserted_count} bills!", "success")
                else:
                    flash("No unique bill data to upload.", "warning")

            # Prepare duplicate message
            if duplicate_bill_keys:
                dup_count = len(duplicate_bill_keys)
                sample_duplicates = ', '.join(list(duplicate_bill_keys)[:5])
                if dup_count > 5:
                    sample_duplicates += f" and {dup_count-5} more"
                flash(f"Skipped {dup_count} duplicate bill keys (sample: {sample_duplicates})", "info")

        except Exception as e:
            connection.rollback()
            flash(f"Error uploading bill data: {str(e)}", "danger")
            app.logger.error(f"Error in upload_bill: {str(e)}", exc_info=True)
        
        finally:
            cursor.close()
            connection.close()

        return redirect('/upload_bill')

    return render_template('upload_bill.html')



    
@app.route('/upload_vat', methods=['GET', 'POST'])
@login_required
# @login_required
@role_required('Super Admin') 
def upload_vat():
    if request.method == 'POST':
        file = request.files['file']
        if not file or not file.filename.endswith('.csv'):
            flash("Please upload a valid CSV file.", "danger")
            return redirect('/upload_vat')

        connection = get_db_connection()
        cursor = connection.cursor()

        duplicate_bill_keys = set()  # Using set for faster lookups
        inserted_count = 0
        csv_bill_keys = set()
        rows = []

        # Define the expected column order
        COLUMNS = [
            'BILL_KEY','VAT_AMOUNT'
        ]

        try:
            # Process the file in memory for faster access
            file_content = file.stream.read().decode('utf-8').splitlines()
            reader = csv.reader(file_content)
            
            # Skip header if exists
            next(reader, None)

            # Single pass processing with immediate duplicate checking
            for row in reader:
                if not row:
                    continue

                try:
                    row_dict = dict(zip(COLUMNS, row))
                    bill_key = row_dict['BILL_KEY']

                    # Check for duplicates in CSV first
                    if bill_key in csv_bill_keys:
                        duplicate_bill_keys.add(bill_key)
                        continue

                    csv_bill_keys.add(bill_key)
                    rows.append(row_dict)
                except (IndexError, KeyError) as e:
                    app.logger.warning(f"Skipping malformed row: {row}. Error: {str(e)}")
                    continue

            # Bulk check for existing keys in database
            if rows:
                # Use batches to avoid too many parameters (Oracle has a limit)
                batch_size = 1000
                existing_keys = set()
                
                for i in range(0, len(csv_bill_keys), batch_size):
                    batch = list(csv_bill_keys)[i:i + batch_size]
                    placeholders = ', '.join([':key'+str(i) for i in range(len(batch))])
                    params = {'key'+str(i): key for i, key in enumerate(batch)}
                    
                    cursor.execute(f"SELECT BILL_KEY FROM bill_vat WHERE BILL_KEY IN ({placeholders})", params)
                    existing_keys.update(row[0] for row in cursor.fetchall())

                # Filter and prepare for insert
                unique_rows = []
                for row in rows:
                    if row['BILL_KEY'] in existing_keys:
                        duplicate_bill_keys.add(row['BILL_KEY'])
                    else:
                        unique_rows.append(row)

                # Bulk insert in batches
                if unique_rows:
                    insert_sql = """
                        INSERT INTO bill_vat (
                            BILL_KEY, VAT_AMOUNT
                        ) VALUES (
                            :BILL_KEY, :VAT_AMOUNT
                        )
                    """
                    
                    # Insert in batches to avoid memory issues
                    for i in range(0, len(unique_rows), batch_size):
                        batch = unique_rows[i:i + batch_size]
                        cursor.executemany(insert_sql, batch)
                    
                    inserted_count = len(unique_rows)
                    connection.commit()
                    flash(f"Successfully uploaded {inserted_count} bills!", "success")
                else:
                    flash("No unique bill data to upload.", "warning")

            # Prepare duplicate message
            if duplicate_bill_keys:
                dup_count = len(duplicate_bill_keys)
                sample_duplicates = ', '.join(list(duplicate_bill_keys)[:5])
                if dup_count > 5:
                    sample_duplicates += f" and {dup_count-5} more"
                flash(f"Skipped {dup_count} duplicate bill keys (sample: {sample_duplicates})", "info")

        except Exception as e:
            connection.rollback()
            flash(f"Error uploading bill data: {str(e)}", "danger")
            app.logger.error(f"Error in upload_bill: {str(e)}", exc_info=True)
        
        finally:
            cursor.close()
            connection.close()

        return redirect('/upload_vat')

    return render_template('upload_vat.html')


# def initialize_reconciliation_tables():
#     conn = get_db_connection()
#     cursor = conn.cursor()
    
#     try:
#         # Check if tables exist first
#         cursor.execute("""
#         BEGIN
#             EXECUTE IMMEDIATE 'CREATE TABLE reconciliation_uploads (
#                 upload_id VARCHAR2(50) PRIMARY KEY,
#                 filename VARCHAR2(255) NOT NULL,
#                 total_records NUMBER NOT NULL,
#                 matched_records NUMBER DEFAULT 0,
#                 unmatched_records NUMBER DEFAULT 0,
#                 duplicate_records NUMBER DEFAULT 0,
#                 uploaded_by VARCHAR2(50) NOT NULL,
#                 uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#                 status VARCHAR2(20) DEFAULT ''PROCESSING''
#             )';
#         EXCEPTION
#             WHEN OTHERS THEN
#                 IF SQLCODE != -955 THEN -- table already exists
#                     RAISE;
#                 END IF;
#         END;
#         """)
        
#         cursor.execute("""
#         BEGIN
#             EXECUTE IMMEDIATE 'CREATE TABLE reconciliation_results (
#                 id NUMBER GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
#                 upload_id VARCHAR2(50) NOT NULL,
#                 billkey VARCHAR2(50) ,
#                 banktxnref VARCHAR2(100) NOT NULL,
#                 status VARCHAR2(20) NOT NULL,
#                 matched_at TIMESTAMP,
#                 uploaded_by VARCHAR2(50) NOT NULL,
#                 uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
#             )';
#         EXCEPTION
#             WHEN OTHERS THEN
#                 IF SQLCODE != -955 THEN -- table already exists
#                     RAISE;
#                 END IF;
#         END;
#         """)
        
#         cursor.execute("""
#         BEGIN
#             EXECUTE IMMEDIATE 'ALTER TABLE reconciliation_results ADD CONSTRAINT uk_reconciliation_unique UNIQUE ( banktxnref)';
#         EXCEPTION
#             WHEN OTHERS THEN
#                 IF SQLCODE != -2261 THEN -- constraint already exists
#                     RAISE;
#                 END IF;
#         END;
#         """)
        
#         conn.commit()
#     except Exception as e:
#         conn.rollback()
#         print(f"Error initializing reconciliation tables: {str(e)}")
#     finally:
#         cursor.close()
#         conn.close()

# # Call this at app startup
# initialize_reconciliation_tables()


# Add these new tables to your initialization function
def initialize_reconciliation_tables():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if tables exist first
        cursor.execute("""
        BEGIN
            EXECUTE IMMEDIATE 'CREATE TABLE reconciliation_uploads (
                upload_id VARCHAR2(50) PRIMARY KEY,
                filename VARCHAR2(255) NOT NULL,
                total_records NUMBER NOT NULL,
                matched_records NUMBER DEFAULT 0,
                unmatched_records NUMBER DEFAULT 0,
                duplicate_records NUMBER DEFAULT 0,
                uploaded_by VARCHAR2(50) NOT NULL,
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status VARCHAR2(20) DEFAULT ''PROCESSING'',
                processed_at TIMESTAMP,
                processing_time NUMBER
            )';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN -- table already exists
                    RAISE;
                END IF;
        END;
        """)
        
        cursor.execute("""
        BEGIN
            EXECUTE IMMEDIATE 'CREATE TABLE reconciliation_results (
                id NUMBER GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
                upload_id VARCHAR2(50) NOT NULL,
                billkey VARCHAR2(50),
                banktxnref VARCHAR2(100) NOT NULL,
                status VARCHAR2(20) NOT NULL,
                matched_at TIMESTAMP,
                uploaded_by VARCHAR2(50) NOT NULL,
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                customer_service_verified NUMBER DEFAULT 0,
                verified_by VARCHAR2(50),
                verified_at TIMESTAMP,
                notes VARCHAR2(500)
            )';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN -- table already exists
                    RAISE;
                END IF;
        END;
        """)
        
        cursor.execute("""
        BEGIN
            EXECUTE IMMEDIATE 'CREATE TABLE reconciliation_downloads (
                download_id NUMBER GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
                upload_id VARCHAR2(50) NOT NULL,
                downloaded_by VARCHAR2(50) NOT NULL,
                downloaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                download_type VARCHAR2(20) NOT NULL, -- 'MATCHED' or 'VERIFIED'
                FOREIGN KEY (upload_id) REFERENCES reconciliation_uploads(upload_id)
            )';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN -- table already exists
                    RAISE;
                END IF;
        END;
        """)
        
        cursor.execute("""
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE reconciliation_results ADD CONSTRAINT uk_reconciliation_unique UNIQUE (banktxnref)';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -2261 THEN -- constraint already exists
                    RAISE;
                END IF;
        END;
        """)
        
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"Error initializing reconciliation tables: {str(e)}")
    finally:
        cursor.close()
        conn.close()

initialize_reconciliation_tables()

# Enhanced reconciliation processing
def process_reconciliation_file(file, uploaded_by):
    conn = get_db_connection()
    cursor = conn.cursor()
    start_time = datetime.now()

    try:
        # Generate unique upload ID
        upload_id = str(uuid.uuid4())
        filename = secure_filename(file.stream.filename)

        # Insert initial tracking record
        cursor.execute("""
            INSERT INTO reconciliation_uploads (
                upload_id, filename, total_records, uploaded_by, status
            ) VALUES (
                :upload_id, :filename, 0, :uploaded_by, 'PROCESSING'
            )
        """, {
            'upload_id': upload_id,
            'filename': filename,
            'uploaded_by': uploaded_by
        })

        # Process file in chunks for memory efficiency
        chunk_size = 1000
        total = matched = unmatched = duplicates = 0
        seen_keys = set()
        batch = []
        warnings = []

        # Use pandas for efficient CSV processing
        for chunk in pd.read_csv(file.stream, chunksize=chunk_size, dtype=str):
            # Process each row in the chunk
            for _, row in chunk.iterrows():
                total += 1
                banktxnref = str(row.get('BANKTXNREF', '')).strip()

                if not banktxnref:
                    continue

                unique_key = banktxnref
                if unique_key in seen_keys:
                    duplicates += 1
                    status = 'DUPLICATE'
                    warnings.append(f"Duplicate in file: banktxnref={banktxnref}")
                else:
                    seen_keys.add(unique_key)
                    
                    # Check if already in DB (optimized query)
                    cursor.execute("""
                        SELECT status FROM reconciliation_results 
                        WHERE banktxnref = :banktxnref
                        AND ROWNUM = 1
                    """, {'banktxnref': banktxnref})
                    existing = cursor.fetchone()
                    
                    if existing:
                        duplicates += 1
                        status = 'DUPLICATE'
                        warnings.append(f"Duplicate in database: banktxnref={banktxnref}")
                    else:
                        # Optimized check in payment table
                        cursor.execute("""
                            SELECT 1 FROM payment 
                            WHERE BANKTXNREF = :banktxnref
                            AND ROWNUM = 1
                        """, {'banktxnref': banktxnref})
                        if cursor.fetchone():
                            matched += 1
                            status = 'MATCHED'
                        else:
                            unmatched += 1
                            status = 'UNMATCHED'

                if status != 'DUPLICATE':
                    batch.append({
                        'upload_id': upload_id,
                        'banktxnref': banktxnref,
                        'status': status,
                        'uploaded_by': uploaded_by,
                        'matched_at': datetime.now() if status == 'MATCHED' else None
                    })

                # Insert in batches
                if len(batch) >= 100:
                    cursor.executemany("""
                        INSERT INTO reconciliation_results (
                            upload_id, banktxnref, status, 
                            matched_at, uploaded_by
                        ) VALUES (
                            :upload_id, :banktxnref, :status,
                            :matched_at, :uploaded_by
                        )
                    """, batch)
                    batch = []

        # Insert remaining records
        if batch:
            cursor.executemany("""
                INSERT INTO reconciliation_results (
                    upload_id, banktxnref, status, 
                    matched_at, uploaded_by
                ) VALUES (
                    :upload_id, :banktxnref, :status,
                    :matched_at, :uploaded_by
                )
            """, batch)

        # Calculate processing time
        processing_time = (datetime.now() - start_time).total_seconds()

        # Update summary with processing time
        cursor.execute("""
            UPDATE reconciliation_uploads
            SET 
                total_records = :total,
                matched_records = :matched,
                unmatched_records = :unmatched,
                duplicate_records = :duplicates,
                status = 'COMPLETED',
                processed_at = CURRENT_TIMESTAMP,
                processing_time = :processing_time
            WHERE upload_id = :upload_id
        """, {
            'total': total,
            'matched': matched,
            'unmatched': unmatched,
            'duplicates': duplicates,
            'processing_time': processing_time,
            'upload_id': upload_id
        })

        conn.commit()

        return {
            'upload_id': upload_id,
            'total': total,
            'matched': matched,
            'unmatched': unmatched,
            'duplicates': duplicates,
            'warnings': warnings,
            'processing_time': processing_time
        }

    except Exception as e:
        conn.rollback()
        # Update status to failed if error occurs
        cursor.execute("""
            UPDATE reconciliation_uploads
            SET status = 'FAILED',
                processed_at = CURRENT_TIMESTAMP,
                processing_time = :processing_time
            WHERE upload_id = :upload_id
        """, {
            'processing_time': (datetime.now() - start_time).total_seconds(),
            'upload_id': upload_id
        })
        conn.commit()
        raise e
    finally:
        cursor.close()
        conn.close()

# New route for customer service to view unmatched transactions
@app.route('/unmatched_transactions', methods=['GET'])
@role_required('Customer Service', 'Super Admin')
def unmatched_transactions():
    page = request.args.get('page', 1, type=int)
    per_page = 50
    search_query = request.args.get('search', '')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Base query
        query = """
            SELECT r.id, r.banktxnref, r.uploaded_at, u.filename, r.uploaded_by
            FROM reconciliation_results r
            JOIN reconciliation_uploads u ON r.upload_id = u.upload_id
            WHERE r.status = 'UNMATCHED'
            AND r.customer_service_verified = 0
        """
        
        params = {}
        
        # Add search filter if provided
        if search_query:
            query += " AND r.banktxnref LIKE :search"
            params['search'] = f'%{search_query}%'
        
        # Count total for pagination
        count_query = f"SELECT COUNT(*) FROM ({query})"
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]
        
        # Add pagination
        query += """
            ORDER BY r.uploaded_at DESC
            OFFSET :offset ROWS FETCH NEXT :per_page ROWS ONLY
        """
        params['offset'] = (page - 1) * per_page
        params['per_page'] = per_page
        
        cursor.execute(query, params)
        transactions = cursor.fetchall()
        
        pages = ceil(total / per_page)
        
        return render_template(
            'unmatched_transactions.html',
            transactions=transactions,
            page=page,
            pages=pages,
            total=total,
            search_query=search_query
        )
        
    except Exception as e:
        flash(f'Error retrieving unmatched transactions: {str(e)}', 'danger')
        return redirect(url_for('navigation'))
    finally:
        cursor.close()
        conn.close()

# Route to mark transaction as verified/paid by customer service
@app.route('/verify_transaction/<int:transaction_id>', methods=['GET', 'POST'])
@role_required('Customer Service', 'Super Admin')
def verify_transaction(transaction_id):
    if request.method == 'POST':
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Get transaction details
            cursor.execute("""
                SELECT banktxnref FROM reconciliation_results
                WHERE id = :id AND status = 'UNMATCHED'
            """, {'id': transaction_id})
            
            transaction = cursor.fetchone()
            
            if not transaction:
                flash('Transaction not found or already processed', 'danger')
                return redirect(url_for('unmatched_transactions'))
            
            banktxnref = transaction[0]
            notes = request.form.get('notes', '')
            
            # Verify the transaction exists in the bank data (additional check)
            cursor.execute("""
                SELECT 1 FROM reconciliation_results
                WHERE banktxnref = :banktxnref
                AND status = 'UNMATCHED'
                AND ROWNUM = 1
            """, {'banktxnref': banktxnref})
            
            if not cursor.fetchone():
                flash('Transaction verification failed', 'danger')
                return redirect(url_for('unmatched_transactions'))
            
            # Update the reconciliation record
            cursor.execute("""
                UPDATE reconciliation_results
                SET 
                    customer_service_verified = 1,
                    verified_by = :verified_by,
                    verified_at = CURRENT_TIMESTAMP,
                    notes = :notes,
                    status = 'VERIFIED'
                WHERE id = :id
            """, {
                'verified_by': session['payroll_number'],
                'notes': notes,
                'id': transaction_id
            })
            
            # Add to payment table (you'll need to adjust this based on your payment table structure)
            try:
                cursor.execute("""
                    INSERT INTO payment (
                        PAYMENTKEY, BILLKEY, CUSTOMERKEY, AMOUNT, BANKTXNREF,
                        PAYMENTCHANNEL, PAYMENTDATE, REQUESTID, CHANNEL,
                        PAYMENTCOMPLETEDAT, VERIFIED_BY, VERIFIED_AT
                    ) VALUES (
                        :paymentkey, :billkey, :customerkey, :amount, :banktxnref,
                        'MANUAL_VERIFICATION', CURRENT_DATE, :requestid, 'MANUAL',
                        CURRENT_TIMESTAMP, :verified_by, CURRENT_TIMESTAMP
                    )
                """, {
                    'paymentkey': str(uuid.uuid4()),
                    'billkey': None,  # You may need to get this from another source
                    'customerkey': None,  # You may need to get this from another source
                    'amount': request.form.get('amount'),  # From form or calculate
                    'banktxnref': banktxnref,
                    'requestid': f"MANUAL_{banktxnref}",
                    'verified_by': session['payroll_number']
                })
            except Exception as e:
                flash(f'Error adding to payment table: {str(e)}', 'warning')
                # Continue even if payment table update fails
            
            conn.commit()
            flash('Transaction verified and marked as paid successfully', 'success')
            
        except Exception as e:
            conn.rollback()
            flash(f'Error verifying transaction: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
        
        return redirect(url_for('unmatched_transactions'))
    
    # GET request - show verification form
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT r.id, r.banktxnref, r.uploaded_at, u.filename
            FROM reconciliation_results r
            JOIN reconciliation_uploads u ON r.upload_id = u.upload_id
            WHERE r.id = :id AND r.status = 'UNMATCHED'
        """, {'id': transaction_id})
        
        transaction = cursor.fetchone()
        
        if not transaction:
            flash('Transaction not found or already processed', 'danger')
            return redirect(url_for('unmatched_transactions'))
        
        return render_template('verify_transaction.html', transaction=transaction)
        
    except Exception as e:
        flash(f'Error retrieving transaction: {str(e)}', 'danger')
        return redirect(url_for('unmatched_transactions'))
    finally:
        cursor.close()
        conn.close()

# Route to download matched/verified transactions with download tracking
@app.route('/download_reconciliation/<upload_id>/<download_type>')
@role_required('Finance', 'Super Admin', 'Customer Service')
def download_reconciliation(upload_id, download_type):
    if download_type not in ['matched', 'verified']:
        flash('Invalid download type', 'danger')
        return redirect(url_for('reconciliation_results', upload_id=upload_id))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if already downloaded
        cursor.execute("""
            SELECT 1 FROM reconciliation_downloads
            WHERE upload_id = :upload_id
            AND download_type = :download_type
            AND ROWNUM = 1
        """, {
            'upload_id': upload_id,
            'download_type': download_type.upper()
        })
        
        if cursor.fetchone():
            flash('This data has already been downloaded previously', 'warning')
        
        # Get the data
        if download_type == 'matched':
            status_condition = "status = 'MATCHED'"
        else:
            status_condition = "status = 'VERIFIED'"
        
        cursor.execute(f"""
            SELECT 
                banktxnref, 
                TO_CHAR(matched_at, 'YYYY-MM-DD HH24:MI:SS') as matched_at,
                status,
                uploaded_by,
                TO_CHAR(uploaded_at, 'YYYY-MM-DD HH24:MI:SS') as uploaded_at,
                CASE WHEN customer_service_verified = 1 THEN 'YES' ELSE 'NO' END as verified,
                verified_by,
                TO_CHAR(verified_at, 'YYYY-MM-DD HH24:MI:SS') as verified_at,
                notes
            FROM reconciliation_results
            WHERE upload_id = :upload_id
            AND {status_condition}
            ORDER BY matched_at DESC
        """, {'upload_id': upload_id})
        
        results = cursor.fetchall()
        
        if not results:
            flash('No data found for download', 'warning')
            return redirect(url_for('reconciliation_results', upload_id=upload_id))
        
        # Create CSV in memory
        si = io.StringIO()
        cw = csv.writer(si)
        
        # Write header
        headers = [
            'BANKTXNREF', 'MATCHED_AT', 'STATUS', 'UPLOADED_BY', 'UPLOADED_AT',
            'VERIFIED', 'VERIFIED_BY', 'VERIFIED_AT', 'NOTES'
        ]
        cw.writerow(headers)
        
        # Write data
        for row in results:
            cw.writerow(row)
        
        # Record the download
        cursor.execute("""
            INSERT INTO reconciliation_downloads (
                upload_id, downloaded_by, download_type
            ) VALUES (
                :upload_id, :downloaded_by, :download_type
            )
        """, {
            'upload_id': upload_id,
            'downloaded_by': session['payroll_number'],
            'download_type': download_type.upper()
        })
        
        conn.commit()
        
        # Prepare response
        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename={download_type}_transactions_{upload_id}.csv"
        output.headers["Content-type"] = "text/csv"
        
        return output
        
    except Exception as e:
        conn.rollback()
        flash(f'Error downloading data: {str(e)}', 'danger')
        return redirect(url_for('reconciliation_results', upload_id=upload_id))
    finally:
        cursor.close()
        conn.close()

# Route to view download history
@app.route('/download_history')
@role_required('Finance', 'Super Admin')
def download_history():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Count total downloads
        cursor.execute("SELECT COUNT(*) FROM reconciliation_downloads")
        total = cursor.fetchone()[0]
        
        # Get downloads with pagination
        cursor.execute("""
            SELECT d.download_id, d.upload_id, u.filename, 
                   d.downloaded_by, d.downloaded_at, d.download_type,
                   u.total_records, u.matched_records, u.unmatched_records
            FROM reconciliation_downloads d
            JOIN reconciliation_uploads u ON d.upload_id = u.upload_id
            ORDER BY d.downloaded_at DESC
            OFFSET :offset ROWS FETCH NEXT :per_page ROWS ONLY
        """, {
            'offset': (page - 1) * per_page,
            'per_page': per_page
        })
        
        downloads = cursor.fetchall()
        pages = ceil(total / per_page)
        
        return render_template(
            'download_history.html',
            downloads=downloads,
            page=page,
            pages=pages,
            total=total
        )
        
    except Exception as e:
        flash(f'Error retrieving download history: {str(e)}', 'danger')
        return redirect(url_for('navigation'))
    finally:
        cursor.close()
        conn.close()


@app.route('/reconciliation', methods=['GET', 'POST'])
@role_required('Finance', 'Super Admin')  # Only authorized roles can access
def reconciliation():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(request.url)
            
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)
            
        if not file.filename.lower().endswith('.csv'):
            flash('Only CSV files are allowed', 'danger')
            return redirect(request.url)
            
        try:
            # Process the file
            result = process_reconciliation_file(file, session['payroll_number'])
            session['reconciliation_warnings'] = result.get('warnings', [])

            
            # Prepare summary message
            msg = (
                f"Reconciliation completed. "
                f"Total: {result['total']}, "
                f"Matched: {result['matched']}, "
                f"Unmatched: {result['unmatched']}, "
                f"Duplicates: {result['duplicates']}"
            )
            flash(msg, 'success')
            
            return redirect(url_for('reconciliation_results', upload_id=result['upload_id']))
            
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'danger')
            app.logger.error(f"Reconciliation error: {str(e)}", exc_info=True)
            return redirect(request.url)
    
    # GET request - show upload form
    return render_template('reconciliation.html')

# def process_reconciliation_file(file, uploaded_by):
#     conn = get_db_connection()
#     cursor = conn.cursor()

#     try:
#         # Generate unique upload ID
#         upload_id = str(uuid.uuid4())
#         filename = secure_filename(file.filename)

#         stream = io.TextIOWrapper(file.stream, encoding='utf-8')
#         reader = csv.DictReader(stream)

#         total = matched = unmatched = duplicates = 0
#         warnings = []
#         seen_keys = set()  # To detect duplicates in the same upload
#         batch = []

#         # Insert initial tracking record
#         cursor.execute("""
#             INSERT INTO reconciliation_uploads (
#                 upload_id, filename, total_records, uploaded_by
#             ) VALUES (
#                 :upload_id, :filename, :total, :uploaded_by
#             )
#         """, {
#             'upload_id': upload_id,
#             'filename': filename,
#             'total': 0,
#             'uploaded_by': uploaded_by
#         })

#         for row in reader:
#             total += 1
#             # billkey = row.get('billkey', '').strip()
#             banktxnref = row.get('BANKTXNREF', '').strip()

#             if  not banktxnref:
#                 continue  # Skip invalid rows

#             unique_key = f"{banktxnref}"
#             if unique_key in seen_keys:
#                 duplicates += 1
#                 status = 'DUPLICATE'
#                 warnings.append(f"Duplicate in file:  banktxnref={banktxnref}")
#             else:
#                 seen_keys.add(unique_key)
#                 # Check if already in DB
#                 cursor.execute("""
#                     SELECT 1 FROM reconciliation_results 
#                     WHERE banktxnref = :banktxnref
#                 """, { 'banktxnref': banktxnref})
#                 if cursor.fetchone():
#                     duplicates += 1
#                     status = 'DUPLICATE'
#                     warnings.append(f"Duplicate in database:  banktxnref={banktxnref}")
#                 else:
#                     # Check if matched in payment table
#                     cursor.execute("""
#                         SELECT 1 FROM payment 
#                         WHERE  BANKTXNREF = :banktxnref
#                     """, { 'banktxnref': banktxnref})
#                     if cursor.fetchone():
#                         matched += 1
#                         status = 'MATCHED'
#                     else:
#                         unmatched += 1
#                         status = 'UNMATCHED'

#             # batch.append({
#             #     'upload_id': upload_id,
#             #     'billkey': billkey,
#             #     'banktxnref': banktxnref,
#             #     'status': status,
#             #     'uploaded_by': uploaded_by,
#             #     'matched_at': datetime.now() if status == 'MATCHED' else None
#             # })
#             if status != 'DUPLICATE':
#                 batch.append({
#         'upload_id': upload_id,
#         'banktxnref': banktxnref,
#         'status': status,
#         'uploaded_by': uploaded_by,
#         'matched_at': datetime.now() if status == 'MATCHED' else None
#     })

#             if len(batch) >= 100:
#                 cursor.executemany("""
#                     INSERT INTO reconciliation_results (
#                         upload_id,  banktxnref, status, 
#                         matched_at, uploaded_by
#                     ) VALUES (
#                         :upload_id,  :banktxnref, :status,
#                         :matched_at, :uploaded_by
#                     )
#                 """, batch)
#                 batch = []

#         if batch:
#             cursor.executemany("""
#                 INSERT INTO reconciliation_results (
#                     upload_id,  banktxnref, status, 
#                     matched_at, uploaded_by
#                 ) VALUES (
#                     :upload_id,  :banktxnref, :status,
#                     :matched_at, :uploaded_by
#                 )
#             """, batch)

#         # Update summary
#         cursor.execute("""
#             UPDATE reconciliation_uploads
#             SET 
#                 total_records = :total,
#                 matched_records = :matched,
#                 unmatched_records = :unmatched,
#                 duplicate_records = :duplicates,
#                 status = 'COMPLETED'
#             WHERE upload_id = :upload_id
#         """, {
#             'total': total,
#             'matched': matched,
#             'unmatched': unmatched,
#             'duplicates': duplicates,
#             'upload_id': upload_id
#         })

#         conn.commit()

#         return {
#             'upload_id': upload_id,
#             'total': total,
#             'matched': matched,
#             'unmatched': unmatched,
#             'duplicates': duplicates,
#             'warnings': warnings  # Show this in the frontend
#         }

#     except Exception as e:
#         conn.rollback()
#         raise e
#     finally:
#         cursor.close()
#         conn.close()



@app.route('/reconciliation/results/<upload_id>')
@role_required('Finance', 'Super Admin')
def reconciliation_results(upload_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    warnings = session.pop('reconciliation_warnings', [])  # Remove after reading
    
    try:
        # Get upload summary
        cursor.execute("""
            SELECT * FROM reconciliation_uploads
            WHERE upload_id = :upload_id
        """, {'upload_id': upload_id})
        
        upload = cursor.fetchone()
        
        if not upload:
            flash('Upload not found', 'danger')
            return redirect(url_for('reconciliation'))
        
        # Get results with pagination
        page = request.args.get('page', 1, type=int)
        per_page = 50
        offset = (page - 1) * per_page
        
        cursor.execute("""
            SELECT  banktxnref, status, 
                   matched_at, uploaded_by
            FROM reconciliation_results
            WHERE upload_id = :upload_id
            ORDER BY status
            OFFSET :offset ROWS FETCH NEXT :per_page ROWS ONLY
        """, {
            'upload_id': upload_id,
            'offset': offset,
            'per_page': per_page
        })
        
        results = cursor.fetchall()
        
        # Get total count for pagination
        cursor.execute("""
            SELECT COUNT(*) FROM reconciliation_results
            WHERE upload_id = :upload_id
        """, {'upload_id': upload_id})
        
        total = cursor.fetchone()[0]
        pages = ceil(total / per_page)
        
        # Convert results to a list of dictionaries for easier template access
        formatted_results = []
        for row in results:
            formatted_results.append({
                'banktxnref': row[1],
                'status': row[2],
                'matched_at': row[3],
                'uploaded_by': row[4]
            })
        
        return render_template(
            'reconciliation_results.html',
            upload={
                'id': upload[0],
                'filename': upload[1],
                'total': upload[2],
                'matched': upload[3],
                'unmatched': upload[4],
                'duplicates': upload[5],
                'uploaded_by': upload[6],
                'uploaded_at': upload[7],
                'status': upload[8]
            },
            results=formatted_results,
            page=page,
            pages=pages,
            total=total, warnings=warnings, upload_id=upload_id
        )
        
    except Exception as e:
        flash(f'Error retrieving results: {str(e)}', 'danger')
        return redirect(url_for('reconciliation'))
    finally:
        cursor.close()
        conn.close()

@app.route('/reconciliation/export/<upload_id>')
@role_required('Finance', 'Super Admin')
def export_reconciliation(upload_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Get upload info
        cursor.execute("""
            SELECT filename FROM reconciliation_uploads
            WHERE upload_id = :upload_id
        """, {'upload_id': upload_id})
        
        upload = cursor.fetchone()
        
        if not upload:
            flash('Upload not found', 'danger')
            return redirect(url_for('reconciliation'))
        
        # Get all results
        cursor.execute("""
            SELECT  banktxnref, status, 
                   TO_CHAR(matched_at, 'YYYY-MM-DD HH24:MI:SS') as matched_at
            FROM reconciliation_results
            WHERE upload_id = :upload_id
            ORDER BY status
        """, {'upload_id': upload_id})
        
        results = cursor.fetchall()
        
        # Create CSV in memory
        si = io.StringIO()
        cw = csv.writer(si)
        
        # Write header
        cw.writerow([ 'BANKTXNREF', 'STATUS', 'MATCHED_AT'])
        
        # Write data
        for row in results:
            cw.writerow(row)
        
        # Prepare response
        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename=reconciliation_{upload_id}.csv"
        output.headers["Content-type"] = "text/csv"
        
        return output
        
    except Exception as e:
        flash(f'Error exporting results: {str(e)}', 'danger')
        return redirect(url_for('reconciliation_results', upload_id=upload_id))
    finally:
        cursor.close()
        conn.close()




@app.route('/logout')
def logout():
    session.pop('payroll_number', None)  # Remove payroll_number from the session
    return redirect(url_for('login'))  # Redirect to the login page

if __name__ == '__main__':
    app.run(debug=True,port=5001,host="0.0.0.0")
