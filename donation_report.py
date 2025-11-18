"""
Donation Report Module
Comprehensive reporting for donations with charts and analytics
"""

from flask import Blueprint, render_template, request, jsonify, send_file, redirect, url_for, flash
from database import get_db_connection
from auth import login_required, role_required
from datetime import datetime, timedelta
import json
import io
from decimal import Decimal
import os

donation_report = Blueprint('donation_report', __name__, url_prefix='/reports/donation')

@donation_report.route('/')
@login_required
@role_required('Super Admin', 'Admin')
def donation_report_page():
    """Main donation report page with charts and statistics"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True, dictionary=True)
    
    try:
        # Get filters
        month = request.args.get('month', '')
        year = request.args.get('year', datetime.now().year)
        status_filter = request.args.get('status', '')
        
        # Default to current month if not specified
        if not month:
            month = datetime.now().month
        
        # Build date filter
        date_filter = f"{year}-{month:02d}-01"
        next_month = int(month) + 1 if int(month) < 12 else 1
        next_year = int(year) if int(month) < 12 else int(year) + 1
        date_filter_end = f"{next_year}-{next_month:02d}-01"
        
        # Total donations collected this month (include Paid, Completed, and Pending since transactions are auto-deducted)
        query_total = """
            SELECT COALESCE(SUM(amount), 0) as total_amount, COUNT(*) as total_count
            FROM donations
            WHERE payment_status IN ('Paid', 'Completed', 'Pending')
            AND DATE(created_at) >= %s AND DATE(created_at) < %s
        """
        cursor.execute(query_total, (date_filter, date_filter_end))
        total_stats = cursor.fetchone()
        
        # Donations by type (include Pending as Completed)
        query_by_type = """
            SELECT dt.name, dt.name_amharic, 
                   COALESCE(SUM(d.amount), 0) as total_amount,
                   COUNT(*) as count
            FROM donation_types dt
            LEFT JOIN donations d ON dt.id = d.donation_type_id 
                AND d.payment_status IN ('Paid', 'Completed', 'Pending')
                AND DATE(d.created_at) >= %s AND DATE(d.created_at) < %s
            WHERE dt.status = 'active' OR dt.is_active = 1
            GROUP BY dt.id, dt.name, dt.name_amharic
            ORDER BY total_amount DESC
        """
        cursor.execute(query_by_type, (date_filter, date_filter_end))
        by_type = cursor.fetchall()
        
        # Donations by status (treat Pending as Completed since transactions are auto-deducted)
        query_by_status = """
            SELECT 
                CASE WHEN payment_status = 'Pending' THEN 'Completed' ELSE payment_status END as payment_status,
                COALESCE(SUM(amount), 0) as total_amount,
                COUNT(*) as count
            FROM donations
            WHERE DATE(created_at) >= %s AND DATE(created_at) < %s
            GROUP BY CASE WHEN payment_status = 'Pending' THEN 'Completed' ELSE payment_status END
            ORDER BY FIELD(CASE WHEN payment_status = 'Pending' THEN 'Completed' ELSE payment_status END, 'Paid', 'Completed', 'Failed')
        """
        cursor.execute(query_by_status, (date_filter, date_filter_end))
        by_status = cursor.fetchall()
        
        # Merge any duplicate Completed entries (in case both Completed and Pending exist)
        status_dict = {}
        for item in by_status:
            status = item['payment_status']
            if status in status_dict:
                status_dict[status]['total_amount'] += item['total_amount']
                status_dict[status]['count'] += item['count']
            else:
                status_dict[status] = item
        by_status = list(status_dict.values())
        
        # Monthly trend (last 12 months) - include Pending as Completed
        query_trend = """
            SELECT DATE_FORMAT(created_at, '%%Y-%%m') as month,
                   COALESCE(SUM(amount), 0) as total_amount,
                   COUNT(*) as count
            FROM donations
            WHERE payment_status IN ('Paid', 'Completed', 'Pending')
            AND created_at >= DATE_SUB(NOW(), INTERVAL 12 MONTH)
            GROUP BY DATE_FORMAT(created_at, '%%Y-%%m')
            ORDER BY month ASC
        """
        cursor.execute(query_trend)
        monthly_trend = cursor.fetchall()
        
        # Recent donations (include christian_name)
        # Automatically convert Pending to Completed for display (since transactions are auto-deducted)
        query_recent = """
            SELECT d.*, dt.name as donation_type_name, dt.name_amharic as donation_type_name_amharic,
                   CASE WHEN d.payment_status = 'Pending' THEN 'Completed' ELSE d.payment_status END as display_status
            FROM donations d
            LEFT JOIN donation_types dt ON d.donation_type_id = dt.id
            WHERE DATE(d.created_at) >= %s AND DATE(d.created_at) < %s
        """
        params_recent = [date_filter, date_filter_end]
        
        if status_filter:
            # If filtering by Completed, also include Pending
            if status_filter == 'Completed':
                query_recent += " AND (d.payment_status = 'Completed' OR d.payment_status = 'Pending')"
            else:
                query_recent += " AND d.payment_status = %s"
                params_recent.append(status_filter)
        
        query_recent += " ORDER BY d.created_at DESC LIMIT 50"
        cursor.execute(query_recent, params_recent)
        recent_donations = cursor.fetchall()
        
        # Update payment_status in results to show Completed instead of Pending
        for donation in recent_donations:
            if donation.get('payment_status') == 'Pending':
                donation['payment_status'] = 'Completed'
        
        # Get all donation types for filter
        cursor.execute("SELECT id, name, name_amharic FROM donation_types WHERE status = 'active' OR is_active = 1 ORDER BY name")
        donation_types = cursor.fetchall()
        
    except Exception as e:
        total_stats = {'total_amount': 0, 'total_count': 0}
        by_type = []
        by_status = []
        monthly_trend = []
        recent_donations = []
        donation_types = []
        flash(f'Error loading report: {str(e)}', 'error')
    finally:
        cursor.close()
        conn.close()
    
    return render_template('admin/donation/report.html',
                         total_stats=total_stats,
                         by_type=by_type,
                         by_status=by_status,
                         monthly_trend=monthly_trend,
                         recent_donations=recent_donations,
                         donation_types=donation_types,
                         selected_month=month,
                         selected_year=year,
                         status_filter=status_filter)

@donation_report.route('/export/excel')
@login_required
@role_required('Super Admin', 'Admin')
def export_excel():
    """Export donation report to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        conn = get_db_connection()
        cursor = conn.cursor(buffered=True, dictionary=True)
        
        # Get filters
        month = request.args.get('month', datetime.now().month)
        year = request.args.get('year', datetime.now().year)
        status_filter = request.args.get('status', '')
        
        date_filter = f"{year}-{month:02d}-01"
        next_month = int(month) + 1 if int(month) < 12 else 1
        next_year = int(year) if int(month) < 12 else int(year) + 1
        date_filter_end = f"{next_year}-{next_month:02d}-01"
        
        # Build query
        query = """
            SELECT d.id, d.donor_name, d.donor_email, d.amount, d.currency,
                   d.payment_status, d.payment_method, d.tx_ref, d.created_at,
                   dt.name as donation_type_name
            FROM donations d
            LEFT JOIN donation_types dt ON d.donation_type_id = dt.id
            WHERE DATE(d.created_at) >= %s AND DATE(d.created_at) < %s
        """
        params = [date_filter, date_filter_end]
        
        if status_filter:
            query += " AND d.payment_status = %s"
            params.append(status_filter)
        
        query += " ORDER BY d.created_at DESC"
        
        cursor.execute(query, params)
        donations = cursor.fetchall()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Donations {year}-{month:02d}"
        
        # Header style
        header_fill = PatternFill(start_color="14860C", end_color="14860C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ['ID', 'Donor Name', 'Email', 'Type', 'Amount', 'Currency', 'Status', 'Method', 'Reference', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        for row_idx, donation in enumerate(donations, 2):
            ws.cell(row=row_idx, column=1, value=donation['id'])
            ws.cell(row=row_idx, column=2, value=donation['donor_name'] or 'Anonymous')
            ws.cell(row=row_idx, column=3, value=donation['donor_email'] or '')
            ws.cell(row=row_idx, column=4, value=donation['donation_type_name'] or '')
            ws.cell(row=row_idx, column=5, value=float(donation['amount']))
            ws.cell(row=row_idx, column=6, value=donation['currency'])
            ws.cell(row=row_idx, column=7, value=donation['payment_status'])
            ws.cell(row=row_idx, column=8, value=donation['payment_method'])
            ws.cell(row=row_idx, column=9, value=donation['tx_ref'] or '')
            ws.cell(row=row_idx, column=10, value=donation['created_at'].strftime('%Y-%m-%d %H:%M') if donation['created_at'] else '')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        cursor.close()
        conn.close()
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"donation_report_{year}_{month:02d}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
        
    except Exception as e:
        flash(f'Error exporting to Excel: {str(e)}', 'error')
        return redirect(url_for('donation_report.donation_report_page'))

@donation_report.route('/export/pdf')
@login_required
@role_required('Super Admin', 'Admin')
def export_pdf():
    """Export donation report to PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        conn = get_db_connection()
        cursor = conn.cursor(buffered=True, dictionary=True)
        
        # Get filters
        month = request.args.get('month', datetime.now().month)
        year = request.args.get('year', datetime.now().year)
        status_filter = request.args.get('status', '')
        
        date_filter = f"{year}-{month:02d}-01"
        next_month = int(month) + 1 if int(month) < 12 else 1
        next_year = int(year) if int(month) < 12 else int(year) + 1
        date_filter_end = f"{next_year}-{next_month:02d}-01"
        
        # Build query
        query = """
            SELECT d.id, d.donor_name, d.christian_name, d.donor_email, d.amount, d.currency,
                   d.payment_status, d.payment_method, d.tx_ref, d.created_at,
                   dt.name as donation_type_name, dt.name_amharic as donation_type_name_amharic
            FROM donations d
            LEFT JOIN donation_types dt ON d.donation_type_id = dt.id
            WHERE DATE(d.created_at) >= %s AND DATE(d.created_at) < %s
        """
        params = [date_filter, date_filter_end]
        
        if status_filter:
            query += " AND d.payment_status = %s"
            params.append(status_filter)
        
        query += " ORDER BY d.created_at DESC"
        
        cursor.execute(query, params)
        donations = cursor.fetchall()
        
        # Get summary statistics
        cursor.execute("""
            SELECT 
                COALESCE(SUM(CASE WHEN payment_status IN ('Paid', 'Completed') THEN amount ELSE 0 END), 0) as total_collected,
                COUNT(*) as total_count,
                COUNT(CASE WHEN payment_status = 'Paid' THEN 1 END) as paid_count,
                COUNT(CASE WHEN payment_status = 'Completed' THEN 1 END) as completed_count,
                COUNT(CASE WHEN payment_status = 'Pending' THEN 1 END) as pending_count,
                COUNT(CASE WHEN payment_status = 'Failed' THEN 1 END) as failed_count
            FROM donations
            WHERE DATE(created_at) >= %s AND DATE(created_at) < %s
        """, (date_filter, date_filter_end))
        stats = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#14860C'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Title
        elements.append(Paragraph("Donation Report / የለግስና ሪፖርት", title_style))
        elements.append(Paragraph(f"Period: {datetime(int(year), int(month), 1).strftime('%B %Y')}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary Statistics
        summary_data = [
            ['Total Collected', f"{stats['total_collected']:.2f} ETB"],
            ['Total Donations', str(stats['total_count'])],
            ['Paid', str(stats['paid_count'])],
            ['Completed', str(stats['completed_count'])],
            ['Pending', str(stats['pending_count'])],
            ['Failed', str(stats['failed_count'])]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#14860C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Donations Table
        if donations:
            table_data = [['ID', 'Donor', 'Type', 'Amount', 'Status', 'Date']]
            
            for donation in donations:
                donor_name = donation['donor_name'] or 'Anonymous'
                if donation['christian_name']:
                    donor_name += f" ({donation['christian_name']})"
                
                table_data.append([
                    str(donation['id']),
                    donor_name[:30],
                    (donation['donation_type_name_amharic'] or donation['donation_type_name'] or '')[:20],
                    f"{donation['amount']:.2f} {donation['currency']}",
                    donation['payment_status'],
                    donation['created_at'].strftime('%Y-%m-%d') if donation['created_at'] else ''
                ])
            
            donations_table = Table(table_data, colWidths=[0.5*inch, 2*inch, 1.5*inch, 1.2*inch, 1*inch, 1*inch])
            donations_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#14860C')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            elements.append(donations_table)
        else:
            elements.append(Paragraph("No donations found for the selected period.", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        filename = f"donation_report_{year}_{month:02d}.pdf"
        return send_file(output, mimetype='application/pdf',
                        as_attachment=True, download_name=filename)
        
    except ImportError:
        flash('PDF export requires reportlab library. Install with: pip install reportlab', 'error')
        return redirect(url_for('donation_report.donation_report_page'))
    except Exception as e:
        flash(f'Error exporting to PDF: {str(e)}', 'error')
        return redirect(url_for('donation_report.donation_report_page'))

@donation_report.route('/chart-data')
@login_required
@role_required('Super Admin', 'Admin')
def chart_data():
    """Get chart data as JSON"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True, dictionary=True)
    
    try:
        chart_type = request.args.get('type', 'monthly')
        
        if chart_type == 'monthly':
            # Monthly trend
            cursor.execute("""
                SELECT DATE_FORMAT(created_at, '%%Y-%%m') as month,
                       COALESCE(SUM(amount), 0) as total_amount,
                       COUNT(*) as count
                FROM donations
                WHERE payment_status IN ('Paid', 'Completed', 'Pending')
                AND created_at >= DATE_SUB(NOW(), INTERVAL 12 MONTH)
                GROUP BY DATE_FORMAT(created_at, '%%Y-%%m')
                ORDER BY month ASC
            """)
            data = cursor.fetchall()
            return jsonify({
                'labels': [row['month'] for row in data],
                'amounts': [float(row['total_amount']) for row in data],
                'counts': [row['count'] for row in data]
            })
        
        elif chart_type == 'by_type':
            # By type pie chart
            cursor.execute("""
                SELECT dt.name_amharic as label, COALESCE(SUM(d.amount), 0) as value
                FROM donation_types dt
                LEFT JOIN donations d ON dt.id = d.donation_type_id 
                    AND d.payment_status IN ('Paid', 'Completed', 'Pending')
                    AND d.created_at >= DATE_SUB(NOW(), INTERVAL 1 MONTH)
                WHERE dt.status = 'active' OR dt.is_active = 1
                GROUP BY dt.id, dt.name_amharic
                HAVING value > 0
                ORDER BY value DESC
            """)
            data = cursor.fetchall()
            return jsonify({
                'labels': [row['label'] or 'Other' for row in data],
                'values': [float(row['value']) for row in data]
            })
        
        elif chart_type == 'by_status':
            # By status (treat Pending as Completed)
            cursor.execute("""
                SELECT 
                    CASE WHEN payment_status = 'Pending' THEN 'Completed' ELSE payment_status END as label,
                    COALESCE(SUM(amount), 0) as value,
                    COUNT(*) as count
                FROM donations
                WHERE created_at >= DATE_SUB(NOW(), INTERVAL 1 MONTH)
                GROUP BY CASE WHEN payment_status = 'Pending' THEN 'Completed' ELSE payment_status END
            """)
            data = cursor.fetchall()
            return jsonify({
                'labels': [row['label'] for row in data],
                'values': [float(row['value']) for row in data],
                'counts': [row['count'] for row in data]
            })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

