"""
Modularized Flask Application for Member and Attendance Management
Main entry point - imports from modules
"""
# Standard library imports
import csv
import io
from io import BytesIO, StringIO
from datetime import datetime, date, timedelta
from math import ceil

# Third-party imports
import pandas as pd
from flask import (
    Flask, render_template, request, redirect, url_for, 
    session, flash, send_file, make_response, jsonify, Response
)
from werkzeug.utils import secure_filename
import mysql.connector
from openpyxl import Workbook
from reportlab.lib.pagesizes import A3, landscape, letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, KeepInFrame
)
from reportlab.pdfgen import canvas
from fpdf import FPDF

# Local imports
from config import Config
from database import (
    get_db_connection, 
    initialize_rbac_tables, 
    initialize_default_roles_and_routes,
    test_db_connection
)
from auth import (
    login_required, 
    role_required, 
    get_authorized_routes,
    verify_password,
    hash_password
)
from utils import get_last_10_weeks_weekends, get_members
from translations import get_text, TRANSLATIONS
from mobile_api import mobile_api

# Initialize Flask app
app = Flask(__name__)

# Register Mobile API Blueprint
app.register_blueprint(mobile_api)
app.config.from_object(Config)

# ========================================
# LANGUAGE SUPPORT
# ========================================

@app.context_processor
def inject_language():
    """Inject language and translation function into all templates"""
    current_lang = session.get('lang', 'am')  # Default to Amharic
    return {
        'lang': current_lang,
        't': lambda key: get_text(key, current_lang),
        'get_text': lambda key, lang=None: get_text(key, lang or current_lang)
    }

@app.route('/set_language/<lang>')
def set_language(lang):
    """Set user's preferred language"""
    if lang in ['am', 'en']:
        session['lang'] = lang
        flash(f'Language changed to {"Amharic (አማርኛ)" if lang == "am" else "English"}', 'success')
    return redirect(request.referrer or url_for('navigation'))

# ========================================
# SESSION MANAGEMENT
# ========================================

@app.before_request
def make_session_permanent():
    """Make session permanent for timeout management"""
    session.permanent = True

def track_session_timeout():
    """Track and enforce session timeout"""
    if 'last_activity' in session:
        now = datetime.now()
        last_activity_str = session['last_activity']
        try:
            last_activity = datetime.fromisoformat(last_activity_str)
            timeout_seconds = Config.SESSION_TIMEOUT_MINUTES * 60
            if (now - last_activity).total_seconds() > timeout_seconds:
                session.clear()
                flash("Session timed out due to inactivity.", "warning")
                return redirect(url_for('login'))
        except:
            pass
    session['last_activity'] = datetime.now().isoformat()

# ========================================
# CONTEXT PROCESSORS
# ========================================

@app.context_processor
def inject_navigation():
    """Inject authorized routes into all templates"""
    return dict(authorized_routes=get_authorized_routes())

# ========================================
# AUTHENTICATION ROUTES
# ========================================

@app.route('/', methods=['GET', 'POST'])
def login():
    """Login page and authentication"""
    if request.method == 'POST':
        payroll_number = request.form['payroll_number']
        password = request.form['password']

        conn = get_db_connection()
        cursor = conn.cursor(buffered=True)
        
        # Get user with hashed password
        cursor.execute("""
            SELECT u.*, r.role_name 
            FROM aawsa_user u
            JOIN roles r ON u.role_id = r.role_id
            WHERE u.payroll_number = %(payroll_number)s
        """, {"payroll_number": payroll_number})
        
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if user and verify_password(user[4], password):  # user[4] is password_hash
            session['payroll_number'] = payroll_number
            session['role'] = user[-1]  # role_name is last column
            session['last_activity'] = datetime.now().isoformat()
            flash("Login successful!", "success")
            return redirect(url_for('navigation'))
        else:
            flash("Invalid credentials", "danger")

    return render_template('login.html')

@app.route('/logout')
def logout():
    """Logout and clear session"""
    session.clear()
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))

# ========================================
# DASHBOARD & NAVIGATION
# ========================================

@app.route('/navigation')
@login_required
def navigation():
    """Main navigation/dashboard page with real statistics"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get total members count
    cursor.execute("SELECT COUNT(*) FROM member_registration")
    total_members = cursor.fetchone()[0] or 0
    
    # Get members by section (for cards)
    cursor.execute("""
        SELECT section_name, COUNT(*) 
        FROM member_registration 
        GROUP BY section_name
    """)
    section_counts = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Get counts for each section card
    children_count = section_counts.get('የሕፃናት ክፍል', 0)
    middle_count = section_counts.get('ማህከላዊያን ክፍል', 0)
    youth_count = section_counts.get('ወጣት ክፍል', 0)
    parent_count = section_counts.get('ወላጅ ክፍል', 0)
    
    # Get attendance statistics for current month
    cursor.execute("""
        SELECT COUNT(*) 
        FROM attendance 
        WHERE MONTH(attendance_date) = MONTH(CURDATE())
        AND YEAR(attendance_date) = YEAR(CURDATE())
    """)
    current_month_attendance = cursor.fetchone()[0] or 0
    
    # Get data for charts - Members by section
    cursor.execute("""
        SELECT section_name, COUNT(*) as count
        FROM member_registration
        GROUP BY section_name
        ORDER BY section_name
    """)
    section_data = cursor.fetchall()
    branch_labels = [row[0] for row in section_data]
    channel_amounts = [row[1] for row in section_data]
    
    # Get gender distribution by section for stacked bar chart
    cursor.execute("""
        SELECT section_name, gender, COUNT(*) as count
        FROM member_registration
        GROUP BY section_name, gender
        ORDER BY section_name, gender
    """)
    gender_data = cursor.fetchall()
    
    # Organize data for stacked chart
    sections = list(set([row[0] for row in gender_data]))
    male_data = []
    female_data = []
    
    for section in sections:
        male_count = sum([row[2] for row in gender_data if row[0] == section and row[1] in ('ወንድ', 'Male')])
        female_count = sum([row[2] for row in gender_data if row[0] == section and row[1] in ('ሴት', 'Female')])
        male_data.append(male_count)
        female_data.append(female_count)
    
    # Get summary report data by section
    cursor.execute("""
        SELECT 
            m.section_name,
            SUM(CASE WHEN m.gender IN ('ወንድ', 'Male') THEN 1 ELSE 0 END) as male_count,
            SUM(CASE WHEN m.gender IN ('ሴት', 'Female') THEN 1 ELSE 0 END) as female_count,
            SUM(CASE WHEN m.marital_status IN ('ያላገባ', 'ያላገባች', 'Single') THEN 1 ELSE 0 END) as single_count,
            SUM(CASE WHEN m.marital_status IN ('ያገባ', 'Married') THEN 1 ELSE 0 END) as married_count,
            SUM(CASE WHEN m.work_status IN ('በሥራ ላይ', 'Employed') THEN 1 ELSE 0 END) as employed_count,
            SUM(CASE WHEN m.work_status IN ('ስራ የለኝም', 'Unemployed') THEN 1 ELSE 0 END) as unemployed_count,
            SUM(CASE WHEN m.work_status IN ('ስራ በመፈለግ ላይ') THEN 1 ELSE 0 END) as seeking_count,
            SUM(CASE WHEN m.work_status IN ('ተማሪ', 'Student') OR m.student IN ('Yes', 'አዎ') THEN 1 ELSE 0 END) as student_count
        FROM member_registration m
        GROUP BY m.section_name
        ORDER BY m.section_name
    """)
    current_settled = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    # Current period (month/year)
    current_period = datetime.now().strftime('%B %Y')
    
    return render_template('navigation.html',
                         # Summary report table - IMPORTANT!
                         current_settled=current_settled,
                         current_delivery=[],
                         current_unsettled=[],
                         recent_sent_bills=[],
                         # Card 1: Total Members
                         current_bill_count=total_members,
                         bill_period=current_period,
                         # Card 2: ማህከላዊያን ክፍል (Middle Section)
                         current_month_total_bill_amnt=middle_count,
                         current_month_bill_amnt='ማህከላዊያን ክፍል',
                         # Card 3: የሕፃናት ክፍል (Children Section)
                         payment_count=children_count,
                         payment_month='የሕፃናት ክፍል',
                         # Card 4: ወጣት ክፍል (Youth Section)
                         payment_total_bill_amnt=youth_count,
                         payment_this_month_bill_amnt='ወጣት ክፍል',
                         # Card 5: ወላጅ ክፍል (Parent Section)
                         parent_section_count=parent_count,
                         # Additional stats
                         total_paid_amount=0,
                         total_sent_amount=0,
                         number_sent=0,
                         number_paid=0,
                         # Chart data - Members by section (pie chart)
                         channel_labels=branch_labels,
                         channel_amounts=channel_amounts,
                         # Chart data - Stacked bar (gender by section)
                         branch_labels=sections,
                         number_sent_data=male_data,
                         total_amount_sent_data=female_data,
                         outstanding_amount=[],
                         this_month_bill_amnt=[],
                         payment_data=[])

@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard page"""
    return render_template('dashboard.html')

@app.route('/dashboard-data')
@login_required
def dashboard_data():
    """API endpoint that returns dashboard data as JSON"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get total members
    cursor.execute("SELECT COUNT(*) FROM member_registration")
    total_members = cursor.fetchone()[0] or 0
    
    # Get members by section (for cards)
    cursor.execute("""
        SELECT section_name, COUNT(*) 
        FROM member_registration 
        GROUP BY section_name
    """)
    section_counts = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Get counts for each section card
    children_count = section_counts.get('የሕፃናት ክፍል', 0)
    middle_count = section_counts.get('ማህከላዊያን ክፍል', 0)
    youth_count = section_counts.get('ወጣት ክፍል', 0)
    parent_count = section_counts.get('ወላጅ ክፍል', 0)
    
    # Get data for charts - Members by section
    cursor.execute("""
        SELECT section_name, COUNT(*) as count
        FROM member_registration
        GROUP BY section_name
        ORDER BY section_name
    """)
    section_data = cursor.fetchall()
    branch_labels = [row[0] for row in section_data]
    channel_amounts = [row[1] for row in section_data]
    
    # Get gender distribution by section for stacked bar chart
    cursor.execute("""
        SELECT section_name, gender, COUNT(*) as count
        FROM member_registration
        GROUP BY section_name, gender
        ORDER BY section_name, gender
    """)
    gender_data = cursor.fetchall()
    
    # Organize data for stacked chart
    sections = list(set([row[0] for row in gender_data]))
    male_data = []
    female_data = []
    
    for section in sections:
        male_count = sum([row[2] for row in gender_data if row[0] == section and row[1] in ('ወንድ', 'Male')])
        female_count = sum([row[2] for row in gender_data if row[0] == section and row[1] in ('ሴት', 'Female')])
        male_data.append(male_count)
        female_data.append(female_count)
    
    # Get summary report data by section
    cursor.execute("""
        SELECT 
            m.section_name,
            SUM(CASE WHEN m.gender IN ('ወንድ', 'Male') THEN 1 ELSE 0 END) as male_count,
            SUM(CASE WHEN m.gender IN ('ሴት', 'Female') THEN 1 ELSE 0 END) as female_count,
            SUM(CASE WHEN m.marital_status IN ('ያላገባ', 'ያላገባች', 'Single') THEN 1 ELSE 0 END) as single_count,
            SUM(CASE WHEN m.marital_status IN ('ያገባ', 'Married') THEN 1 ELSE 0 END) as married_count,
            SUM(CASE WHEN m.work_status IN ('በሥራ ላይ', 'Employed') THEN 1 ELSE 0 END) as employed_count,
            SUM(CASE WHEN m.work_status IN ('ስራ የለኝም', 'Unemployed') THEN 1 ELSE 0 END) as unemployed_count,
            SUM(CASE WHEN m.work_status IN ('ስራ በመፈለግ ላይ') THEN 1 ELSE 0 END) as seeking_count,
            SUM(CASE WHEN m.work_status IN ('ተማሪ', 'Student') OR m.student IN ('Yes', 'አዎ') THEN 1 ELSE 0 END) as student_count
        FROM member_registration m
        GROUP BY m.section_name
        ORDER BY m.section_name
    """)
    current_settled = [[col for col in row] for row in cursor.fetchall()]
    
    cursor.close()
    conn.close()
    
    # Current period (month/year)
    current_period = datetime.now().strftime('%B %Y')
    
    data = {
        'current_bill_count': total_members,
        'bill_period': current_period,
        'current_month_total_bill_amnt': middle_count,
        'current_month_bill_amnt': middle_count,
        'payment_count': children_count,
        'payment_month': 'የሕፃናት ክፍል',
        'payment_total_bill_amnt': youth_count,
        'payment_this_month_bill_amnt': 'ወጣት ክፍል',
        'parent_section_count': parent_count,
        'current_settled': current_settled,
        'current_delivery': [],
        'current_unsettled': [],
        'branch_labels': sections,
        'number_sent_data': male_data,
        'total_amount_sent_data': female_data,
        'outstanding_amount': [],
        'this_month_bill_amnt': [],
        'channel_labels': branch_labels,
        'channel_amounts': channel_amounts
    }
    
    # Debug logging
    print(f"Dashboard data - Summary report rows: {len(current_settled)}")
    if current_settled:
        print(f"First row: {current_settled[0]}")
    
    return jsonify(data)

@app.route('/admin_dashboard')
@login_required
@role_required('Super Admin')
def admin_dashboard():
    """Admin dashboard"""
    return render_template('admin_dashboard.html')

# ========================================
# MEMBER MANAGEMENT ROUTES
# ========================================

@app.route('/manage_members', methods=['GET', 'POST'])
@login_required
def manage_members():
    """Manage member registrations - CRUD operations"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    member_data = {}

    if request.method == 'POST':
        form_data = request.form.to_dict()

        # Rename spiritual_education -> education
        form_data['education'] = form_data.pop('spiritual_education', None)

        # If ID is not provided, make sure it's not in the dict
        member_id = form_data.get('id')
        if not member_id:
            form_data.pop('id', None)

        if member_id:  # UPDATE
            query = """
                UPDATE member_registration SET
                    full_name=%(full_name)s,
                    father_of_repentance=%(father_of_repentance)s,
                    mother_name=%(mother_name)s,
                    parish=%(parish)s,
                    christian_name=%(christian_name)s,
                    age_of_birth=%(age_of_birth)s,
                    subcity=%(subcity)s,
                    woreda=%(woreda)s,
                    house_number=%(house_number)s,
                    special_place=%(special_place)s,
                    phone=%(phone)s,
                    leving=%(leving)s,
                    work_status=%(work_status)s,
                    email=%(email)s,
                    member_year=%(member_year)s,
                    education=%(education)s,
                    `rank`=%(rank)s,
                    education_status=%(education_status)s,
                    work=%(work)s,
                    student=%(student)s,
                    career=%(career)s,
                    marital_status=%(marital_status)s,
                    section_name=%(section_name)s,
                    gender=%(gender)s
                WHERE id=%(id)s
            """
        else:  # INSERT
            query = """
                INSERT INTO member_registration (
                    full_name, father_of_repentance, mother_name, parish, christian_name,
                    age_of_birth, subcity, woreda, house_number, special_place,
                    phone, leving, work_status, email, member_year, education,
                    `rank`, education_status, work, student, career, marital_status,section_name,gender
                ) VALUES (
                    %(full_name)s, %(father_of_repentance)s, %(mother_name)s, %(parish)s, %(christian_name)s,
                    %(age_of_birth)s, %(subcity)s, %(woreda)s, %(house_number)s, %(special_place)s,
                    %(phone)s, %(leving)s, %(work_status)s, %(email)s, %(member_year)s, %(education)s,
                    %(rank)s, %(education_status)s, %(work)s, %(student)s, %(career)s, %(marital_status)s, %(section_name)s, %(gender)s
                )
            """

        try:
            cursor.execute(query, form_data)
            conn.commit()
            flash("Member saved successfully.", "success")
            return redirect(url_for('manage_members'))
        except Exception as e:
            flash(f"Database Error: {str(e)}", "danger")

    # DELETE
    delete_id = request.args.get('delete')
    if delete_id:
        try:
            cursor.execute("DELETE FROM member_registration WHERE id = %(id)s", {'id': delete_id})
            conn.commit()
            flash("Member deleted.", "info")
            return redirect(url_for('manage_members'))
        except Exception as e:
            flash(f"Error deleting member: {str(e)}", "danger")

    # EDIT
    edit_id = request.args.get('edit')
    if edit_id:
        cursor.execute("SELECT * FROM member_registration WHERE id = %(id)s", {'id': edit_id})
        row = cursor.fetchone()
        if row:
            col_names = [desc[0].lower() for desc in cursor.description]
            member_data = dict(zip(col_names, row))

    # Get filters
    search = request.args.get('search', '')
    section_filter = request.args.get('section', '')
    gender_filter = request.args.get('gender', '')
    
    # Build query with filters
    query = "SELECT * FROM member_registration WHERE 1=1"
    params = []
    
    if search:
        query += " AND full_name LIKE %s"
        params.append(f'%{search}%')
    
    if section_filter:
        query += " AND section_name = %s"
        params.append(section_filter)
    
    if gender_filter:
        query += " AND gender = %s"
        params.append(gender_filter)
    
    query += " ORDER BY id DESC"
    
    cursor.execute(query, params)
    members = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("manage_members.html", 
                         members=members, 
                         member_data=member_data,
                         search=search,
                         section_filter=section_filter,
                         gender_filter=gender_filter)

@app.route('/member_report')
@login_required
def member_report():
    """Comprehensive member report with statistics and charts"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filters
    section_filter = request.args.get('section', '')
    gender_filter = request.args.get('gender', '')
    year_filter = request.args.get('year', datetime.now().year)
    
    # Overall statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_members,
            COUNT(CASE WHEN gender = 'ወንድ' THEN 1 END) as male_count,
            COUNT(CASE WHEN gender = 'ሴት' THEN 1 END) as female_count,
            COUNT(CASE WHEN marital_status IN ('ያላገባ', 'ያላገባች') THEN 1 END) as single_count,
            COUNT(CASE WHEN marital_status = 'ያገባ' THEN 1 END) as married_count,
            COUNT(CASE WHEN work_status = 'በሥራ ላይ' THEN 1 END) as employed_count,
            COUNT(CASE WHEN work_status = 'ስራ የለኝም' THEN 1 END) as unemployed_count,
            COUNT(CASE WHEN work_status = 'ተማሪ' OR student = 'Yes' THEN 1 END) as student_count
        FROM member_registration
    """)
    stats = cursor.fetchone()
    overall_stats = {
        'total_members': stats[0] or 0,
        'male_count': stats[1] or 0,
        'female_count': stats[2] or 0,
        'single_count': stats[3] or 0,
        'married_count': stats[4] or 0,
        'employed_count': stats[5] or 0,
        'unemployed_count': stats[6] or 0,
        'student_count': stats[7] or 0
    }
    
    # Section-wise statistics
    cursor.execute("""
        SELECT 
            section_name,
            COUNT(*) as total,
            COUNT(CASE WHEN gender = 'ወንድ' THEN 1 END) as male,
            COUNT(CASE WHEN gender = 'ሴት' THEN 1 END) as female
        FROM member_registration
        GROUP BY section_name
        ORDER BY section_name
    """)
    section_stats = cursor.fetchall()
    
    # Age distribution (approximate)
    cursor.execute("""
        SELECT 
            CASE 
                WHEN TIMESTAMPDIFF(YEAR, age_of_birth, CURDATE()) < 18 THEN 'Under 18'
                WHEN TIMESTAMPDIFF(YEAR, age_of_birth, CURDATE()) BETWEEN 18 AND 30 THEN '18-30'
                WHEN TIMESTAMPDIFF(YEAR, age_of_birth, CURDATE()) BETWEEN 31 AND 50 THEN '31-50'
                WHEN TIMESTAMPDIFF(YEAR, age_of_birth, CURDATE()) > 50 THEN 'Over 50'
                ELSE 'Unknown'
            END as age_group,
            COUNT(*) as count
        FROM member_registration
        WHERE age_of_birth IS NOT NULL
        GROUP BY age_group
        ORDER BY age_group
    """)
    age_distribution = cursor.fetchall()
    
    # Education statistics
    cursor.execute("""
        SELECT 
            education_status,
            COUNT(*) as count
        FROM member_registration
        WHERE education_status IS NOT NULL AND education_status != ''
        GROUP BY education_status
        ORDER BY count DESC
    """)
    education_stats = cursor.fetchall()
    
    # Subcity distribution
    cursor.execute("""
        SELECT 
            subcity,
            COUNT(*) as count
        FROM member_registration
        WHERE subcity IS NOT NULL AND subcity != ''
        GROUP BY subcity
        ORDER BY count DESC
        LIMIT 10
    """)
    subcity_stats = cursor.fetchall()
    
    # Build filtered query for detailed list
    query = """
        SELECT id, full_name, gender, section_name, phone, email, 
               subcity, marital_status, work_status, education_status
        FROM member_registration
        WHERE 1=1
    """
    params = []
    
    if section_filter:
        query += " AND section_name = %s"
        params.append(section_filter)
    
    if gender_filter:
        query += " AND gender = %s"
        params.append(gender_filter)
    
    query += " ORDER BY section_name, full_name LIMIT 100"
    
    cursor.execute(query, params)
    members_list = cursor.fetchall()
    
    sections = ['የሕፃናት ክፍል', 'ማህከላዊያን ክፍል', 'ወጣት ክፍል', 'ወላጅ ክፍል']
    
    cursor.close()
    conn.close()
    
    return render_template('member_report.html',
                         overall_stats=overall_stats,
                         section_stats=section_stats,
                         age_distribution=age_distribution,
                         education_stats=education_stats,
                         subcity_stats=subcity_stats,
                         members_list=members_list,
                         sections=sections,
                         section_filter=section_filter,
                         gender_filter=gender_filter,
                         year_filter=year_filter)

@app.route('/upload_member_registration', methods=['GET', 'POST'])
@login_required
def upload_member_registration():
    """Upload member registrations from CSV/Excel file with validation"""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash("Please select a file to upload.", "warning")
            return redirect(url_for('upload_member_registration'))
        
        # Check file extension
        filename = file.filename.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            flash("Please upload a CSV or Excel file (.csv, .xlsx, .xls)", "warning")
            return redirect(url_for('upload_member_registration'))

        try:
            # Read file with pandas
            if filename.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            
            # Validate required columns
            required_columns = ['full_name', 'phone', 'section_name', 'gender']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                flash(f"Missing required columns: {', '.join(missing_columns)}", "danger")
                return redirect(url_for('upload_member_registration'))
            
            conn = get_db_connection()
            cursor = conn.cursor(buffered=True)
            
            # Track results
            success_count = 0
            duplicate_count = 0
            error_count = 0
            duplicates = []
            errors = []
            
            for index, row in df.iterrows():
                row_num = index + 2  # Excel row number (1-indexed + header)
                
                try:
                    full_name = str(row.get('full_name', '')).strip()
                    phone = str(row.get('phone', '')).strip()
                    
                    # Validate required fields
                    if not full_name or full_name == 'nan':
                        errors.append(f"Row {row_num}: Missing full name")
                        error_count += 1
                        continue
                    
                    # Check for duplicates by phone or full name
                    cursor.execute("""
                        SELECT id, full_name, phone FROM member_registration 
                        WHERE phone = %s OR full_name = %s
                    """, (phone if phone and phone != 'nan' else None, full_name))
                    
                    existing = cursor.fetchone()
                    if existing:
                        duplicates.append({
                            'row': row_num,
                            'name': full_name,
                            'phone': phone if phone and phone != 'nan' else 'N/A',
                            'existing_name': existing[1],
                            'existing_phone': existing[2] or 'N/A'
                        })
                        duplicate_count += 1
                        continue
                    
                    # Prepare data with defaults
                    data = {
                        'full_name': full_name,
                        'father_of_repentance': str(row.get('father_of_repentance', '')).strip() if pd.notna(row.get('father_of_repentance')) else None,
                        'mother_name': str(row.get('mother_name', '')).strip() if pd.notna(row.get('mother_name')) else None,
                        'parish': str(row.get('parish', '')).strip() if pd.notna(row.get('parish')) else None,
                        'christian_name': str(row.get('christian_name', '')).strip() if pd.notna(row.get('christian_name')) else None,
                        'age_of_birth': row.get('age_of_birth') if pd.notna(row.get('age_of_birth')) else None,
                        'subcity': str(row.get('subcity', '')).strip() if pd.notna(row.get('subcity')) else None,
                        'woreda': str(row.get('woreda', '')).strip() if pd.notna(row.get('woreda')) else None,
                        'house_number': str(row.get('house_number', '')).strip() if pd.notna(row.get('house_number')) else None,
                        'special_place': str(row.get('special_place', '')).strip() if pd.notna(row.get('special_place')) else None,
                        'phone': phone if phone and phone != 'nan' else None,
                        'leving': str(row.get('leving', '')).strip() if pd.notna(row.get('leving')) else None,
                        'work_status': str(row.get('work_status', '')).strip() if pd.notna(row.get('work_status')) else None,
                        'email': str(row.get('email', '')).strip() if pd.notna(row.get('email')) else None,
                        'member_year': row.get('member_year') if pd.notna(row.get('member_year')) else None,
                        'education': str(row.get('education', '')).strip() if pd.notna(row.get('education')) else None,
                        'rank': str(row.get('rank', '')).strip() if pd.notna(row.get('rank')) else None,
                        'education_status': str(row.get('education_status', '')).strip() if pd.notna(row.get('education_status')) else None,
                        'work': str(row.get('work', '')).strip() if pd.notna(row.get('work')) else None,
                        'student': str(row.get('student', '')).strip() if pd.notna(row.get('student')) else None,
                        'career': str(row.get('career', '')).strip() if pd.notna(row.get('career')) else None,
                        'marital_status': str(row.get('marital_status', '')).strip() if pd.notna(row.get('marital_status')) else None,
                        'section_name': str(row.get('section_name', '')).strip() if pd.notna(row.get('section_name')) else None,
                        'gender': str(row.get('gender', '')).strip() if pd.notna(row.get('gender')) else None
                    }
                    
                    # Insert member
                    cursor.execute("""
                        INSERT INTO member_registration (
                            full_name, father_of_repentance, mother_name, parish, christian_name,
                            age_of_birth, subcity, woreda, house_number, special_place,
                            phone, leving, work_status, email, member_year,
                            education, `rank`, education_status, work, student,
                            career, marital_status, section_name, gender
                        ) VALUES (
                            %(full_name)s, %(father_of_repentance)s, %(mother_name)s, %(parish)s, %(christian_name)s,
                            %(age_of_birth)s, %(subcity)s, %(woreda)s, %(house_number)s, %(special_place)s,
                            %(phone)s, %(leving)s, %(work_status)s, %(email)s, %(member_year)s,
                            %(education)s, %(rank)s, %(education_status)s, %(work)s, %(student)s,
                            %(career)s, %(marital_status)s, %(section_name)s, %(gender)s
                        )
                    """, data)
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
                    continue
            
            conn.commit()
            cursor.close()
            conn.close()
            
            # Return results page with details
            return render_template('upload_member_registration.html',
                                 upload_complete=True,
                                 success_count=success_count,
                                 duplicate_count=duplicate_count,
                                 error_count=error_count,
                                 total_rows=len(df),
                                 duplicates=duplicates,
                                 errors=errors)
            
        except Exception as e:
            flash(f"Failed to process file: {str(e)}", "danger")
            return redirect(url_for('upload_member_registration'))

    return render_template('upload_member_registration.html', upload_complete=False)

@app.route('/download_member_template')
@login_required
def download_member_template():
    """Download Excel template with dropdowns for member upload"""
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Create workbook with hidden reference sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    
    # Create a hidden sheet for dropdown values
    ws_ref = wb.create_sheet("DropdownValues")
    ws_ref.sheet_state = 'hidden'
    
    # Add dropdown value lists to hidden sheet
    section_values = ['የሕፃናት ክፍል', 'ወጣት ክፍል', 'ማህከላዊያን ክፍል', 'ወላጅ ክፍል']
    gender_values = ['ወንድ', 'ሴት']
    work_values = ['በሥራ ላይ', 'ስራ የለኝም', 'በመፈለግ ላይ', 'ተማሪ']
    marital_values = ['ያገባ', 'ያላገባ', 'ያላገባች', 'የፈታ', 'የፈታች']
    subcity_values = ['Addis Ketema', 'Akaki Kality', 'Arada', 'Bole', 'Gullele', 
                      'Kirkos', 'Kolfe Keranio', 'Lemi Kura', 'Lideta', 'Nifas Silk-Lafto', 'Yeka']
    education_status_values = ['Elementary', 'High School', 'Diploma', 'Degree', 'Masters', 'PhD', 'Other']
    student_values = ['Yes', 'No']
    leving_values = ['ያገባ', 'ያላገባ', 'ያላገባች', 'የፈታ', 'የፈታች']
    
    # Write values to hidden sheet (each column)
    for i, val in enumerate(section_values, 1):
        ws_ref.cell(row=i, column=1, value=val)
    for i, val in enumerate(gender_values, 1):
        ws_ref.cell(row=i, column=2, value=val)
    for i, val in enumerate(work_values, 1):
        ws_ref.cell(row=i, column=3, value=val)
    for i, val in enumerate(marital_values, 1):
        ws_ref.cell(row=i, column=4, value=val)
    for i, val in enumerate(subcity_values, 1):
        ws_ref.cell(row=i, column=5, value=val)
    for i, val in enumerate(education_status_values, 1):
        ws_ref.cell(row=i, column=6, value=val)
    for i, val in enumerate(student_values, 1):
        ws_ref.cell(row=i, column=7, value=val)
    for i, val in enumerate(leving_values, 1):
        ws_ref.cell(row=i, column=8, value=val)
    
    # Define columns in order
    columns = [
        'full_name', 'father_of_repentance', 'mother_name', 'parish', 'christian_name',
        'age_of_birth', 'subcity', 'woreda', 'house_number', 'special_place',
        'phone', 'leving', 'work_status', 'email', 'member_year',
        'education', 'rank', 'education_status', 'work', 'student',
        'career', 'marital_status', 'section_name', 'gender'
    ]
    
    # Write headers with formatting
    header_fill = PatternFill(start_color="14860C", end_color="14860C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Auto-adjust column width
        ws.column_dimensions[cell.column_letter].width = 18
    
    # Add instruction row
    instruction_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    instructions = [
        'REQUIRED', 'Optional', 'Optional', 'Optional', 'Optional',
        'YYYY-MM-DD', 'Use dropdown ↓', 'Optional', 'Optional', 'Optional',
        'REQUIRED', 'Use dropdown ↓', 'Use dropdown ↓', 'Optional', 'Year',
        'Optional', 'Optional', 'Use dropdown ↓', 'Optional', 'Use dropdown ↓',
        'Optional', 'Use dropdown ↓', 'Use dropdown ↓', 'Use dropdown ↓'
    ]
    
    for col_idx, instruction in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col_idx, value=instruction)
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(horizontal='center')
    
    # Add example data rows
    example_data = [
        ['John Doe', 'Father Name', 'Mother Name', 'St. Mary', 'Giorgis',
         '1990-01-01', 'Addis Ketema', '01', '123', 'Near Church',
         '0911234567', 'ያገባ', 'በሥራ ላይ', 'john@example.com', 2020,
         'University', 'Deacon', 'Graduate', 'Engineer', 'No',
         'Software', 'ያገባ', 'የሕፃናት ክፍል', 'ወንድ'],
        ['Jane Smith', 'Father Name', 'Mother Name', 'St. George', 'Mariam',
         '1992-05-15', 'Bole', '02', '456', 'Behind School',
         '0922345678', 'ያላገባ', 'ስራ የለኝም', 'jane@example.com', 2021,
         'High School', 'None', 'Student', 'Teacher', 'Yes',
         'Education', 'ያላገባች', 'ወጣት ክ�ፍል', 'ሴት']
    ]
    
    for row_idx, row_data in enumerate(example_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Find column indices for dropdowns
    section_col = columns.index('section_name') + 1  # 23
    gender_col = columns.index('gender') + 1  # 24
    work_col = columns.index('work_status') + 1  # 13
    marital_col = columns.index('marital_status') + 1  # 22
    subcity_col = columns.index('subcity') + 1  # 7
    education_status_col = columns.index('education_status') + 1  # 18
    student_col = columns.index('student') + 1  # 20
    leving_col = columns.index('leving') + 1  # 12
    
    # Convert column index to letter
    from openpyxl.utils import get_column_letter
    
    section_letter = get_column_letter(section_col)
    gender_letter = get_column_letter(gender_col)
    work_letter = get_column_letter(work_col)
    marital_letter = get_column_letter(marital_col)
    subcity_letter = get_column_letter(subcity_col)
    education_status_letter = get_column_letter(education_status_col)
    student_letter = get_column_letter(student_col)
    leving_letter = get_column_letter(leving_col)
    
    # Add data validation (dropdowns) using hidden sheet reference
    # Section Name dropdown
    section_dv = DataValidation(
        type="list",
        formula1='DropdownValues!$A$1:$A$4',
        allow_blank=True
    )
    section_dv.error = 'Options: የሕፃናት ክፍል, ወጣት ክፍል, ማህከላዊያን ክፍል, ወላጅ ክፍል'
    section_dv.errorTitle = 'Invalid Section'
    section_dv.prompt = 'Click dropdown arrow to select'
    section_dv.promptTitle = 'Select Section'
    ws.add_data_validation(section_dv)
    section_dv.add(f'{section_letter}3:{section_letter}1000')
    
    # Gender dropdown
    gender_dv = DataValidation(
        type="list",
        formula1='DropdownValues!$B$1:$B$2',
        allow_blank=True
    )
    gender_dv.error = 'Options: ወንድ, ሴት'
    gender_dv.errorTitle = 'Invalid Gender'
    gender_dv.prompt = 'Click dropdown arrow to select'
    gender_dv.promptTitle = 'Select Gender'
    ws.add_data_validation(gender_dv)
    gender_dv.add(f'{gender_letter}3:{gender_letter}1000')
    
    # Work Status dropdown
    work_dv = DataValidation(
        type="list",
        formula1='DropdownValues!$C$1:$C$4',
        allow_blank=True
    )
    work_dv.error = 'Options: በሥራ ላይ, ስራ የለኝም, በመፈለግ ላይ, ተማሪ'
    work_dv.errorTitle = 'Invalid Work Status'
    work_dv.prompt = 'Click dropdown arrow to select'
    work_dv.promptTitle = 'Select Work Status'
    ws.add_data_validation(work_dv)
    work_dv.add(f'{work_letter}3:{work_letter}1000')
    
    # Marital Status dropdown
    marital_dv = DataValidation(
        type="list",
        formula1='DropdownValues!$D$1:$D$5',
        allow_blank=True
    )
    marital_dv.error = 'Options: ያገባ, ያላገባ, ያላገባች, የፈታ, የፈታች'
    marital_dv.errorTitle = 'Invalid Marital Status'
    marital_dv.prompt = 'Click dropdown arrow to select'
    marital_dv.promptTitle = 'Select Marital Status'
    ws.add_data_validation(marital_dv)
    marital_dv.add(f'{marital_letter}3:{marital_letter}1000')
    
    # Subcity dropdown
    subcity_dv = DataValidation(
        type="list",
        formula1='DropdownValues!$E$1:$E$11',
        allow_blank=True
    )
    subcity_dv.error = 'Select from: Addis Ketema, Akaki Kality, Arada, Bole, Gullele, Kirkos, Kolfe Keranio, Lemi Kura, Lideta, Nifas Silk-Lafto, Yeka'
    subcity_dv.errorTitle = 'Invalid Subcity'
    subcity_dv.prompt = 'Select Addis Ababa subcity'
    subcity_dv.promptTitle = 'Subcity'
    ws.add_data_validation(subcity_dv)
    subcity_dv.add(f'{subcity_letter}3:{subcity_letter}1000')
    
    # Education Status dropdown
    education_status_dv = DataValidation(
        type="list",
        formula1='DropdownValues!$F$1:$F$7',
        allow_blank=True
    )
    education_status_dv.error = 'Options: Elementary, High School, Diploma, Degree, Masters, PhD, Other'
    education_status_dv.errorTitle = 'Invalid Education Status'
    education_status_dv.prompt = 'Select education level'
    education_status_dv.promptTitle = 'Education Status'
    ws.add_data_validation(education_status_dv)
    education_status_dv.add(f'{education_status_letter}3:{education_status_letter}1000')
    
    # Student dropdown
    student_dv = DataValidation(
        type="list",
        formula1='DropdownValues!$G$1:$G$2',
        allow_blank=True
    )
    student_dv.error = 'Options: Yes, No'
    student_dv.errorTitle = 'Invalid Value'
    student_dv.prompt = 'Is the person a student?'
    student_dv.promptTitle = 'Student'
    ws.add_data_validation(student_dv)
    student_dv.add(f'{student_letter}3:{student_letter}1000')
    
    # Leving dropdown (living status)
    leving_dv = DataValidation(
        type="list",
        formula1='DropdownValues!$H$1:$H$5',
        allow_blank=True
    )
    leving_dv.error = 'Options: ያገባ, ያላገባ, ያላገባች, የፈታ, የፈታች'
    leving_dv.errorTitle = 'Invalid Living Status'
    leving_dv.prompt = 'Select living status'
    leving_dv.promptTitle = 'Living Status'
    ws.add_data_validation(leving_dv)
    leving_dv.add(f'{leving_letter}3:{leving_letter}1000')
    
    # Freeze first row (header)
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='member_upload_template.xlsx'
    )

@app.route('/download_member_template_csv')
@login_required
def download_member_template_csv():
    """Download CSV template with instructions for member upload"""
    # Create CSV with example data and comments
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header row
    writer.writerow([
        'full_name', 'father_of_repentance', 'mother_name', 'parish', 'christian_name',
        'age_of_birth', 'subcity', 'woreda', 'house_number', 'special_place',
        'phone', 'leving', 'work_status', 'email', 'member_year',
        'education', 'rank', 'education_status', 'work', 'student',
        'career', 'marital_status', 'section_name', 'gender'
    ])
    
    # Write instructions row
    writer.writerow([
        'REQUIRED', 'Optional', 'Optional', 'Optional', 'Optional',
        'YYYY-MM-DD', 'Optional', 'Optional', 'Optional', 'Optional',
        'REQUIRED', 'Optional', 'See values below', 'Optional', 'Year',
        'Optional', 'Optional', 'Optional', 'Optional', 'Yes/No',
        'Optional', 'See values below', 'REQUIRED - See values', 'REQUIRED - ወንድ or ሴት'
    ])
    
    # Write example data rows
    writer.writerow([
        'John Doe', 'Father Name', 'Mother Name', 'St. Mary', 'Giorgis',
        '1990-01-01', 'Addis Ketema', '01', '123', 'Near Church',
        '0911234567', 'ያገባ', 'በሥራ ላይ', 'john@example.com', '2020',
        'University', 'Deacon', 'Graduate', 'Engineer', 'No',
        'Software', 'ያገባ', 'የሕፃናት ክፍል', 'ወንድ'
    ])
    
    writer.writerow([
        'Jane Smith', 'Father Name', 'Mother Name', 'St. George', 'Mariam',
        '1992-05-15', 'Bole', '02', '456', 'Behind School',
        '0922345678', 'ያላገባ', 'ስራ የለኝም', 'jane@example.com', '2021',
        'High School', 'None', 'Student', 'Teacher', 'Yes',
        'Education', 'ያላገባች', 'ወጣት ክፍል', 'ሴት'
    ])
    
    # Add blank rows for separation
    writer.writerow([])
    writer.writerow([])
    
    # Add valid values reference
    writer.writerow(['*** VALID VALUES FOR DROPDOWN FIELDS ***'])
    writer.writerow([])
    writer.writerow(['section_name (ክፍል):', 'የሕፃናት ክፍል', 'ወጣት ክፍል', 'ማህከላዊያን ክፍል', 'ወላጅ ክፍል'])
    writer.writerow(['gender (ፆታ):', 'ወንድ', 'ሴት'])
    writer.writerow(['work_status (የስራ ሁኔታ):', 'በሥራ ላይ', 'ስራ የለኝም', 'በመፈለግ ላይ', 'ተማሪ'])
    writer.writerow(['marital_status (የትዳር ሁኔታ):', 'ያገባ', 'ያላገባ', 'ያላገባች', 'የፈታ', 'የፈታች'])
    writer.writerow([])
    writer.writerow(['*** IMPORTANT ***'])
    writer.writerow(['1. Delete the instruction rows and example rows before uploading'])
    writer.writerow(['2. Delete this reference section before uploading'])
    writer.writerow(['3. Use exact values from the lists above (copy-paste recommended)'])
    writer.writerow(['4. Required fields: full_name, phone, section_name, gender'])
    
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=member_upload_template.csv'}
    )

# ========================================
# ATTENDANCE ROUTES  
# ========================================

@app.route('/attendance', methods=['GET', 'POST'])
@login_required
def attendance():
    """Mark attendance for members by section"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)

    # Fetch distinct section names for dropdown
    cursor.execute("SELECT DISTINCT section_name FROM member_registration ORDER BY section_name")
    section_names = [row[0] for row in cursor.fetchall()]

    selected_section = request.args.get('section')

    if request.method == 'POST':
        # On submit, section name comes as hidden input
        selected_section = request.form.get('selected_section')
        form_data = request.form.to_dict()
        try:
            for key, value in form_data.items():
                if not key.startswith('attendance_'):
                    continue
                if value == '':
                    continue
                _, member_id_str, att_date_str = key.split('_', 2)
                member_id = int(member_id_str)
                attendance_date = att_date_str
                present_status = value
                day_name = date.fromisoformat(attendance_date).strftime('%A')
                remarks = ''

                cursor.execute("""
                    SELECT id FROM attendance 
                    WHERE member_id = %(member_id)s AND DATE(attendance_date) = %(attendance_date)s
                """, {'member_id': member_id, 'attendance_date': attendance_date})
                row = cursor.fetchone()

                if row:
                    cursor.execute("""
                        UPDATE attendance SET
                            present_status = %(present_status)s,
                            day_name = %(day_name)s,
                            remarks = %(remarks)s
                        WHERE id = %(id)s
                    """, {'present_status': present_status, 'day_name': day_name, 'remarks': remarks, 'id': row[0]})
                else:
                    cursor.execute("""
                        INSERT INTO attendance (
                            member_id, attendance_date, present_status, day_name, remarks
                        ) VALUES (
                            %(member_id)s, %(attendance_date)s, %(present_status)s, %(day_name)s, %(remarks)s
                        )
                    """, {'member_id': member_id, 'attendance_date': attendance_date, 'present_status': present_status, 'day_name': day_name, 'remarks': remarks})
            conn.commit()
            flash("Attendance saved successfully.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Error saving attendance: {e}", "danger")
        return redirect(url_for('attendance', section=selected_section))

    members = []
    attendance_data = {}
    dates = []

    if selected_section:
        # Fetch members only in selected section
        cursor.execute("""
            SELECT id, full_name FROM member_registration WHERE section_name = %(section)s ORDER BY full_name
        """, {'section': selected_section})
        members_rows = cursor.fetchall()
        members = [{'member_id': m[0], 'full_name': m[1]} for m in members_rows]

        # Dates for attendance
        dates = get_last_10_weeks_weekends()

        if members:
            member_ids = [m['member_id'] for m in members]
            format_ids = ",".join(str(id) for id in member_ids)
            format_dates = ",".join(f"'{d}'" for d in dates)

            query = f"""
                SELECT member_id, DATE_FORMAT(attendance_date, '%Y-%m-%d') AS att_date, present_status
                FROM attendance
                WHERE member_id IN ({format_ids})
                  AND DATE(attendance_date) IN ({format_dates})
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            attendance_data = {f"{r[0]}_{r[1]}": r[2] for r in rows}

    cursor.close()
    conn.close()

    return render_template('attendance_section.html',
                           section_names=section_names,
                           selected_section=selected_section,
                           members=members,
                           dates=dates,
                           attendance_data=attendance_data)

@app.route('/attendance_report', methods=['GET', 'POST'])
@login_required
def attendance_report():
    """View and export attendance reports"""
    search_query = request.args.get('search', '')
    selected_branch = request.args.get('branch', 'All')
    export_format = request.args.get('format', None)
    page = request.args.get('page', 1, type=int)
    per_page = Config.ITEMS_PER_PAGE
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    if request.method == 'POST':
        search_query = request.form.get('search', '')
        start_date = request.form.get('start_date', '')
        end_date = request.form.get('end_date', '')

    # Database connection
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)

    # Base query for filtering
    query = """
      SELECT FULL_NAME, GENDER, CHRISTIAN_NAME, PHONE, SECTION_NAME, ATTENDANCE_DATE, PRESENT_STATUS 
      FROM MEMBER_REGISTRATION a, attendance b 
      WHERE a.id=b.member_id
    """

    params = {}

    # Add search filter
    if search_query:
        query += " AND (a.FULL_NAME LIKE %(search)s OR a.PHONE LIKE %(search)s)"
        params['search'] = f'%{search_query}%'

    # Add branch/section filter
    if selected_branch != 'All':
        query += " AND a.SECTION_NAME = %(branch)s"
        params['branch'] = selected_branch

    # Add date range filter
    if start_date and end_date:
        try:
            if 'T' in start_date:
                start_date = start_date.replace('T', ' ') + ':00'
            if 'T' in end_date:
                end_date = end_date.replace('T', ' ') + ':00'

            dt_format = '%Y-%m-%d '
            start_date_obj = datetime.strptime(start_date, dt_format)
            end_date_obj = datetime.strptime(end_date, dt_format)

            query += """
            AND b.ATTENDANCE_DATE 
            BETWEEN %(start_date)s 
            AND %(end_date)s
            """
            params['start_date'] = start_date_obj.strftime(dt_format)
            params['end_date'] = end_date_obj.strftime(dt_format)

        except ValueError:
            flash("Invalid date format", "danger")

    # Finalize query order
    query += " ORDER BY FULL_NAME, ATTENDANCE_DATE"

    # Execute query
    cursor.execute(query, params)
    filtered_records = cursor.fetchall()

    # Pagination logic
    total = len(filtered_records)
    pages = ceil(total / per_page)
    start = (page - 1) * per_page
    end = start + per_page
    payments_paginated = filtered_records[start:end]

    # Export functionality
    if export_format == 'csv':
        return export_attendance_csv(filtered_records)
    elif export_format == 'pdf':
        return attendance_report_pdf(filtered_records)
    elif export_format == 'excel':
        return attendance_report_excel(filtered_records)

    # Close database connection
    cursor.close()
    conn.close()

    # Fetch branches/sections for filter dropdown
    branches = get_members()

    # Render template
    return render_template(
        'attendance_report.html',
        payments=payments_paginated,
        search_query=search_query,
        branches=branches,
        selected_branch=selected_branch,
        start_date=start_date,
        end_date=end_date,
        page=page,
        pages=pages,
        total=total
    )

# ========================================
# EXPORT FUNCTIONS
# ========================================

def export_attendance_csv(data):
    """Export attendance data to CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Full Name', 'Gender', 'Christian Name', 'Phone', 'Section', 'Date', 'Status'])
    
    # Write data
    for row in data:
        writer.writerow(row)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=attendance_report.csv'
    return response

def attendance_report_pdf(data):
    """Export attendance report to PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A3)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Attendance Report</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Table data
    table_data = [['Full Name', 'Gender', 'Christian Name', 'Phone', 'Section', 'Date', 'Status']]
    for row in data:
        table_data.append(list(row))
    
    # Create table
    t = Table(table_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(t)
    doc.build(elements)
    
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, 
                     download_name='attendance_report.pdf')

def attendance_report_excel(data):
    """Export attendance report to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Write header
    headers = ['Full Name', 'Gender', 'Christian Name', 'Phone', 'Section', 'Date', 'Status']
    ws.append(headers)
    
    # Write data
    for row in data:
        ws.append(list(row))
    
    # Create response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='attendance_report.xlsx')

# ========================================
# USER MANAGEMENT ROUTES
# ========================================

@app.route('/user_management')
@login_required
@role_required('Super Admin')
def user_management():
    """Manage application users"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    cursor.execute("""
        SELECT u.user_id, u.payroll_number, u.username, u.email, u.full_name, 
               u.branch, u.is_active, u.last_login, r.role_name
        FROM aawsa_user u
        LEFT JOIN roles r ON u.role_id = r.role_id
        ORDER BY u.user_id DESC
    """)
    
    users = []
    for row in cursor.fetchall():
        users.append({
            'user_id': row[0],
            'payroll_number': row[1],
            'username': row[2],
            'email': row[3],
            'full_name': row[4],
            'branch': row[5],
            'is_active': row[6],
            'last_login': row[7],
            'role_name': row[8]
        })
    
    # Get roles for dropdown
    cursor.execute("SELECT role_id, role_name FROM roles ORDER BY role_name")
    roles = [{'role_id': r[0], 'role_name': r[1]} for r in cursor.fetchall()]
    
    cursor.close()
    conn.close()
    
    return render_template('user_management.html', users=users, roles=roles)

@app.route('/create_user', methods=['POST'])
@login_required
@role_required('Super Admin')
def create_user():
    """Create a new user"""
    payroll_number = request.form.get('payroll_number')
    username = request.form.get('username')
    email = request.form.get('email')
    full_name = request.form.get('full_name')
    password = request.form.get('password')
    role_id = request.form.get('role_id')
    branch = request.form.get('branch')
    
    if not all([payroll_number, username, password, role_id]):
        flash('Please fill in all required fields', 'danger')
        return redirect(url_for('user_management'))
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(buffered=True)
        
        password_hash = hash_password(password)
        
        cursor.execute("""
            INSERT INTO aawsa_user (payroll_number, username, email, full_name, password_hash, role_id, branch, is_active)
            VALUES (%(payroll_number)s, %(username)s, %(email)s, %(full_name)s, %(password_hash)s, %(role_id)s, %(branch)s, TRUE)
        """, {
            'payroll_number': payroll_number,
            'username': username,
            'email': email,
            'full_name': full_name,
            'password_hash': password_hash,
            'role_id': role_id,
            'branch': branch
        })
        
        conn.commit()
        flash(f'User {username} created successfully', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error creating user: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('user_management'))

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin')
def edit_user(user_id):
    """Edit existing user"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        full_name = request.form.get('full_name')
        role_id = request.form.get('role_id')
        branch = request.form.get('branch')
        is_active = request.form.get('is_active') == 'on'
        password = request.form.get('password')
        
        try:
            if password:
                password_hash = hash_password(password)
                cursor.execute("""
                    UPDATE aawsa_user 
                    SET username=%(username)s, email=%(email)s, full_name=%(full_name)s, 
                        role_id=%(role_id)s, branch=%(branch)s, is_active=%(is_active)s, password_hash=%(password_hash)s
                    WHERE user_id=%(user_id)s
                """, {
                    'username': username, 'email': email, 'full_name': full_name,
                    'role_id': role_id, 'branch': branch, 'is_active': is_active,
                    'password_hash': password_hash, 'user_id': user_id
                })
            else:
                cursor.execute("""
                    UPDATE aawsa_user 
                    SET username=%(username)s, email=%(email)s, full_name=%(full_name)s, 
                        role_id=%(role_id)s, branch=%(branch)s, is_active=%(is_active)s
                    WHERE user_id=%(user_id)s
                """, {
                    'username': username, 'email': email, 'full_name': full_name,
                    'role_id': role_id, 'branch': branch, 'is_active': is_active, 'user_id': user_id
                })
            
            conn.commit()
            flash('User updated successfully', 'success')
            return redirect(url_for('user_management'))
        except Exception as e:
            conn.rollback()
            flash(f'Error updating user: {str(e)}', 'danger')
    
    # GET request - load user data
    cursor.execute("""
        SELECT u.*, r.role_name
        FROM aawsa_user u
        LEFT JOIN roles r ON u.role_id = r.role_id
        WHERE u.user_id = %(user_id)s
    """, {'user_id': user_id})
    
    user_row = cursor.fetchone()
    if user_row:
        user = {
            'user_id': user_row[0],
            'payroll_number': user_row[1],
            'username': user_row[2],
            'email': user_row[3],
            'password_hash': user_row[4],
            'full_name': user_row[5],
            'branch': user_row[6],
            'last_login': user_row[7],
            'role_id': user_row[8],
            'is_active': user_row[11],
            'role_name': user_row[-1]
        }
    else:
        flash('User not found', 'danger')
        return redirect(url_for('user_management'))
    
    # Get roles for dropdown
    cursor.execute("SELECT role_id, role_name FROM roles ORDER BY role_name")
    roles = [{'role_id': r[0], 'role_name': r[1]} for r in cursor.fetchall()]
    
    cursor.close()
    conn.close()
    
    return render_template('edit_user.html', user=user, roles=roles)

@app.route('/delete_user/<int:user_id>', methods=['POST'])
@login_required
@role_required('Super Admin')
def delete_user(user_id):
    """Delete a user"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    try:
        cursor.execute("DELETE FROM aawsa_user WHERE user_id = %(user_id)s", {'user_id': user_id})
        conn.commit()
        flash('User deleted successfully', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting user: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('user_management'))

# ========================================
# ROLE MANAGEMENT ROUTES
# ========================================

@app.route('/manage_roles')
@login_required
@role_required('Super Admin')
def manage_roles():
    """Manage roles"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    cursor.execute("SELECT role_id, role_name, description FROM roles ORDER BY role_name")
    roles = []
    for row in cursor.fetchall():
        role_id = row[0]
        
        # Get assigned routes count for each role
        cursor.execute("""
            SELECT COUNT(*) FROM role_routes WHERE role_id = %s
        """, (role_id,))
        routes_count = cursor.fetchone()[0]
        
        roles.append({
            'role_id': role_id,
            'role_name': row[1],
            'description': row[2],
            'routes_count': routes_count
        })
    
    cursor.close()
    conn.close()
    
    return render_template('manage_roles.html', roles=roles)

@app.route('/add_role', methods=['POST'])
@login_required
@role_required('Super Admin')
def add_role():
    """Add a new role"""
    role_name = request.form.get('role_name')
    description = request.form.get('description')
    
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    try:
        cursor.execute(
            "INSERT INTO roles (role_name, description) VALUES (%(role_name)s, %(description)s)",
            {'role_name': role_name, 'description': description}
        )
        conn.commit()
        flash(f'Role "{role_name}" created successfully', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error creating role: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('manage_roles'))

@app.route('/delete_role/<int:role_id>', methods=['POST'])
@login_required
@role_required('Super Admin')
def delete_role(role_id):
    """Delete a role"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    try:
        cursor.execute("DELETE FROM roles WHERE role_id = %(role_id)s", {'role_id': role_id})
        conn.commit()
        flash('Role deleted successfully', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting role: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('manage_roles'))

@app.route('/roles/<int:role_id>/routes', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin')
def manage_role_routes(role_id):
    """Manage routes and CRUD permissions for a specific role"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        # Get all routes and their permissions from form
        selected_routes = request.form.getlist('routes')
        
        try:
            # Delete existing role-route mappings
            cursor.execute("DELETE FROM role_routes WHERE role_id = %s", (role_id,))
            
            # Insert new mappings with CRUD permissions
            for route_id in selected_routes:
                can_create = 1 if f'can_create_{route_id}' in request.form else 0
                can_read = 1 if f'can_read_{route_id}' in request.form else 0
                can_update = 1 if f'can_update_{route_id}' in request.form else 0
                can_delete = 1 if f'can_delete_{route_id}' in request.form else 0
                can_approve = 1 if f'can_approve_{route_id}' in request.form else 0
                can_export = 1 if f'can_export_{route_id}' in request.form else 0
                
                cursor.execute("""
                    INSERT INTO role_routes 
                    (role_id, route_id, can_create, can_read, can_update, can_delete, can_approve, can_export)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (role_id, route_id, can_create, can_read, can_update, can_delete, can_approve, can_export))
            
            conn.commit()
            flash('Role permissions updated successfully with CRUD controls', 'success')
            return redirect(url_for('manage_roles'))
        except Exception as e:
            conn.rollback()
            flash(f'Error updating permissions: {str(e)}', 'danger')
    
    # Get role info
    cursor.execute("SELECT role_name FROM roles WHERE role_id = %s", (role_id,))
    role = cursor.fetchone()
    if not role:
        flash('Role not found', 'danger')
        return redirect(url_for('manage_roles'))
    
    role_name = role[0]
    
    # Get all routes
    cursor.execute("SELECT route_id, route_name, description FROM routes ORDER BY route_name")
    all_routes = [{'route_id': r[0], 'route_name': r[1], 'description': r[2]} for r in cursor.fetchall()]
    
    # Get current role routes with CRUD permissions
    cursor.execute("""
        SELECT route_id, can_create, can_read, can_update, can_delete, can_approve, can_export
        FROM role_routes 
        WHERE role_id = %s
    """, (role_id,))
    
    route_permissions = {}
    for row in cursor.fetchall():
        route_permissions[row[0]] = {
            'can_create': bool(row[1]),
            'can_read': bool(row[2]),
            'can_update': bool(row[3]),
            'can_delete': bool(row[4]),
            'can_approve': bool(row[5]),
            'can_export': bool(row[6])
        }
    
    assigned_route_ids = list(route_permissions.keys())
    
    cursor.close()
    conn.close()
    
    return render_template('manage_role_routes.html', 
                         role_id=role_id, 
                         role_name=role_name, 
                         all_routes=all_routes, 
                         assigned_route_ids=assigned_route_ids,
                         route_permissions=route_permissions)

@app.route('/routes', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin')
def manage_routes():
    """Manage application routes/endpoints"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        route_name = request.form.get('route_name')
        endpoint = request.form.get('endpoint')
        description = request.form.get('description')
        
        try:
            cursor.execute("""
                INSERT INTO routes (route_name, endpoint, description)
                VALUES (%s, %s, %s)
            """, (route_name, endpoint, description))
            conn.commit()
            flash('Endpoint created successfully', 'success')
        except Exception as e:
            conn.rollback()
            flash(f'Error creating endpoint: {str(e)}', 'danger')
    
    # Get all routes with usage count
    cursor.execute("""
        SELECT r.route_id, r.route_name, r.endpoint, r.description,
               COUNT(rr.role_id) as usage_count
        FROM routes r
        LEFT JOIN role_routes rr ON r.route_id = rr.route_id
        GROUP BY r.route_id, r.route_name, r.endpoint, r.description
        ORDER BY r.route_name
    """)
    
    routes = []
    for row in cursor.fetchall():
        routes.append({
            'route_id': row[0],
            'route_name': row[1],
            'endpoint': row[2],
            'description': row[3],
            'usage_count': row[4]
        })
    
    cursor.close()
    conn.close()
    
    return render_template('manage_routes.html', routes=routes)

@app.route('/routes/delete/<int:route_id>', methods=['POST'])
@login_required
@role_required('Super Admin')
def delete_route(route_id):
    """Delete a route/endpoint"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    try:
        # Check if route is in use
        cursor.execute("SELECT COUNT(*) FROM role_routes WHERE route_id = %s", (route_id,))
        usage_count = cursor.fetchone()[0]
        
        if usage_count > 0:
            flash(f'Cannot delete endpoint: It is assigned to {usage_count} role(s). Remove assignments first.', 'warning')
        else:
            cursor.execute("DELETE FROM routes WHERE route_id = %s", (route_id,))
            conn.commit()
            flash('Endpoint deleted successfully', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting endpoint: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('manage_routes'))

# ========================================
# INVENTORY MANAGEMENT ROUTES
# ========================================

@app.route('/manage_inventory', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin', 'Inventory Manager')
def manage_inventory():
    """Manage inventory items - CRUD operations"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get search and filter parameters
    search = request.args.get('search', '')
    category_filter = request.args.get('category', '')
    location_filter = request.args.get('location', '')
    status_filter = request.args.get('status', 'Active')
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            try:
                data = {
                    'item_name': request.form.get('item_name'),
                    'category': request.form.get('category'),
                    'quantity': int(request.form.get('quantity', 0)),
                    'unit': request.form.get('unit'),
                    'location': request.form.get('location'),
                    'supplier': request.form.get('supplier'),
                    'purchase_date': request.form.get('purchase_date') or None,
                    'unit_price': float(request.form.get('unit_price', 0)) if request.form.get('unit_price') else None,
                    'min_stock_level': int(request.form.get('min_stock_level', 10)),
                    'description': request.form.get('description'),
                    'created_by': session.get('username', 'Unknown')
                }
                
                cursor.execute("""
                    INSERT INTO inventory_items (
                        item_name, category, quantity, unit, location, supplier,
                        purchase_date, unit_price, min_stock_level, description, created_by
                    ) VALUES (
                        %(item_name)s, %(category)s, %(quantity)s, %(unit)s, %(location)s, %(supplier)s,
                        %(purchase_date)s, %(unit_price)s, %(min_stock_level)s, %(description)s, %(created_by)s
                    )
                """, data)
                conn.commit()
                flash('Inventory item added successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error adding item: {str(e)}', 'danger')
        
        elif action == 'update':
            try:
                item_id = request.form.get('item_id')
                data = {
                    'id': item_id,
                    'item_name': request.form.get('item_name'),
                    'category': request.form.get('category'),
                    'unit': request.form.get('unit'),
                    'location': request.form.get('location'),
                    'supplier': request.form.get('supplier'),
                    'purchase_date': request.form.get('purchase_date') or None,
                    'unit_price': float(request.form.get('unit_price', 0)) if request.form.get('unit_price') else None,
                    'min_stock_level': int(request.form.get('min_stock_level', 10)),
                    'description': request.form.get('description'),
                    'status': request.form.get('status')
                }
                
                cursor.execute("""
                    UPDATE inventory_items SET
                        item_name = %(item_name)s,
                        category = %(category)s,
                        unit = %(unit)s,
                        location = %(location)s,
                        supplier = %(supplier)s,
                        purchase_date = %(purchase_date)s,
                        unit_price = %(unit_price)s,
                        min_stock_level = %(min_stock_level)s,
                        description = %(description)s,
                        status = %(status)s
                    WHERE id = %(id)s
                """, data)
                conn.commit()
                flash('Inventory item updated successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error updating item: {str(e)}', 'danger')
        
        elif action == 'delete':
            try:
                item_id = request.form.get('item_id')
                cursor.execute("DELETE FROM inventory_items WHERE id = %s", (item_id,))
                conn.commit()
                flash('Inventory item deleted successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error deleting item: {str(e)}', 'danger')
        
        return redirect(url_for('manage_inventory'))
    
    # Build query with filters
    query = """
        SELECT id, item_name, category, quantity, unit, location, supplier,
               purchase_date, unit_price, min_stock_level, description, status,
               created_at
        FROM inventory_items
        WHERE 1=1
    """
    params = []
    
    if search:
        query += " AND (item_name LIKE %s OR description LIKE %s OR supplier LIKE %s)"
        search_term = f'%{search}%'
        params.extend([search_term, search_term, search_term])
    
    if category_filter:
        query += " AND category = %s"
        params.append(category_filter)
    
    if location_filter:
        query += " AND location = %s"
        params.append(location_filter)
    
    if status_filter:
        query += " AND status = %s"
        params.append(status_filter)
    
    query += " ORDER BY item_name"
    
    cursor.execute(query, params)
    items = cursor.fetchall()
    
    # Get unique categories and locations for filters
    cursor.execute("SELECT DISTINCT category FROM inventory_items ORDER BY category")
    categories = [row[0] for row in cursor.fetchall()]
    
    cursor.execute("SELECT DISTINCT location FROM inventory_items WHERE location IS NOT NULL ORDER BY location")
    locations = [row[0] for row in cursor.fetchall()]
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_items,
            SUM(quantity) as total_quantity,
            COUNT(CASE WHEN quantity <= min_stock_level THEN 1 END) as low_stock_count,
            COUNT(CASE WHEN quantity = 0 THEN 1 END) as out_of_stock_count
        FROM inventory_items
        WHERE status = 'Active'
    """)
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('manage_inventory.html',
                         items=items,
                         categories=categories,
                         locations=locations,
                         stats=stats,
                         search=search,
                         category_filter=category_filter,
                         location_filter=location_filter,
                         status_filter=status_filter)

@app.route('/inventory_transactions', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin', 'Inventory Manager')
def inventory_transactions():
    """Record and view inventory transactions"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        try:
            item_id = int(request.form.get('item_id'))
            transaction_type = request.form.get('transaction_type')
            quantity = int(request.form.get('quantity'))
            transaction_date = request.form.get('transaction_date')
            reference_number = request.form.get('reference_number')
            responsible_user = request.form.get('responsible_user')
            recipient = request.form.get('recipient')
            purpose = request.form.get('purpose')
            notes = request.form.get('notes')
            
            # Get current quantity
            cursor.execute("SELECT quantity, item_name FROM inventory_items WHERE id = %s", (item_id,))
            result = cursor.fetchone()
            if not result:
                flash('Item not found', 'danger')
                return redirect(url_for('inventory_transactions'))
            
            current_qty = result[0]
            item_name = result[1]
            
            # Calculate new quantity based on transaction type
            if transaction_type == 'Incoming':
                new_qty = current_qty + quantity
            elif transaction_type == 'Outgoing':
                if quantity > current_qty:
                    flash(f'Insufficient stock! Available: {current_qty}, Requested: {quantity}', 'danger')
                    return redirect(url_for('inventory_transactions'))
                new_qty = current_qty - quantity
            elif transaction_type == 'Adjustment':
                # Adjustment can be positive or negative
                adjustment_type = request.form.get('adjustment_direction', 'increase')
                if adjustment_type == 'decrease':
                    quantity = -quantity
                new_qty = current_qty + quantity
                if new_qty < 0:
                    new_qty = 0
            else:
                flash('Invalid transaction type', 'danger')
                return redirect(url_for('inventory_transactions'))
            
            # Record transaction
            cursor.execute("""
                INSERT INTO inventory_transactions (
                    item_id, transaction_type, quantity, transaction_date,
                    reference_number, responsible_user, recipient, purpose, notes,
                    previous_quantity, new_quantity, created_by
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (item_id, transaction_type, abs(quantity), transaction_date,
                  reference_number, responsible_user, recipient, purpose, notes,
                  current_qty, new_qty, session.get('username', 'Unknown')))
            
            # Update item quantity
            cursor.execute("UPDATE inventory_items SET quantity = %s WHERE id = %s", (new_qty, item_id))
            
            conn.commit()
            flash(f'Transaction recorded: {item_name} quantity updated from {current_qty} to {new_qty}', 'success')
        except Exception as e:
            conn.rollback()
            flash(f'Error recording transaction: {str(e)}', 'danger')
        
        return redirect(url_for('inventory_transactions'))
    
    # Get all active inventory items for dropdown
    cursor.execute("""
        SELECT id, item_name, category, quantity, unit 
        FROM inventory_items 
        WHERE status = 'Active'
        ORDER BY item_name
    """)
    items = cursor.fetchall()
    
    # Get recent transactions
    cursor.execute("""
        SELECT t.id, i.item_name, t.transaction_type, t.quantity, t.transaction_date,
               t.reference_number, t.responsible_user, t.recipient, t.purpose,
               t.previous_quantity, t.new_quantity, t.created_at
        FROM inventory_transactions t
        JOIN inventory_items i ON t.item_id = i.id
        ORDER BY t.transaction_date DESC, t.created_at DESC
        LIMIT 50
    """)
    transactions = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_transactions,
            COUNT(CASE WHEN transaction_type = 'Incoming' THEN 1 END) as incoming_count,
            COUNT(CASE WHEN transaction_type = 'Outgoing' THEN 1 END) as outgoing_count,
            COUNT(CASE WHEN transaction_type = 'Adjustment' THEN 1 END) as adjustment_count
        FROM inventory_transactions
        WHERE transaction_date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
    """)
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('inventory_transactions.html',
                         items=items,
                         transactions=transactions,
                         stats=stats)

@app.route('/inventory_stock_report')
@login_required
def inventory_stock_report():
    """Inventory stock report with low stock alerts"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filters
    category_filter = request.args.get('category', '')
    location_filter = request.args.get('location', '')
    stock_alert = request.args.get('alert', '')  # low/out/all
    
    # Build query
    query = """
        SELECT id, item_name, category, quantity, unit, location, supplier,
               purchase_date, unit_price, min_stock_level, status,
               CASE 
                   WHEN quantity = 0 THEN 'Out of Stock'
                   WHEN quantity <= min_stock_level THEN 'Low Stock'
                   ELSE 'In Stock'
               END as stock_status
        FROM inventory_items
        WHERE 1=1
    """
    params = []
    
    if category_filter:
        query += " AND category = %s"
        params.append(category_filter)
    
    if location_filter:
        query += " AND location = %s"
        params.append(location_filter)
    
    if stock_alert == 'low':
        query += " AND quantity <= min_stock_level AND quantity > 0"
    elif stock_alert == 'out':
        query += " AND quantity = 0"
    
    query += " ORDER BY item_name"
    
    cursor.execute(query, params)
    items = cursor.fetchall()
    
    # Get categories and locations
    cursor.execute("SELECT DISTINCT category FROM inventory_items ORDER BY category")
    categories = [row[0] for row in cursor.fetchall()]
    
    cursor.execute("SELECT DISTINCT location FROM inventory_items WHERE location IS NOT NULL ORDER BY location")
    locations = [row[0] for row in cursor.fetchall()]
    
    # Get summary statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_items,
            SUM(quantity) as total_quantity,
            COUNT(CASE WHEN quantity <= min_stock_level AND quantity > 0 THEN 1 END) as low_stock,
            COUNT(CASE WHEN quantity = 0 THEN 1 END) as out_of_stock,
            SUM(CASE WHEN unit_price IS NOT NULL THEN quantity * unit_price ELSE 0 END) as total_value
        FROM inventory_items
        WHERE status = 'Active'
    """)
    summary = cursor.fetchone()
    
    # Get category-wise breakdown
    cursor.execute("""
        SELECT category, COUNT(*) as item_count, SUM(quantity) as total_qty
        FROM inventory_items
        WHERE status = 'Active'
        GROUP BY category
        ORDER BY category
    """)
    category_stats = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('inventory_stock_report.html',
                         items=items,
                         categories=categories,
                         locations=locations,
                         summary=summary,
                         category_stats=category_stats,
                         category_filter=category_filter,
                         location_filter=location_filter,
                         stock_alert=stock_alert)

@app.route('/inventory_movement_report')
@login_required
def inventory_movement_report():
    """Inventory movement/transaction history report"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filters
    item_filter = request.args.get('item', '')
    type_filter = request.args.get('type', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = """
        SELECT t.id, i.item_name, i.category, t.transaction_type, t.quantity,
               t.transaction_date, t.reference_number, t.responsible_user,
               t.recipient, t.purpose, t.previous_quantity, t.new_quantity,
               t.notes, t.created_at
        FROM inventory_transactions t
        JOIN inventory_items i ON t.item_id = i.id
        WHERE 1=1
    """
    params = []
    
    if item_filter:
        query += " AND t.item_id = %s"
        params.append(item_filter)
    
    if type_filter:
        query += " AND t.transaction_type = %s"
        params.append(type_filter)
    
    if date_from:
        query += " AND t.transaction_date >= %s"
        params.append(date_from)
    
    if date_to:
        query += " AND t.transaction_date <= %s"
        params.append(date_to)
    
    query += " ORDER BY t.transaction_date DESC, t.created_at DESC LIMIT 200"
    
    cursor.execute(query, params)
    transactions = cursor.fetchall()
    
    # Get items for filter dropdown
    cursor.execute("SELECT id, item_name FROM inventory_items ORDER BY item_name")
    items = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_transactions,
            SUM(CASE WHEN transaction_type = 'Incoming' THEN quantity ELSE 0 END) as total_incoming,
            SUM(CASE WHEN transaction_type = 'Outgoing' THEN quantity ELSE 0 END) as total_outgoing,
            SUM(CASE WHEN transaction_type = 'Adjustment' THEN quantity ELSE 0 END) as total_adjustments
        FROM inventory_transactions t
        WHERE 1=1
    """ + (" AND transaction_date >= %s" if date_from else "") + (" AND transaction_date <= %s" if date_to else ""))
    
    filter_params = []
    if date_from:
        filter_params.append(date_from)
    if date_to:
        filter_params.append(date_to)
    
    cursor.execute(query.replace("SELECT t.id", "SELECT COUNT(*)").split("ORDER BY")[0], params)
    total_filtered = cursor.fetchone()[0] if params else 0
    
    cursor.execute("""
        SELECT 
            COUNT(*) as total_transactions,
            SUM(CASE WHEN transaction_type = 'Incoming' THEN quantity ELSE 0 END) as total_incoming,
            SUM(CASE WHEN transaction_type = 'Outgoing' THEN quantity ELSE 0 END) as total_outgoing,
            SUM(CASE WHEN transaction_type = 'Adjustment' THEN ABS(quantity) ELSE 0 END) as total_adjustments
        FROM inventory_transactions
    """)
    summary = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('inventory_movement_report.html',
                         transactions=transactions,
                         items=items,
                         summary=summary,
                         item_filter=item_filter,
                         type_filter=type_filter,
                         date_from=date_from,
                         date_to=date_to)

# ========================================
# FIXED ASSET MANAGEMENT ROUTES
# ========================================

@app.route('/manage_fixed_assets', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin', 'Asset Manager')
def manage_fixed_assets():
    """Manage fixed assets register - CRUD operations"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get search and filters
    search = request.args.get('search', '')
    category_filter = request.args.get('category', '')
    department_filter = request.args.get('department', '')
    status_filter = request.args.get('status', 'Active')
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            try:
                purchase_cost = float(request.form.get('purchase_cost'))
                useful_life = int(request.form.get('useful_life_years', 5))
                salvage = float(request.form.get('salvage_value', 0))
                
                # Calculate initial book value
                book_value = purchase_cost
                
                data = {
                    'asset_name': request.form.get('asset_name'),
                    'category': request.form.get('category'),
                    'purchase_date': request.form.get('purchase_date'),
                    'purchase_cost': purchase_cost,
                    'current_location': request.form.get('current_location'),
                    'condition_status': request.form.get('condition_status', 'Good'),
                    'assigned_department': request.form.get('assigned_department'),
                    'assigned_user': request.form.get('assigned_user'),
                    'serial_number': request.form.get('serial_number'),
                    'useful_life_years': useful_life,
                    'depreciation_method': request.form.get('depreciation_method', 'Straight-Line'),
                    'salvage_value': salvage,
                    'book_value': book_value,
                    'description': request.form.get('description'),
                    'created_by': session.get('username', 'Unknown')
                }
                
                cursor.execute("""
                    INSERT INTO fixed_assets (
                        asset_name, category, purchase_date, purchase_cost, current_location,
                        condition_status, assigned_department, assigned_user, serial_number,
                        useful_life_years, depreciation_method, salvage_value, book_value,
                        description, created_by
                    ) VALUES (
                        %(asset_name)s, %(category)s, %(purchase_date)s, %(purchase_cost)s, %(current_location)s,
                        %(condition_status)s, %(assigned_department)s, %(assigned_user)s, %(serial_number)s,
                        %(useful_life_years)s, %(depreciation_method)s, %(salvage_value)s, %(book_value)s,
                        %(description)s, %(created_by)s
                    )
                """, data)
                conn.commit()
                flash('Fixed asset added successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error adding asset: {str(e)}', 'danger')
        
        elif action == 'update':
            try:
                asset_id = request.form.get('asset_id')
                data = {
                    'id': asset_id,
                    'asset_name': request.form.get('asset_name'),
                    'category': request.form.get('category'),
                    'purchase_date': request.form.get('purchase_date'),
                    'purchase_cost': float(request.form.get('purchase_cost')),
                    'current_location': request.form.get('current_location'),
                    'condition_status': request.form.get('condition_status'),
                    'assigned_department': request.form.get('assigned_department'),
                    'assigned_user': request.form.get('assigned_user'),
                    'serial_number': request.form.get('serial_number'),
                    'useful_life_years': int(request.form.get('useful_life_years')),
                    'depreciation_method': request.form.get('depreciation_method'),
                    'salvage_value': float(request.form.get('salvage_value', 0)),
                    'description': request.form.get('description'),
                    'status': request.form.get('status')
                }
                
                cursor.execute("""
                    UPDATE fixed_assets SET
                        asset_name = %(asset_name)s,
                        category = %(category)s,
                        purchase_date = %(purchase_date)s,
                        purchase_cost = %(purchase_cost)s,
                        current_location = %(current_location)s,
                        condition_status = %(condition_status)s,
                        assigned_department = %(assigned_department)s,
                        assigned_user = %(assigned_user)s,
                        serial_number = %(serial_number)s,
                        useful_life_years = %(useful_life_years)s,
                        depreciation_method = %(depreciation_method)s,
                        salvage_value = %(salvage_value)s,
                        description = %(description)s,
                        status = %(status)s
                    WHERE id = %(id)s
                """, data)
                conn.commit()
                flash('Fixed asset updated successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error updating asset: {str(e)}', 'danger')
        
        elif action == 'delete':
            try:
                asset_id = request.form.get('asset_id')
                cursor.execute("DELETE FROM fixed_assets WHERE id = %s", (asset_id,))
                conn.commit()
                flash('Fixed asset deleted successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error deleting asset: {str(e)}', 'danger')
        
        return redirect(url_for('manage_fixed_assets'))
    
    # Build query with filters
    query = """
        SELECT id, asset_name, category, purchase_date, purchase_cost, current_location,
               condition_status, assigned_department, assigned_user, serial_number,
               useful_life_years, depreciation_method, book_value, description, status,
               created_at
        FROM fixed_assets
        WHERE 1=1
    """
    params = []
    
    if search:
        query += " AND (asset_name LIKE %s OR description LIKE %s OR serial_number LIKE %s)"
        search_term = f'%{search}%'
        params.extend([search_term, search_term, search_term])
    
    if category_filter:
        query += " AND category = %s"
        params.append(category_filter)
    
    if department_filter:
        query += " AND assigned_department = %s"
        params.append(department_filter)
    
    if status_filter:
        query += " AND status = %s"
        params.append(status_filter)
    
    query += " ORDER BY asset_name"
    
    cursor.execute(query, params)
    assets = cursor.fetchall()
    
    # Get unique values for filters
    cursor.execute("SELECT DISTINCT category FROM fixed_assets ORDER BY category")
    categories = [row[0] for row in cursor.fetchall()]
    
    cursor.execute("SELECT DISTINCT assigned_department FROM fixed_assets WHERE assigned_department IS NOT NULL ORDER BY assigned_department")
    departments = [row[0] for row in cursor.fetchall()]
    
    # Get all members for assignment dropdown
    cursor.execute("SELECT id, full_name, section_name, phone FROM member_registration ORDER BY full_name")
    members = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_assets,
            SUM(purchase_cost) as total_cost,
            SUM(book_value) as total_book_value,
            COUNT(CASE WHEN status = 'Active' THEN 1 END) as active_count,
            COUNT(CASE WHEN condition_status = 'Poor' THEN 1 END) as poor_condition_count
        FROM fixed_assets
    """)
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('manage_fixed_assets.html',
                         assets=assets,
                         categories=categories,
                         departments=departments,
                         members=members,
                         stats=stats,
                         search=search,
                         category_filter=category_filter,
                         department_filter=department_filter,
                         status_filter=status_filter)

@app.route('/asset_movements', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin', 'Asset Manager')
def asset_movements():
    """Track asset movements, assignments, and maintenance"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        try:
            asset_id = int(request.form.get('asset_id'))
            movement_type = request.form.get('movement_type')
            movement_date = request.form.get('movement_date')
            
            # Get current asset info
            cursor.execute("""
                SELECT current_location, assigned_department, assigned_user, condition_status
                FROM fixed_assets WHERE id = %s
            """, (asset_id,))
            current = cursor.fetchone()
            
            if not current:
                flash('Asset not found', 'danger')
                return redirect(url_for('asset_movements'))
            
            from_location, from_dept, from_user, condition_before = current
            
            # Get new values from form
            to_location = request.form.get('to_location') or from_location
            to_department = request.form.get('to_department') or from_dept
            to_user = request.form.get('to_user') or from_user
            condition_after = request.form.get('condition_after') or condition_before
            repair_cost = float(request.form.get('repair_cost', 0)) if request.form.get('repair_cost') else None
            
            # Record movement
            cursor.execute("""
                INSERT INTO asset_movements (
                    asset_id, movement_type, movement_date, from_location, to_location,
                    from_department, to_department, from_user, to_user,
                    responsible_person, condition_before, condition_after,
                    repair_cost, remarks, created_by
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (asset_id, movement_type, movement_date, from_location, to_location,
                  from_dept, to_department, from_user, to_user,
                  request.form.get('responsible_person'),
                  condition_before, condition_after, repair_cost,
                  request.form.get('remarks'),
                  session.get('username', 'Unknown')))
            
            # Update asset record
            update_data = {
                'id': asset_id,
                'current_location': to_location,
                'assigned_department': to_department,
                'assigned_user': to_user,
                'condition_status': condition_after
            }
            
            # If disposal, update status
            if movement_type == 'Disposal':
                update_data['status'] = 'Disposed'
                update_data['disposal_date'] = movement_date
                update_data['disposal_method'] = request.form.get('disposal_method')
                update_data['disposal_value'] = float(request.form.get('disposal_value', 0)) if request.form.get('disposal_value') else None
                
                cursor.execute("""
                    UPDATE fixed_assets SET
                        current_location = %(current_location)s,
                        assigned_department = %(assigned_department)s,
                        assigned_user = %(assigned_user)s,
                        condition_status = %(condition_status)s,
                        status = %(status)s,
                        disposal_date = %(disposal_date)s,
                        disposal_method = %(disposal_method)s,
                        disposal_value = %(disposal_value)s
                    WHERE id = %(id)s
                """, update_data)
            elif movement_type == 'Repair':
                update_data['status'] = 'Under Repair'
                cursor.execute("""
                    UPDATE fixed_assets SET
                        current_location = %(current_location)s,
                        condition_status = %(condition_status)s,
                        status = %(status)s
                    WHERE id = %(id)s
                """, update_data)
            else:
                cursor.execute("""
                    UPDATE fixed_assets SET
                        current_location = %(current_location)s,
                        assigned_department = %(assigned_department)s,
                        assigned_user = %(assigned_user)s,
                        condition_status = %(condition_status)s
                    WHERE id = %(id)s
                """, update_data)
            
            conn.commit()
            flash(f'Asset movement recorded successfully ({movement_type})', 'success')
        except Exception as e:
            conn.rollback()
            flash(f'Error recording movement: {str(e)}', 'danger')
        
        return redirect(url_for('asset_movements'))
    
    # Get all active assets
    cursor.execute("""
        SELECT id, asset_name, category, current_location, assigned_department,
               assigned_user, serial_number, condition_status
        FROM fixed_assets
        WHERE status != 'Disposed'
        ORDER BY asset_name
    """)
    assets = cursor.fetchall()
    
    # Get recent movements
    cursor.execute("""
        SELECT m.id, a.asset_name, m.movement_type, m.movement_date,
               m.from_location, m.to_location, m.from_department, m.to_department,
               m.responsible_person, m.condition_before, m.condition_after,
               m.repair_cost, m.remarks, m.created_at
        FROM asset_movements m
        JOIN fixed_assets a ON m.asset_id = a.id
        ORDER BY m.movement_date DESC, m.created_at DESC
        LIMIT 50
    """)
    movements = cursor.fetchall()
    
    # Get unique departments
    cursor.execute("SELECT DISTINCT assigned_department FROM fixed_assets WHERE assigned_department IS NOT NULL")
    departments = [row[0] for row in cursor.fetchall()]
    
    # Get all members for assignment dropdown
    cursor.execute("SELECT id, full_name, section_name, phone FROM member_registration ORDER BY full_name")
    members = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('asset_movements.html',
                         assets=assets,
                         movements=movements,
                         departments=departments,
                         members=members)

@app.route('/asset_register_report')
@login_required
def asset_register_report():
    """Asset register report with current status"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filters
    category_filter = request.args.get('category', '')
    department_filter = request.args.get('department', '')
    condition_filter = request.args.get('condition', '')
    
    # Build query
    query = """
        SELECT id, asset_name, category, purchase_date, purchase_cost,
               current_location, condition_status, assigned_department,
               assigned_user, serial_number, book_value, status,
               YEAR(CURDATE()) - YEAR(purchase_date) as age_years
        FROM fixed_assets
        WHERE 1=1
    """
    params = []
    
    if category_filter:
        query += " AND category = %s"
        params.append(category_filter)
    
    if department_filter:
        query += " AND assigned_department = %s"
        params.append(department_filter)
    
    if condition_filter:
        query += " AND condition_status = %s"
        params.append(condition_filter)
    
    query += " ORDER BY asset_name"
    
    cursor.execute(query, params)
    assets = cursor.fetchall()
    
    # Get filter options
    cursor.execute("SELECT DISTINCT category FROM fixed_assets ORDER BY category")
    categories = [row[0] for row in cursor.fetchall()]
    
    cursor.execute("SELECT DISTINCT assigned_department FROM fixed_assets WHERE assigned_department IS NOT NULL ORDER BY assigned_department")
    departments = [row[0] for row in cursor.fetchall()]
    
    # Get summary
    cursor.execute("""
        SELECT 
            COUNT(*) as total_assets,
            SUM(purchase_cost) as total_purchase_cost,
            SUM(book_value) as total_book_value,
            SUM(accumulated_depreciation) as total_depreciation,
            COUNT(CASE WHEN status = 'Active' THEN 1 END) as active_count,
            COUNT(CASE WHEN status = 'Disposed' THEN 1 END) as disposed_count
        FROM fixed_assets
    """)
    summary = cursor.fetchone()
    
    # Category breakdown
    cursor.execute("""
        SELECT category, COUNT(*) as count, SUM(purchase_cost) as total_cost, SUM(book_value) as total_value
        FROM fixed_assets
        WHERE status = 'Active'
        GROUP BY category
        ORDER BY category
    """)
    category_stats = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('asset_register_report.html',
                         assets=assets,
                         categories=categories,
                         departments=departments,
                         summary=summary,
                         category_stats=category_stats,
                         category_filter=category_filter,
                         department_filter=department_filter,
                         condition_filter=condition_filter)

@app.route('/asset_movement_report')
@login_required
def asset_movement_report():
    """Asset movement history report"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filters
    asset_filter = request.args.get('asset', '')
    type_filter = request.args.get('type', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = """
        SELECT m.id, a.asset_name, a.category, m.movement_type, m.movement_date,
               m.from_location, m.to_location, m.from_department, m.to_department,
               m.from_user, m.to_user, m.responsible_person,
               m.condition_before, m.condition_after, m.repair_cost, m.remarks
        FROM asset_movements m
        JOIN fixed_assets a ON m.asset_id = a.id
        WHERE 1=1
    """
    params = []
    
    if asset_filter:
        query += " AND m.asset_id = %s"
        params.append(asset_filter)
    
    if type_filter:
        query += " AND m.movement_type = %s"
        params.append(type_filter)
    
    if date_from:
        query += " AND m.movement_date >= %s"
        params.append(date_from)
    
    if date_to:
        query += " AND m.movement_date <= %s"
        params.append(date_to)
    
    query += " ORDER BY m.movement_date DESC LIMIT 200"
    
    cursor.execute(query, params)
    movements = cursor.fetchall()
    
    # Get assets for filter
    cursor.execute("SELECT id, asset_name FROM fixed_assets ORDER BY asset_name")
    assets = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_movements,
            COUNT(CASE WHEN movement_type = 'Assignment' THEN 1 END) as assignments,
            COUNT(CASE WHEN movement_type = 'Relocation' THEN 1 END) as relocations,
            COUNT(CASE WHEN movement_type = 'Repair' THEN 1 END) as repairs,
            COUNT(CASE WHEN movement_type = 'Disposal' THEN 1 END) as disposals
        FROM asset_movements
    """)
    summary = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('asset_movement_report.html',
                         movements=movements,
                         assets=assets,
                         summary=summary,
                         asset_filter=asset_filter,
                         type_filter=type_filter,
                         date_from=date_from,
                         date_to=date_to)

@app.route('/asset_depreciation_report')
@login_required
def asset_depreciation_report():
    """Asset depreciation analysis report"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get all active assets with depreciation calculation
    cursor.execute("""
        SELECT 
            id, asset_name, category, purchase_date, purchase_cost,
            useful_life_years, depreciation_method, salvage_value,
            book_value, accumulated_depreciation,
            YEAR(CURDATE()) - YEAR(purchase_date) as age_years,
            DATEDIFF(CURDATE(), purchase_date) / 365.25 as age_decimal
        FROM fixed_assets
        WHERE status = 'Active'
        ORDER BY category, asset_name
    """)
    assets = cursor.fetchall()
    
    # Calculate depreciation for each asset
    depreciation_data = []
    for asset in assets:
        asset_id, name, category, purchase_date, cost, useful_life, method, salvage, book_val, accum_dep, age_years, age_decimal = asset
        
        # Straight-line depreciation calculation
        if method == 'Straight-Line' and useful_life > 0:
            annual_depreciation = (cost - salvage) / useful_life
            calculated_accum_dep = min(annual_depreciation * age_decimal, cost - salvage)
            calculated_book_value = cost - calculated_accum_dep
        else:
            annual_depreciation = 0
            calculated_accum_dep = accum_dep or 0
            calculated_book_value = book_val or cost
        
        depreciation_data.append({
            'id': asset_id,
            'name': name,
            'category': category,
            'purchase_date': purchase_date,
            'purchase_cost': cost,
            'useful_life': useful_life,
            'age_years': age_years,
            'annual_depreciation': annual_depreciation,
            'accumulated_depreciation': calculated_accum_dep,
            'book_value': calculated_book_value,
            'salvage_value': salvage
        })
    
    # Summary statistics
    cursor.execute("""
        SELECT 
            SUM(purchase_cost) as total_cost,
            SUM(book_value) as total_book_value,
            SUM(accumulated_depreciation) as total_depreciation,
            AVG(YEAR(CURDATE()) - YEAR(purchase_date)) as avg_age
        FROM fixed_assets
        WHERE status = 'Active'
    """)
    summary = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('asset_depreciation_report.html',
                         depreciation_data=depreciation_data,
                         summary=summary)

# ========================================
# DEPARTMENT & POSITION MANAGEMENT ROUTES
# ========================================

@app.route('/manage_departments', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin')
def manage_departments():
    """Manage organizational departments"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            try:
                data = {
                    'department_name': request.form.get('department_name'),
                    'department_code': request.form.get('department_code'),
                    'parent_department_id': int(request.form.get('parent_department_id')) if request.form.get('parent_department_id') else None,
                    'head_member_id': int(request.form.get('head_member_id')) if request.form.get('head_member_id') else None,
                    'description': request.form.get('description'),
                    'created_by': session.get('username', 'Unknown')
                }
                
                cursor.execute("""
                    INSERT INTO departments (department_name, department_code, parent_department_id, 
                                           head_member_id, description, created_by)
                    VALUES (%(department_name)s, %(department_code)s, %(parent_department_id)s, 
                           %(head_member_id)s, %(description)s, %(created_by)s)
                """, data)
                
                department_id = cursor.lastrowid
                
                # If a head is assigned, create history record
                if data['head_member_id']:
                    cursor.execute("""
                        INSERT INTO department_head_history (department_id, member_id, start_date, 
                                                            appointment_reason, created_by)
                        VALUES (%s, %s, CURDATE(), %s, %s)
                    """, (department_id, data['head_member_id'], 
                          'Initial appointment', session.get('username', 'Unknown')))
                
                conn.commit()
                flash('Department created successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error creating department: {str(e)}', 'danger')
        
        elif action == 'update':
            try:
                data = {
                    'id': request.form.get('department_id'),
                    'department_name': request.form.get('department_name'),
                    'department_code': request.form.get('department_code'),
                    'parent_department_id': int(request.form.get('parent_department_id')) if request.form.get('parent_department_id') else None,
                    'head_member_id': int(request.form.get('head_member_id')) if request.form.get('head_member_id') else None,
                    'description': request.form.get('description'),
                    'status': request.form.get('status')
                }
                
                # Get current head to check if it's changing
                cursor.execute("SELECT head_member_id FROM departments WHERE id = %s", (data['id'],))
                current_head = cursor.fetchone()
                old_head_id = current_head[0] if current_head else None
                
                # Check if department head is changing
                if old_head_id != data['head_member_id']:
                    # End previous head's tenure if exists
                    if old_head_id:
                        cursor.execute("""
                            UPDATE department_head_history 
                            SET end_date = CURDATE(), is_current = 0,
                                termination_reason = %s
                            WHERE department_id = %s AND member_id = %s AND is_current = 1
                        """, (request.form.get('change_reason', 'Department head changed'), 
                              data['id'], old_head_id))
                    
                    # Create new head history record if new head is assigned
                    if data['head_member_id']:
                        cursor.execute("""
                            INSERT INTO department_head_history (department_id, member_id, start_date, 
                                                                appointment_reason, created_by)
                            VALUES (%s, %s, CURDATE(), %s, %s)
                        """, (data['id'], data['head_member_id'], 
                              request.form.get('appointment_reason', 'New appointment'), 
                              session.get('username', 'Unknown')))
                
                cursor.execute("""
                    UPDATE departments SET
                        department_name = %(department_name)s,
                        department_code = %(department_code)s,
                        parent_department_id = %(parent_department_id)s,
                        head_member_id = %(head_member_id)s,
                        description = %(description)s,
                        status = %(status)s
                    WHERE id = %(id)s
                """, data)
                conn.commit()
                flash('Department updated successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error updating department: {str(e)}', 'danger')
        
        elif action == 'delete':
            try:
                dept_id = request.form.get('department_id')
                cursor.execute("DELETE FROM departments WHERE id = %s", (dept_id,))
                conn.commit()
                flash('Department deleted successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error deleting department: {str(e)}', 'danger')
        
        return redirect(url_for('manage_departments'))
    
    # Get all departments with head member info
    cursor.execute("""
        SELECT d.id, d.department_name, d.department_code, d.parent_department_id,
               pd.department_name as parent_name, d.head_member_id, m.full_name as head_name,
               d.description, d.status, d.created_at
        FROM departments d
        LEFT JOIN departments pd ON d.parent_department_id = pd.id
        LEFT JOIN member_registration m ON d.head_member_id = m.id
        ORDER BY d.department_name
    """)
    departments = cursor.fetchall()
    
    # Get all members for head selection
    cursor.execute("SELECT id, full_name, section_name FROM member_registration ORDER BY full_name")
    members = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_departments,
            COUNT(CASE WHEN status = 'Active' THEN 1 END) as active_count,
            COUNT(CASE WHEN head_member_id IS NOT NULL THEN 1 END) as with_head_count
        FROM departments
    """)
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('manage_departments.html',
                         departments=departments,
                         members=members,
                         stats=stats)

@app.route('/manage_positions', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin')
def manage_positions():
    """Manage job positions with dynamic creation features"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            try:
                data = {
                    'position_title': request.form.get('position_title'),
                    'department_id': int(request.form.get('department_id')) if request.form.get('department_id') else None,
                    'position_level': request.form.get('position_level'),
                    'responsibilities': request.form.get('responsibilities'),
                    'reporting_to': int(request.form.get('reporting_to')) if request.form.get('reporting_to') else None,
                    'position_type': request.form.get('position_type', 'Regular'),
                    'is_leadership': 1 if request.form.get('is_leadership') == 'on' else 0,
                    'max_holders': int(request.form.get('max_holders', 1)),
                    'min_experience_years': int(request.form.get('min_experience_years', 0)),
                    'required_skills': request.form.get('required_skills'),
                    'salary_range': request.form.get('salary_range'),
                    'created_by': session.get('username', 'Unknown')
                }
                
                cursor.execute("""
                    INSERT INTO positions (position_title, department_id, position_level, 
                                         responsibilities, reporting_to, position_type, 
                                         is_leadership, max_holders, min_experience_years, 
                                         required_skills, salary_range, created_by)
                    VALUES (%(position_title)s, %(department_id)s, %(position_level)s, 
                           %(responsibilities)s, %(reporting_to)s, %(position_type)s, 
                           %(is_leadership)s, %(max_holders)s, %(min_experience_years)s, 
                           %(required_skills)s, %(salary_range)s, %(created_by)s)
                """, data)
                conn.commit()
                flash('Position created successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error creating position: {str(e)}', 'danger')
        
        elif action == 'create_from_template':
            try:
                template_id = request.form.get('template_id')
                department_id = int(request.form.get('department_id')) if request.form.get('department_id') else None
                
                # Get template data
                cursor.execute("""
                    SELECT template_name, position_title, position_level, responsibilities, 
                           position_type, is_leadership, max_holders, min_experience_years, 
                           required_skills, description
                    FROM position_templates WHERE id = %s
                """, (template_id,))
                template = cursor.fetchone()
                
                if template:
                    # Create position from template
                    cursor.execute("""
                        INSERT INTO positions (position_title, department_id, position_level, 
                                             responsibilities, position_type, is_leadership, 
                                             max_holders, min_experience_years, required_skills, 
                                             created_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (template[1], department_id, template[2], template[3], template[4], 
                          template[5], template[6], template[7], template[8], 
                          session.get('username', 'Unknown')))
                    
                    # Update template usage count
                    cursor.execute("""
                        UPDATE position_templates SET usage_count = usage_count + 1 
                        WHERE id = %s
                    """, (template_id,))
                    
                    conn.commit()
                    flash(f'Position "{template[1]}" created from template successfully', 'success')
                else:
                    flash('Template not found', 'danger')
            except Exception as e:
                conn.rollback()
                flash(f'Error creating position from template: {str(e)}', 'danger')
        
        elif action == 'bulk_create':
            try:
                department_id = int(request.form.get('department_id')) if request.form.get('department_id') else None
                template_ids = request.form.getlist('template_ids')
                
                created_count = 0
                for template_id in template_ids:
                    cursor.execute("""
                        SELECT template_name, position_title, position_level, responsibilities, 
                               position_type, is_leadership, max_holders, min_experience_years, 
                               required_skills, description
                        FROM position_templates WHERE id = %s
                    """, (template_id,))
                    template = cursor.fetchone()
                    
                    if template:
                        cursor.execute("""
                            INSERT INTO positions (position_title, department_id, position_level, 
                                                 responsibilities, position_type, is_leadership, 
                                                 max_holders, min_experience_years, required_skills, 
                                                 created_by)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (template[1], department_id, template[2], template[3], template[4], 
                              template[5], template[6], template[7], template[8], 
                              session.get('username', 'Unknown')))
                        
                        cursor.execute("""
                            UPDATE position_templates SET usage_count = usage_count + 1 
                            WHERE id = %s
                        """, (template_id,))
                        created_count += 1
                
                conn.commit()
                flash(f'{created_count} positions created from templates successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error creating bulk positions: {str(e)}', 'danger')
        
        elif action == 'clone':
            try:
                source_position_id = request.form.get('source_position_id')
                new_title = request.form.get('new_title')
                department_id = int(request.form.get('department_id')) if request.form.get('department_id') else None
                
                # Get source position data
                cursor.execute("""
                    SELECT position_title, position_level, responsibilities, position_type, 
                           is_leadership, max_holders, min_experience_years, required_skills, 
                           salary_range
                    FROM positions WHERE id = %s
                """, (source_position_id,))
                source = cursor.fetchone()
                
                if source:
                    cursor.execute("""
                        INSERT INTO positions (position_title, department_id, position_level, 
                                             responsibilities, position_type, is_leadership, 
                                             max_holders, min_experience_years, required_skills, 
                                             salary_range, created_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (new_title, department_id, source[1], source[2], source[3], 
                          source[4], source[5], source[6], source[7], source[8], 
                          session.get('username', 'Unknown')))
                    conn.commit()
                    flash(f'Position "{new_title}" cloned successfully', 'success')
                else:
                    flash('Source position not found', 'danger')
            except Exception as e:
                conn.rollback()
                flash(f'Error cloning position: {str(e)}', 'danger')
        
        elif action == 'update':
            try:
                data = {
                    'id': request.form.get('position_id'),
                    'position_title': request.form.get('position_title'),
                    'department_id': int(request.form.get('department_id')) if request.form.get('department_id') else None,
                    'position_level': request.form.get('position_level'),
                    'responsibilities': request.form.get('responsibilities'),
                    'reporting_to': int(request.form.get('reporting_to')) if request.form.get('reporting_to') else None,
                    'position_type': request.form.get('position_type'),
                    'is_leadership': 1 if request.form.get('is_leadership') == 'on' else 0,
                    'max_holders': int(request.form.get('max_holders', 1)),
                    'min_experience_years': int(request.form.get('min_experience_years', 0)),
                    'required_skills': request.form.get('required_skills'),
                    'salary_range': request.form.get('salary_range'),
                    'status': request.form.get('status')
                }
                
                cursor.execute("""
                    UPDATE positions SET
                        position_title = %(position_title)s,
                        department_id = %(department_id)s,
                        position_level = %(position_level)s,
                        responsibilities = %(responsibilities)s,
                        reporting_to = %(reporting_to)s,
                        position_type = %(position_type)s,
                        is_leadership = %(is_leadership)s,
                        max_holders = %(max_holders)s,
                        min_experience_years = %(min_experience_years)s,
                        required_skills = %(required_skills)s,
                        salary_range = %(salary_range)s,
                        status = %(status)s
                    WHERE id = %(id)s
                """, data)
                conn.commit()
                flash('Position updated successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error updating position: {str(e)}', 'danger')
        
        elif action == 'delete':
            try:
                position_id = request.form.get('position_id')
                
                # Check if position has active assignments
                cursor.execute("""
                    SELECT COUNT(*) FROM member_positions 
                    WHERE position_id = %s AND is_current = 1
                """, (position_id,))
                active_assignments = cursor.fetchone()[0]
                
                if active_assignments > 0:
                    flash(f'Cannot delete position. It has {active_assignments} active member assignment(s).', 'danger')
                else:
                    cursor.execute("DELETE FROM positions WHERE id = %s", (position_id,))
                    conn.commit()
                    flash('Position deleted successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error deleting position: {str(e)}', 'danger')
        
        return redirect(url_for('manage_positions'))
    
    # Get all positions with department names
    cursor.execute("""
        SELECT p.id, p.position_title, p.department_id, d.department_name, 
               p.position_level, p.responsibilities, p.reporting_to, 
               p.position_type, p.is_leadership, p.max_holders, 
               p.min_experience_years, p.required_skills, p.salary_range, 
               p.status, p.created_at,
               (SELECT COUNT(*) FROM member_positions mp WHERE mp.position_id = p.id AND mp.is_current = 1) as current_holders,
               (SELECT pt.position_title FROM positions pt WHERE pt.id = p.reporting_to) as reports_to_title
        FROM positions p
        LEFT JOIN departments d ON p.department_id = d.id
        ORDER BY p.position_level DESC, p.position_title
    """)
    positions = cursor.fetchall()
    
    # Get departments for dropdown
    cursor.execute("SELECT id, department_name FROM departments WHERE status = 'Active' ORDER BY department_name")
    departments = cursor.fetchall()
    
    # Get position templates
    cursor.execute("""
        SELECT id, template_name, category, position_title, position_level, 
               responsibilities, position_type, is_leadership, max_holders, 
               min_experience_years, required_skills, description, usage_count
        FROM position_templates 
        WHERE is_active = 1
        ORDER BY category, position_title
    """)
    templates = cursor.fetchall()
    
    # Get positions for reporting hierarchy dropdown
    cursor.execute("""
        SELECT p.id, p.position_title, p.department_id, d.department_name
        FROM positions p
        LEFT JOIN departments d ON p.department_id = d.id
        WHERE p.status = 'Active'
        ORDER BY p.position_level DESC, p.position_title
    """)
    reporting_positions = cursor.fetchall()
    
    # Calculate statistics
    cursor.execute("SELECT COUNT(*) FROM positions")
    total_positions = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM positions WHERE status = 'Active'")
    active_positions = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(DISTINCT department_id) FROM positions WHERE department_id IS NOT NULL")
    departments_with_positions = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM positions WHERE is_leadership = 1")
    leadership_positions = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM position_templates WHERE is_active = 1")
    available_templates = cursor.fetchone()[0]
    
    stats = [total_positions, active_positions, departments_with_positions, leadership_positions, available_templates]
    
    cursor.close()
    conn.close()
    
    return render_template('manage_positions.html',
                         positions=positions,
                         departments=departments,
                         templates=templates,
                         reporting_positions=reporting_positions,
                         stats=stats)

@app.route('/assign_member_positions', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin')
def assign_member_positions():
    """Assign members to positions in departments"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'assign':
            try:
                member_id = int(request.form.get('member_id'))
                position_id = int(request.form.get('position_id'))
                department_id = int(request.form.get('department_id'))
                start_date = request.form.get('start_date')
                appointment_type = request.form.get('appointment_type')
                notes = request.form.get('notes')
                
                # Set previous assignments for this member to not current
                cursor.execute("""
                    UPDATE member_positions SET is_current = 0, end_date = %s
                    WHERE member_id = %s AND is_current = 1
                """, (start_date, member_id))
                
                # Create new assignment
                cursor.execute("""
                    INSERT INTO member_positions (member_id, position_id, department_id, 
                                                start_date, appointment_type, notes, created_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (member_id, position_id, department_id, start_date, appointment_type, notes,
                      session.get('username', 'Unknown')))
                
                conn.commit()
                flash('Member assigned to position successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error assigning position: {str(e)}', 'danger')
        
        elif action == 'end_assignment':
            try:
                assignment_id = request.form.get('assignment_id')
                end_date = request.form.get('end_date')
                
                cursor.execute("""
                    UPDATE member_positions SET is_current = 0, end_date = %s
                    WHERE id = %s
                """, (end_date, assignment_id))
                conn.commit()
                flash('Assignment ended successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error ending assignment: {str(e)}', 'danger')
        
        return redirect(url_for('assign_member_positions'))
    
    # Get search filter
    search = request.args.get('search', '')
    department_filter = request.args.get('department', '')
    
    # Get current assignments
    query = """
        SELECT mp.id, m.id as member_id, m.full_name, m.section_name,
               p.position_title, d.department_name, mp.start_date, mp.end_date,
               mp.appointment_type, mp.is_current
        FROM member_positions mp
        JOIN member_registration m ON mp.member_id = m.id
        JOIN positions p ON mp.position_id = p.id
        JOIN departments d ON mp.department_id = d.id
        WHERE mp.is_current = 1
    """
    params = []
    
    if search:
        query += " AND m.full_name LIKE %s"
        params.append(f'%{search}%')
    
    if department_filter:
        query += " AND mp.department_id = %s"
        params.append(department_filter)
    
    query += " ORDER BY d.department_name, p.position_title, m.full_name"
    
    cursor.execute(query, params)
    current_assignments = cursor.fetchall()
    
    # Get members without current positions
    cursor.execute("""
        SELECT m.id, m.full_name, m.section_name, m.phone
        FROM member_registration m
        LEFT JOIN member_positions mp ON m.id = mp.member_id AND mp.is_current = 1
        WHERE mp.id IS NULL
        ORDER BY m.full_name
    """)
    unassigned_members = cursor.fetchall()
    
    # Get all members, positions, departments for dropdowns
    cursor.execute("SELECT id, full_name, section_name FROM member_registration ORDER BY full_name")
    all_members = cursor.fetchall()
    
    cursor.execute("SELECT id, position_title, department_id FROM positions WHERE status = 'Active' ORDER BY position_title")
    positions = cursor.fetchall()
    
    cursor.execute("SELECT id, department_name FROM departments WHERE status = 'Active' ORDER BY department_name")
    departments = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(DISTINCT member_id) as members_with_positions,
            COUNT(*) as total_current_assignments,
            COUNT(DISTINCT department_id) as active_departments
        FROM member_positions
        WHERE is_current = 1
    """)
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('assign_member_positions.html',
                         current_assignments=current_assignments,
                         unassigned_members=unassigned_members,
                         all_members=all_members,
                         positions=positions,
                         departments=departments,
                         stats=stats,
                         search=search,
                         department_filter=department_filter)

@app.route('/member_career_history/<int:member_id>')
@login_required
def member_career_history(member_id):
    """View member's complete career/position history"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get member basic info
    cursor.execute("""
        SELECT id, full_name, section_name, phone, email, gender
        FROM member_registration WHERE id = %s
    """, (member_id,))
    member = cursor.fetchone()
    
    if not member:
        flash('Member not found', 'danger')
        return redirect(url_for('manage_members'))
    
    # Get current positions
    cursor.execute("""
        SELECT mp.id, p.position_title, d.department_name, mp.start_date, 
               mp.appointment_type, p.position_level, p.is_leadership,
               p.position_type, mp.notes
        FROM member_positions mp
        JOIN positions p ON mp.position_id = p.id
        LEFT JOIN departments d ON mp.department_id = d.id
        WHERE mp.member_id = %s AND mp.is_current = 1
        ORDER BY mp.start_date DESC
    """, (member_id,))
    current_positions = cursor.fetchall()
    
    # Get position history
    cursor.execute("""
        SELECT mp.id, p.position_title, d.department_name, mp.start_date, 
               mp.end_date, mp.appointment_type, p.position_level, p.is_leadership,
               p.position_type, mp.notes,
               DATEDIFF(COALESCE(mp.end_date, CURDATE()), mp.start_date) as days_served
        FROM member_positions mp
        JOIN positions p ON mp.position_id = p.id
        LEFT JOIN departments d ON mp.department_id = d.id
        WHERE mp.member_id = %s AND mp.is_current = 0
        ORDER BY mp.end_date DESC, mp.start_date DESC
    """, (member_id,))
    position_history = cursor.fetchall()
    
    # Get department head history
    cursor.execute("""
        SELECT dh.id, d.department_name, dh.start_date, dh.end_date, 
               dh.appointment_reason, dh.termination_reason, dh.is_current,
               DATEDIFF(COALESCE(dh.end_date, CURDATE()), dh.start_date) as days_served
        FROM department_head_history dh
        JOIN departments d ON dh.department_id = d.id
        WHERE dh.member_id = %s
        ORDER BY dh.start_date DESC
    """, (member_id,))
    dept_head_history = cursor.fetchall()
    
    # Calculate statistics
    total_positions_held = len(current_positions) + len(position_history)
    leadership_count = len([p for p in current_positions if p[6]]) + len([p for p in position_history if p[7]])
    total_service_days = sum([p[10] for p in position_history if p[10]])
    
    # Timeline data (combine all positions for timeline view)
    cursor.execute("""
        SELECT 'Position' as type, p.position_title as title, d.department_name as department,
               mp.start_date, mp.end_date, mp.is_current, p.is_leadership,
               mp.appointment_type as details
        FROM member_positions mp
        JOIN positions p ON mp.position_id = p.id
        LEFT JOIN departments d ON mp.department_id = d.id
        WHERE mp.member_id = %s
        
        UNION ALL
        
        SELECT 'Dept Head' as type, d.department_name as title, '' as department,
               dh.start_date, dh.end_date, dh.is_current, 1 as is_leadership,
               dh.appointment_reason as details
        FROM department_head_history dh
        JOIN departments d ON dh.department_id = d.id
        WHERE dh.member_id = %s
        
        ORDER BY start_date DESC, end_date DESC
    """, (member_id, member_id))
    timeline = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('member_career_history.html',
                         member=member,
                         current_positions=current_positions,
                         position_history=position_history,
                         dept_head_history=dept_head_history,
                         timeline=timeline,
                         total_positions=total_positions_held,
                         leadership_count=leadership_count,
                         total_service_days=total_service_days)

@app.route('/organizational_chart')
@login_required
def organizational_chart():
    """View organizational structure and member positions"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get all departments with hierarchy
    cursor.execute("""
        SELECT d.id, d.department_name, d.department_code, d.parent_department_id,
               pd.department_name as parent_name, d.head_member_id, m.full_name as head_name,
               d.description
        FROM departments d
        LEFT JOIN departments pd ON d.parent_department_id = pd.id
        LEFT JOIN member_registration m ON d.head_member_id = m.id
        WHERE d.status = 'Active'
        ORDER BY d.parent_department_id, d.department_name
    """)
    departments = cursor.fetchall()
    
    # Get all current member positions
    cursor.execute("""
        SELECT mp.id, d.id as dept_id, d.department_name, p.position_title, p.position_level,
               m.full_name, m.phone, mp.start_date, mp.appointment_type
        FROM member_positions mp
        JOIN member_registration m ON mp.member_id = m.id
        JOIN positions p ON mp.position_id = p.id
        JOIN departments d ON mp.department_id = d.id
        WHERE mp.is_current = 1
        ORDER BY d.department_name, p.position_level, p.position_title
    """)
    member_positions = cursor.fetchall()
    
    # Get summary stats
    cursor.execute("""
        SELECT 
            COUNT(DISTINCT d.id) as total_departments,
            COUNT(DISTINCT p.id) as total_positions,
            COUNT(DISTINCT mp.member_id) as members_with_positions
        FROM departments d
        LEFT JOIN positions p ON d.id = p.department_id AND p.status = 'Active'
        LEFT JOIN member_positions mp ON mp.is_current = 1
        WHERE d.status = 'Active'
    """)
    summary = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('organizational_chart.html',
                         departments=departments,
                         member_positions=member_positions,
                         summary=summary)

# ========================================
# APPLICATION INITIALIZATION
# ========================================

def initialize_app():
    """Initialize database tables and default data"""
    print("\n" + "="*60)
    print("Starting Application - Database Check")
    print("="*60)
    
    # Test database connection
    success, message = test_db_connection()
    if success:
        print(f"✓ {message}")
    else:
        print(f"✗ {message}")
        print("⚠ WARNING: Database is not available!")
        print("⚠ Application will start but you need to:")
        print("   1. Start your MySQL database (localhost:3306)")
        print("   2. Restart the application after database is up")
        print("="*60)
        return
    
    # Initialize RBAC tables
    if initialize_rbac_tables():
        print("✓ Database initialization completed")
    else:
        print("⚠ Database initialization failed but app will continue")
    
    print("="*60)
    
    # Initialize default roles and routes
    print("\nInitializing default roles and routes...")
    if initialize_default_roles_and_routes():
        print("✓ Default roles and routes setup completed")
    else:
        print("⚠ Default roles and routes setup failed but app will continue")
    print()

# ========================================
# MAIN APPLICATION ENTRY POINT
# ========================================

# ========================================
# LIBRARY MANAGEMENT ROUTES
# ========================================

@app.route('/manage_books', methods=['GET', 'POST'])
@login_required
def manage_books():
    """Manage library books - CRUD operations"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    book_data = {}
    
    if request.method == 'POST':
        action = request.form.get('action', 'add')
        
        if action == 'delete':
            book_id = request.form.get('id')
            try:
                cursor.execute("DELETE FROM library_books WHERE id = %s", (book_id,))
                conn.commit()
                flash('Book deleted successfully!', 'success')
            except Exception as e:
                flash(f'Error deleting book: {str(e)}', 'danger')
        
        else:
            form_data = {
                'title': request.form.get('title'),
                'author': request.form.get('author'),
                'category': request.form.get('category'),
                'isbn': request.form.get('isbn'),
                'publisher': request.form.get('publisher'),
                'publication_year': request.form.get('publication_year') or None,
                'language': request.form.get('language'),
                'total_copies': request.form.get('total_copies', 1),
                'available_copies': request.form.get('available_copies', 1),
                'shelf_location': request.form.get('shelf_location'),
                'description': request.form.get('description'),
                'status': request.form.get('status', 'Active')
            }
            
            book_id = request.form.get('id')
            
            try:
                if book_id and action == 'edit':  # UPDATE
                    query = """
                        UPDATE library_books SET
                            title = %s, author = %s, category = %s, isbn = %s,
                            publisher = %s, publication_year = %s, language = %s,
                            total_copies = %s, available_copies = %s,
                            shelf_location = %s, description = %s, status = %s
                        WHERE id = %s
                    """
                    values = (
                        form_data['title'], form_data['author'], form_data['category'],
                        form_data['isbn'], form_data['publisher'], form_data['publication_year'],
                        form_data['language'], form_data['total_copies'], form_data['available_copies'],
                        form_data['shelf_location'], form_data['description'], form_data['status'],
                        book_id
                    )
                    cursor.execute(query, values)
                    flash('Book updated successfully!', 'success')
                
                else:  # INSERT
                    query = """
                        INSERT INTO library_books (
                            title, author, category, isbn, publisher, publication_year,
                            language, total_copies, available_copies, shelf_location,
                            description, status, created_by
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    values = (
                        form_data['title'], form_data['author'], form_data['category'],
                        form_data['isbn'], form_data['publisher'], form_data['publication_year'],
                        form_data['language'], form_data['total_copies'], form_data['available_copies'],
                        form_data['shelf_location'], form_data['description'], form_data['status'],
                        session.get('payroll_number', 'SYSTEM')
                    )
                    cursor.execute(query, values)
                    flash('Book added successfully!', 'success')
                
                conn.commit()
                
            except mysql.connector.IntegrityError as e:
                if 'Duplicate entry' in str(e) and 'isbn' in str(e):
                    flash('Error: A book with this ISBN already exists!', 'danger')
                else:
                    flash(f'Database error: {str(e)}', 'danger')
            except Exception as e:
                flash(f'Error saving book: {str(e)}', 'danger')
    
    # Handle edit request
    edit_id = request.args.get('edit_id')
    if edit_id:
        cursor.execute("SELECT * FROM library_books WHERE id = %s", (edit_id,))
        book_row = cursor.fetchone()
        if book_row:
            book_data = {
                'id': book_row[0],
                'title': book_row[1],
                'author': book_row[2],
                'category': book_row[3],
                'isbn': book_row[4],
                'publisher': book_row[5],
                'publication_year': book_row[6],
                'language': book_row[7],
                'total_copies': book_row[8],
                'available_copies': book_row[9],
                'borrowed_copies': book_row[10],
                'shelf_location': book_row[11],
                'description': book_row[12],
                'cover_image': book_row[13],
                'status': book_row[14]
            }
    
    # Get all books with pagination and search
    search_query = request.args.get('search', '').strip()
    category_filter = request.args.get('category', '').strip()
    status_filter = request.args.get('status', '').strip()
    
    query = "SELECT * FROM library_books WHERE 1=1"
    params = []
    
    if search_query:
        query += " AND (title LIKE %s OR author LIKE %s OR isbn LIKE %s)"
        search_param = f"%{search_query}%"
        params.extend([search_param, search_param, search_param])
    
    if category_filter:
        query += " AND category = %s"
        params.append(category_filter)
    
    if status_filter:
        query += " AND status = %s"
        params.append(status_filter)
    
    query += " ORDER BY id DESC LIMIT 50"
    
    cursor.execute(query, params)
    books = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_books,
            SUM(total_copies) as total_copies,
            SUM(available_copies) as available_copies,
            SUM(borrowed_copies) as borrowed_copies,
            COUNT(DISTINCT category) as total_categories
        FROM library_books
        WHERE status = 'Active'
    """)
    stats = cursor.fetchone()
    library_stats = {
        'total_books': stats[0] or 0,
        'total_copies': stats[1] or 0,
        'available_copies': stats[2] or 0,
        'borrowed_copies': stats[3] or 0,
        'total_categories': stats[4] or 0
    }
    
    # Get categories for filter dropdown
    cursor.execute("SELECT DISTINCT category FROM library_books WHERE category IS NOT NULL ORDER BY category")
    categories = [row[0] for row in cursor.fetchall()]
    
    cursor.close()
    conn.close()
    
    return render_template('library_books.html',
                         books=books,
                         book_data=book_data,
                         library_stats=library_stats,
                         categories=categories,
                         search_query=search_query,
                         category_filter=category_filter,
                         status_filter=status_filter)

# ========================================
# BORROW MANAGEMENT ROUTES
# ========================================

@app.route('/borrow_management', methods=['GET', 'POST'])
@login_required
def borrow_management():
    """Manage book borrowing and returns"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'borrow':
            member_id = request.form.get('member_id')
            book_id = request.form.get('book_id')
            borrow_date = request.form.get('borrow_date') or date.today()
            due_date = request.form.get('due_date')
            notes = request.form.get('notes', '')
            
            try:
                # Check if book is available
                cursor.execute("SELECT available_copies FROM library_books WHERE id = %s", (book_id,))
                result = cursor.fetchone()
                
                if result and result[0] > 0:
                    # Insert borrowing record
                    cursor.execute("""
                        INSERT INTO book_borrowing (
                            member_id, book_id, borrow_date, due_date, status, notes, borrowed_by
                        ) VALUES (%s, %s, %s, %s, 'Borrowed', %s, %s)
                    """, (member_id, book_id, borrow_date, due_date, notes, session.get('payroll_number')))
                    
                    # Update book availability
                    cursor.execute("""
                        UPDATE library_books 
                        SET available_copies = available_copies - 1,
                            borrowed_copies = borrowed_copies + 1
                        WHERE id = %s
                    """, (book_id,))
                    
                    conn.commit()
                    flash('Book borrowed successfully!', 'success')
                else:
                    flash('This book is not available for borrowing.', 'warning')
                    
            except Exception as e:
                flash(f'Error processing borrow: {str(e)}', 'danger')
                conn.rollback()
        
        elif action == 'return':
            borrow_id = request.form.get('borrow_id')
            return_date = request.form.get('return_date') or date.today()
            
            try:
                # Get borrow record
                cursor.execute("SELECT book_id FROM book_borrowing WHERE id = %s", (borrow_id,))
                result = cursor.fetchone()
                
                if result:
                    book_id = result[0]
                    
                    # Update borrowing record
                    cursor.execute("""
                        UPDATE book_borrowing 
                        SET return_date = %s, status = 'Returned', returned_to = %s
                        WHERE id = %s
                    """, (return_date, session.get('payroll_number'), borrow_id))
                    
                    # Update book availability
                    cursor.execute("""
                        UPDATE library_books 
                        SET available_copies = available_copies + 1,
                            borrowed_copies = borrowed_copies - 1
                        WHERE id = %s
                    """, (book_id,))
                    
                    conn.commit()
                    flash('Book returned successfully!', 'success')
                else:
                    flash('Borrow record not found.', 'danger')
                    
            except Exception as e:
                flash(f'Error processing return: {str(e)}', 'danger')
                conn.rollback()
    
    # Get all members for dropdown
    cursor.execute("SELECT id, full_name FROM member_registration ORDER BY full_name")
    members = cursor.fetchall()
    
    # Get available books for dropdown
    cursor.execute("""
        SELECT id, title, author, available_copies 
        FROM library_books 
        WHERE status = 'Active' AND available_copies > 0
        ORDER BY title
    """)
    available_books = cursor.fetchall()
    
    # Get all borrowing records with member and book details
    cursor.execute("""
        SELECT 
            bb.id,
            m.full_name,
            lb.title,
            lb.author,
            bb.borrow_date,
            bb.due_date,
            bb.return_date,
            bb.status,
            bb.notes,
            DATEDIFF(CURDATE(), bb.due_date) as days_overdue
        FROM book_borrowing bb
        JOIN member_registration m ON bb.member_id = m.id
        JOIN library_books lb ON bb.book_id = lb.id
        ORDER BY bb.id DESC
        LIMIT 100
    """)
    borrowings = cursor.fetchall()
    
    # Update overdue status
    cursor.execute("""
        UPDATE book_borrowing 
        SET status = 'Overdue'
        WHERE status = 'Borrowed' AND due_date < CURDATE()
    """)
    conn.commit()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(CASE WHEN status = 'Borrowed' THEN 1 END) as currently_borrowed,
            COUNT(CASE WHEN status = 'Overdue' THEN 1 END) as overdue,
            COUNT(CASE WHEN status = 'Returned' THEN 1 END) as returned,
            COUNT(*) as total_transactions
        FROM book_borrowing
    """)
    stats = cursor.fetchone()
    borrow_stats = {
        'currently_borrowed': stats[0] or 0,
        'overdue': stats[1] or 0,
        'returned': stats[2] or 0,
        'total_transactions': stats[3] or 0
    }
    
    cursor.close()
    conn.close()
    
    return render_template('borrow_management.html',
                         members=members,
                         available_books=available_books,
                         borrowings=borrowings,
                         borrow_stats=borrow_stats)

# ========================================
# LIBRARY REPORTS
# ========================================

@app.route('/book_report')
@login_required
def book_report():
    """Library book report with filters and export"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filters
    category_filter = request.args.get('category', '')
    status_filter = request.args.get('status', '')
    language_filter = request.args.get('language', '')
    
    # Build query
    query = "SELECT * FROM library_books WHERE 1=1"
    params = []
    
    if category_filter:
        query += " AND category = %s"
        params.append(category_filter)
    
    if status_filter:
        query += " AND status = %s"
        params.append(status_filter)
    
    if language_filter:
        query += " AND language = %s"
        params.append(language_filter)
    
    query += " ORDER BY category, title"
    
    cursor.execute(query, params)
    books = cursor.fetchall()
    
    # Get summary statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_books,
            SUM(total_copies) as total_copies,
            SUM(available_copies) as available_copies,
            SUM(borrowed_copies) as borrowed_copies,
            COUNT(DISTINCT category) as categories,
            COUNT(CASE WHEN available_copies = 0 THEN 1 END) as out_of_stock
        FROM library_books
        WHERE status = 'Active'
    """)
    stats = cursor.fetchone()
    summary_stats = {
        'total_books': stats[0] or 0,
        'total_copies': stats[1] or 0,
        'available': stats[2] or 0,
        'borrowed': stats[3] or 0,
        'categories': stats[4] or 0,
        'out_of_stock': stats[5] or 0
    }
    
    # Get filter options
    cursor.execute("SELECT DISTINCT category FROM library_books WHERE category IS NOT NULL ORDER BY category")
    categories = [row[0] for row in cursor.fetchall()]
    
    cursor.execute("SELECT DISTINCT language FROM library_books WHERE language IS NOT NULL ORDER BY language")
    languages = [row[0] for row in cursor.fetchall()]
    
    cursor.close()
    conn.close()
    
    return render_template('book_report.html',
                         books=books,
                         summary_stats=summary_stats,
                         categories=categories,
                         languages=languages,
                         category_filter=category_filter,
                         status_filter=status_filter,
                         language_filter=language_filter)

@app.route('/borrow_report')
@login_required
def borrow_report():
    """Borrow transactions report with filters and export"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filters
    member_filter = request.args.get('member_id', '')
    book_filter = request.args.get('book_id', '')
    status_filter = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = """
        SELECT 
            bb.id,
            m.full_name as member_name,
            lb.title as book_title,
            lb.author,
            bb.borrow_date,
            bb.due_date,
            bb.return_date,
            bb.status,
            DATEDIFF(CURDATE(), bb.due_date) as days_overdue,
            bb.notes
        FROM book_borrowing bb
        JOIN member_registration m ON bb.member_id = m.id
        JOIN library_books lb ON bb.book_id = lb.id
        WHERE 1=1
    """
    params = []
    
    if member_filter:
        query += " AND bb.member_id = %s"
        params.append(member_filter)
    
    if book_filter:
        query += " AND bb.book_id = %s"
        params.append(book_filter)
    
    if status_filter:
        query += " AND bb.status = %s"
        params.append(status_filter)
    
    if date_from:
        query += " AND bb.borrow_date >= %s"
        params.append(date_from)
    
    if date_to:
        query += " AND bb.borrow_date <= %s"
        params.append(date_to)
    
    query += " ORDER BY bb.borrow_date DESC"
    
    cursor.execute(query, params)
    transactions = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(CASE WHEN status = 'Borrowed' THEN 1 END) as borrowed,
            COUNT(CASE WHEN status = 'Overdue' THEN 1 END) as overdue,
            COUNT(CASE WHEN status = 'Returned' THEN 1 END) as returned,
            COUNT(*) as total
        FROM book_borrowing
    """)
    stats = cursor.fetchone()
    summary_stats = {
        'borrowed': stats[0] or 0,
        'overdue': stats[1] or 0,
        'returned': stats[2] or 0,
        'total': stats[3] or 0
    }
    
    # Get members and books for filter dropdowns
    cursor.execute("SELECT id, full_name FROM member_registration ORDER BY full_name")
    members = cursor.fetchall()
    
    cursor.execute("SELECT id, title FROM library_books ORDER BY title")
    books_list = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('borrow_report.html',
                         transactions=transactions,
                         summary_stats=summary_stats,
                         members=members,
                         books_list=books_list,
                         member_filter=member_filter,
                         book_filter=book_filter,
                         status_filter=status_filter,
                         date_from=date_from,
                         date_to=date_to)

# ========================================
# MEWACO (CONTRIBUTION) MANAGEMENT ROUTES
# ========================================

@app.route('/mewaco_types', methods=['GET', 'POST'])
@login_required
def mewaco_types():
    """Manage MEWACO contribution types"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    type_data = {}
    
    if request.method == 'POST':
        action = request.form.get('action', 'add')
        
        if action == 'delete':
            type_id = request.form.get('id')
            try:
                cursor.execute("DELETE FROM mewaco_types WHERE id = %s", (type_id,))
                conn.commit()
                flash('Contribution type deleted successfully!', 'success')
            except Exception as e:
                flash(f'Error deleting type: {str(e)}', 'danger')
        
        else:
            form_data = {
                'type_name': request.form.get('type_name'),
                'description': request.form.get('description'),
                'default_amount': request.form.get('default_amount', 0),
                'status': request.form.get('status', 'Active')
            }
            
            type_id = request.form.get('id')
            
            try:
                if type_id and action == 'edit':  # UPDATE
                    query = """
                        UPDATE mewaco_types SET
                            type_name = %s, description = %s, 
                            default_amount = %s, status = %s
                        WHERE id = %s
                    """
                    values = (
                        form_data['type_name'], form_data['description'],
                        form_data['default_amount'], form_data['status'], type_id
                    )
                    cursor.execute(query, values)
                    flash('Contribution type updated successfully!', 'success')
                
                else:  # INSERT
                    query = """
                        INSERT INTO mewaco_types (
                            type_name, description, default_amount, status, created_by
                        ) VALUES (%s, %s, %s, %s, %s)
                    """
                    values = (
                        form_data['type_name'], form_data['description'],
                        form_data['default_amount'], form_data['status'],
                        session.get('payroll_number', 'SYSTEM')
                    )
                    cursor.execute(query, values)
                    flash('Contribution type added successfully!', 'success')
                
                conn.commit()
                
            except mysql.connector.IntegrityError as e:
                if 'Duplicate entry' in str(e):
                    flash('Error: A contribution type with this name already exists!', 'danger')
                else:
                    flash(f'Database error: {str(e)}', 'danger')
            except Exception as e:
                flash(f'Error saving type: {str(e)}', 'danger')
    
    # Handle edit request
    edit_id = request.args.get('edit_id')
    if edit_id:
        cursor.execute("SELECT * FROM mewaco_types WHERE id = %s", (edit_id,))
        type_row = cursor.fetchone()
        if type_row:
            type_data = {
                'id': type_row[0],
                'type_name': type_row[1],
                'description': type_row[2],
                'default_amount': type_row[3],
                'status': type_row[4]
            }
    
    # Get all contribution types
    cursor.execute("SELECT * FROM mewaco_types ORDER BY id DESC")
    types = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_types,
            COUNT(CASE WHEN status = 'Active' THEN 1 END) as active_types,
            SUM(default_amount) as total_default_amount
        FROM mewaco_types
    """)
    stats = cursor.fetchone()
    type_stats = {
        'total_types': stats[0] or 0,
        'active_types': stats[1] or 0,
        'total_default': stats[2] or 0
    }
    
    cursor.close()
    conn.close()
    
    return render_template('mewaco_types.html',
                         types=types,
                         type_data=type_data,
                         type_stats=type_stats)

@app.route('/mewaco_contributions', methods=['GET', 'POST'])
@login_required
def mewaco_contributions():
    """Manage monthly/user-based contributions"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        action = request.form.get('action', 'add')
        
        if action == 'delete':
            contrib_id = request.form.get('id')
            try:
                cursor.execute("DELETE FROM mewaco_contributions WHERE id = %s", (contrib_id,))
                conn.commit()
                flash('Contribution deleted successfully!', 'success')
            except Exception as e:
                flash(f'Error deleting contribution: {str(e)}', 'danger')
        
        else:
            form_data = {
                'member_id': request.form.get('member_id'),
                'mewaco_type_id': request.form.get('mewaco_type_id'),
                'contribution_date': request.form.get('contribution_date') or date.today(),
                'amount': request.form.get('amount', 0),
                'payment_method': request.form.get('payment_method', 'Cash'),
                'receipt_number': request.form.get('receipt_number'),
                'notes': request.form.get('notes', '')
            }
            
            contrib_id = request.form.get('id')
            
            try:
                if contrib_id and action == 'edit':  # UPDATE
                    query = """
                        UPDATE mewaco_contributions SET
                            member_id = %s, mewaco_type_id = %s, contribution_date = %s,
                            amount = %s, payment_method = %s, receipt_number = %s, notes = %s
                        WHERE id = %s
                    """
                    values = (
                        form_data['member_id'], form_data['mewaco_type_id'], 
                        form_data['contribution_date'], form_data['amount'],
                        form_data['payment_method'], form_data['receipt_number'],
                        form_data['notes'], contrib_id
                    )
                    cursor.execute(query, values)
                    flash('Contribution updated successfully!', 'success')
                
                else:  # INSERT
                    query = """
                        INSERT INTO mewaco_contributions (
                            member_id, mewaco_type_id, contribution_date, amount,
                            payment_method, receipt_number, notes, recorded_by
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    values = (
                        form_data['member_id'], form_data['mewaco_type_id'],
                        form_data['contribution_date'], form_data['amount'],
                        form_data['payment_method'], form_data['receipt_number'],
                        form_data['notes'], session.get('payroll_number', 'SYSTEM')
                    )
                    cursor.execute(query, values)
                    flash('Contribution recorded successfully!', 'success')
                
                conn.commit()
                
            except Exception as e:
                flash(f'Error saving contribution: {str(e)}', 'danger')
                conn.rollback()
    
    # Get filters
    member_filter = request.args.get('member_id', '')
    type_filter = request.args.get('type_id', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query for contributions
    query = """
        SELECT 
            mc.id, m.full_name, mt.type_name, mc.contribution_date,
            mc.amount, mc.payment_method, mc.receipt_number, mc.notes
        FROM mewaco_contributions mc
        JOIN member_registration m ON mc.member_id = m.id
        JOIN mewaco_types mt ON mc.mewaco_type_id = mt.id
        WHERE 1=1
    """
    params = []
    
    if member_filter:
        query += " AND mc.member_id = %s"
        params.append(member_filter)
    
    if type_filter:
        query += " AND mc.mewaco_type_id = %s"
        params.append(type_filter)
    
    if date_from:
        query += " AND mc.contribution_date >= %s"
        params.append(date_from)
    
    if date_to:
        query += " AND mc.contribution_date <= %s"
        params.append(date_to)
    
    query += " ORDER BY mc.contribution_date DESC LIMIT 100"
    
    cursor.execute(query, params)
    contributions = cursor.fetchall()
    
    # Get members and types for dropdowns
    cursor.execute("SELECT id, full_name FROM member_registration ORDER BY full_name")
    members = cursor.fetchall()
    
    cursor.execute("SELECT id, type_name, default_amount FROM mewaco_types WHERE status = 'Active' ORDER BY type_name")
    mewaco_types = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_contributions,
            SUM(amount) as total_amount,
            COUNT(DISTINCT member_id) as unique_contributors,
            COUNT(DISTINCT mewaco_type_id) as types_used
        FROM mewaco_contributions
    """)
    stats = cursor.fetchone()
    contrib_stats = {
        'total_contributions': stats[0] or 0,
        'total_amount': stats[1] or 0,
        'unique_contributors': stats[2] or 0,
        'types_used': stats[3] or 0
    }
    
    cursor.close()
    conn.close()
    
    return render_template('mewaco_contributions.html',
                         members=members,
                         mewaco_types=mewaco_types,
                         contributions=contributions,
                         contrib_stats=contrib_stats,
                         member_filter=member_filter,
                         type_filter=type_filter,
                         date_from=date_from,
                         date_to=date_to)

@app.route('/monthly_contributions', methods=['GET', 'POST'])
@login_required
def monthly_contributions():
    """Bulk monthly contribution collection page"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get selected month/year and type
    selected_month = request.args.get('month') or datetime.now().strftime('%Y-%m')
    selected_type_id = request.args.get('type_id', '')
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'bulk_save':
            # Process bulk contribution data
            contribution_type_id = request.form.get('contribution_type_id')
            contribution_month = request.form.get('contribution_month')
            
            try:
                saved_count = 0
                for key in request.form:
                    if key.startswith('amount_'):
                        member_id = key.split('_')[1]
                        amount = request.form.get(key)
                        status = request.form.get(f'status_{member_id}', 'Unpaid')
                        
                        if amount and float(amount) > 0 and status == 'Paid':
                            # Record contribution
                            cursor.execute("""
                                INSERT INTO mewaco_contributions (
                                    member_id, mewaco_type_id, contribution_date, amount,
                                    payment_method, notes, recorded_by
                                ) VALUES (%s, %s, %s, %s, 'Cash', %s, %s)
                            """, (
                                member_id, contribution_type_id, contribution_month + '-01',
                                amount, f'Monthly contribution - {contribution_month}',
                                session.get('payroll_number', 'SYSTEM')
                            ))
                            saved_count += 1
                
                conn.commit()
                flash(f'Successfully recorded {saved_count} contributions!', 'success')
                
            except Exception as e:
                flash(f'Error saving contributions: {str(e)}', 'danger')
                conn.rollback()
        
        elif action == 'bulk_upload':
            # Handle Excel/CSV upload
            if 'file' in request.files:
                file = request.files['file']
                if file and file.filename:
                    try:
                        df = pd.read_excel(file) if file.filename.endswith('.xlsx') else pd.read_csv(file)
                        
                        saved_count = 0
                        for _, row in df.iterrows():
                            cursor.execute("""
                                SELECT id FROM member_registration WHERE full_name = %s
                            """, (row['Member Name'],))
                            member = cursor.fetchone()
                            
                            if member:
                                cursor.execute("""
                                    INSERT INTO mewaco_contributions (
                                        member_id, mewaco_type_id, contribution_date, amount,
                                        payment_method, recorded_by
                                    ) VALUES (%s, %s, %s, %s, %s, %s)
                                """, (
                                    member[0], request.form.get('upload_type_id'),
                                    row['Date'], row['Amount'], row.get('Payment Method', 'Cash'),
                                    session.get('payroll_number', 'SYSTEM')
                                ))
                                saved_count += 1
                        
                        conn.commit()
                        flash(f'Successfully uploaded {saved_count} contributions!', 'success')
                        
                    except Exception as e:
                        flash(f'Error processing upload: {str(e)}', 'danger')
                        conn.rollback()
    
    # Get all members
    cursor.execute("SELECT id, full_name, section_name FROM member_registration ORDER BY section_name, full_name")
    members = cursor.fetchall()
    
    # Get active contribution types
    cursor.execute("SELECT id, type_name, default_amount FROM mewaco_types WHERE status = 'Active' ORDER BY type_name")
    mewaco_types = cursor.fetchall()
    
    # Get existing contributions for selected month and type
    existing_contributions = {}
    if selected_type_id and selected_month:
        cursor.execute("""
            SELECT member_id, SUM(amount) as total_amount
            FROM mewaco_contributions
            WHERE mewaco_type_id = %s 
            AND DATE_FORMAT(contribution_date, '%%Y-%%m') = %s
            GROUP BY member_id
        """, (selected_type_id, selected_month))
        existing_contributions = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Calculate summary
    if selected_type_id and members:
        cursor.execute("SELECT default_amount FROM mewaco_types WHERE id = %s", (selected_type_id,))
        default_amt = cursor.fetchone()
        default_amount = default_amt[0] if default_amt else 0
        
        total_expected = len(members) * default_amount
        total_collected = sum(existing_contributions.values())
        total_outstanding = total_expected - total_collected
        paid_count = len(existing_contributions)
        unpaid_count = len(members) - paid_count
    else:
        total_expected = 0
        total_collected = 0
        total_outstanding = 0
        paid_count = 0
        unpaid_count = len(members) if members else 0
    
    summary = {
        'total_expected': total_expected,
        'total_collected': total_collected,
        'total_outstanding': total_outstanding,
        'paid_count': paid_count,
        'unpaid_count': unpaid_count
    }
    
    cursor.close()
    conn.close()
    
    return render_template('monthly_contributions.html',
                         members=members,
                         mewaco_types=mewaco_types,
                         selected_month=selected_month,
                         selected_type_id=selected_type_id,
                         existing_contributions=existing_contributions,
                         summary=summary)

@app.route('/contribution_report_monthly')
@login_required
def contribution_report_monthly():
    """Monthly contribution report with trends"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filters
    type_filter = request.args.get('type_id', '')
    month_filter = request.args.get('month', '')
    year_filter = request.args.get('year', datetime.now().year)
    
    # Get contribution types for dropdown
    cursor.execute("SELECT id, type_name FROM mewaco_types ORDER BY type_name")
    contrib_types = cursor.fetchall()
    
    # Build query
    query = """
        SELECT 
            DATE_FORMAT(mc.contribution_date, '%%Y-%%m') as month,
            mt.type_name,
            COUNT(DISTINCT mc.member_id) as contributor_count,
            SUM(mc.amount) as total_amount,
            COUNT(*) as transaction_count
        FROM mewaco_contributions mc
        JOIN mewaco_types mt ON mc.mewaco_type_id = mt.id
        WHERE YEAR(mc.contribution_date) = %s
    """
    params = [year_filter]
    
    if type_filter:
        query += " AND mc.mewaco_type_id = %s"
        params.append(type_filter)
    
    if month_filter:
        query += " AND DATE_FORMAT(mc.contribution_date, '%%Y-%%m') = %s"
        params.append(month_filter)
    
    query += " GROUP BY month, mt.type_name ORDER BY month DESC, mt.type_name"
    
    cursor.execute(query, params)
    monthly_data = cursor.fetchall()
    
    # Get chart data for trends (last 12 months)
    cursor.execute("""
        SELECT 
            DATE_FORMAT(contribution_date, '%%Y-%%m') as month,
            SUM(amount) as total
        FROM mewaco_contributions
        WHERE contribution_date >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
        GROUP BY month
        ORDER BY month
    """)
    chart_data = cursor.fetchall()
    chart_labels = [row[0] for row in chart_data]
    chart_amounts = [float(row[1]) for row in chart_data]
    
    # Overall statistics
    cursor.execute("""
        SELECT 
            SUM(amount) as grand_total,
            COUNT(*) as total_transactions,
            COUNT(DISTINCT member_id) as unique_contributors
        FROM mewaco_contributions
        WHERE YEAR(contribution_date) = %s
    """, (year_filter,))
    stats = cursor.fetchone()
    overall_stats = {
        'grand_total': stats[0] or 0,
        'total_transactions': stats[1] or 0,
        'unique_contributors': stats[2] or 0
    }
    
    cursor.close()
    conn.close()
    
    return render_template('contribution_report_monthly.html',
                         monthly_data=monthly_data,
                         contrib_types=contrib_types,
                         chart_labels=chart_labels,
                         chart_amounts=chart_amounts,
                         overall_stats=overall_stats,
                         type_filter=type_filter,
                         month_filter=month_filter,
                         year_filter=year_filter)

@app.route('/member_contribution_summary')
@login_required
def member_contribution_summary():
    """Individual member contribution summary and history"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filters
    member_filter = request.args.get('member_id', '')
    year_filter = request.args.get('year', datetime.now().year)
    
    # Get all members for dropdown
    cursor.execute("SELECT id, full_name FROM member_registration ORDER BY full_name")
    members = cursor.fetchall()
    
    # Get member contribution summary
    query = """
        SELECT 
            m.full_name,
            mt.type_name,
            COUNT(*) as payment_count,
            SUM(mc.amount) as total_paid,
            MIN(mc.contribution_date) as first_payment,
            MAX(mc.contribution_date) as last_payment
        FROM mewaco_contributions mc
        JOIN member_registration m ON mc.member_id = m.id
        JOIN mewaco_types mt ON mc.mewaco_type_id = mt.id
        WHERE YEAR(mc.contribution_date) = %s
    """
    params = [year_filter]
    
    if member_filter:
        query += " AND mc.member_id = %s"
        params.append(member_filter)
    
    query += " GROUP BY m.full_name, mt.type_name ORDER BY m.full_name, mt.type_name"
    
    cursor.execute(query, params)
    member_summary = cursor.fetchall()
    
    # Get detailed transaction history if member is selected
    transaction_history = []
    if member_filter:
        cursor.execute("""
            SELECT 
                mc.contribution_date,
                mt.type_name,
                mc.amount,
                mc.payment_method,
                mc.receipt_number,
                mc.notes
            FROM mewaco_contributions mc
            JOIN mewaco_types mt ON mc.mewaco_type_id = mt.id
            WHERE mc.member_id = %s
            AND YEAR(mc.contribution_date) = %s
            ORDER BY mc.contribution_date DESC
        """, (member_filter, year_filter))
        transaction_history = cursor.fetchall()
    
    # Get overall statistics
    cursor.execute("""
        SELECT 
            COUNT(DISTINCT m.id) as total_contributors,
            SUM(mc.amount) as grand_total,
            AVG(mc.amount) as average_contribution
        FROM mewaco_contributions mc
        JOIN member_registration m ON mc.member_id = m.id
        WHERE YEAR(mc.contribution_date) = %s
    """, (year_filter,))
    stats = cursor.fetchone()
    summary_stats = {
        'total_contributors': stats[0] or 0,
        'grand_total': stats[1] or 0,
        'average_contribution': stats[2] or 0
    }
    
    cursor.close()
    conn.close()
    
    return render_template('member_contribution_summary.html',
                         members=members,
                         member_summary=member_summary,
                         transaction_history=transaction_history,
                         summary_stats=summary_stats,
                         member_filter=member_filter,
                         year_filter=year_filter)

# ========================================
# MEDEBE (SECTION & SUB-SECTION) MANAGEMENT ROUTES
# ========================================

@app.route('/medebe_management', methods=['GET', 'POST'])
@login_required
def medebe_management():
    """Manage medebe (sub-sections)"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    medebe_data = {}
    
    if request.method == 'POST':
        action = request.form.get('action', 'add')
        
        if action == 'delete':
            medebe_id = request.form.get('id')
            try:
                cursor.execute("DELETE FROM medebe WHERE id = %s", (medebe_id,))
                conn.commit()
                flash('Medebe deleted successfully!', 'success')
            except Exception as e:
                flash(f'Error deleting medebe: {str(e)}', 'danger')
        
        else:
            form_data = {
                'medebe_name': request.form.get('medebe_name'),
                'section_name': request.form.get('section_name'),
                'description': request.form.get('description', ''),
                'created_date': request.form.get('created_date') or date.today()
            }
            
            medebe_id = request.form.get('id')
            
            try:
                if medebe_id and action == 'edit':  # UPDATE
                    query = """
                        UPDATE medebe SET
                            medebe_name = %s, section_name = %s, 
                            description = %s, created_date = %s
                        WHERE id = %s
                    """
                    values = (
                        form_data['medebe_name'], form_data['section_name'],
                        form_data['description'], form_data['created_date'], medebe_id
                    )
                    cursor.execute(query, values)
                    flash('Medebe updated successfully!', 'success')
                
                else:  # INSERT
                    query = """
                        INSERT INTO medebe (
                            medebe_name, section_name, description, created_date, created_by
                        ) VALUES (%s, %s, %s, %s, %s)
                    """
                    values = (
                        form_data['medebe_name'], form_data['section_name'],
                        form_data['description'], form_data['created_date'],
                        session.get('payroll_number', 'SYSTEM')
                    )
                    cursor.execute(query, values)
                    flash('Medebe added successfully!', 'success')
                
                conn.commit()
                
            except Exception as e:
                flash(f'Error saving medebe: {str(e)}', 'danger')
    
    # Handle edit request
    edit_id = request.args.get('edit_id')
    if edit_id:
        cursor.execute("SELECT * FROM medebe WHERE id = %s", (edit_id,))
        medebe_row = cursor.fetchone()
        if medebe_row:
            medebe_data = {
                'id': medebe_row[0],
                'medebe_name': medebe_row[1],
                'section_name': medebe_row[2],
                'description': medebe_row[3],
                'created_date': medebe_row[4]
            }
    
    # Get filters
    section_filter = request.args.get('section', '')
    search = request.args.get('search', '')
    
    # Build query
    query = "SELECT * FROM medebe WHERE 1=1"
    params = []
    
    if section_filter:
        query += " AND section_name = %s"
        params.append(section_filter)
    
    if search:
        query += " AND (medebe_name LIKE %s OR description LIKE %s)"
        params.extend([f'%{search}%', f'%{search}%'])
    
    query += " ORDER BY section_name, id DESC"
    
    cursor.execute(query, params)
    medebes = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_medebes,
            COUNT(DISTINCT section_name) as sections_with_medebes
        FROM medebe
    """)
    stats = cursor.fetchone()
    medebe_stats = {
        'total_medebes': stats[0] or 0,
        'sections_with_medebes': stats[1] or 0
    }
    
    # Get member counts per medebe
    cursor.execute("""
        SELECT m.id, m.medebe_name, COUNT(mma.id) as member_count
        FROM medebe m
        LEFT JOIN member_medebe_assignment mma ON m.id = mma.medebe_id
        GROUP BY m.id, m.medebe_name
    """)
    member_counts = {row[0]: row[2] for row in cursor.fetchall()}
    
    cursor.close()
    conn.close()
    
    return render_template('medebe_management.html',
                         medebes=medebes,
                         medebe_data=medebe_data,
                         medebe_stats=medebe_stats,
                         member_counts=member_counts,
                         section_filter=section_filter,
                         search=search)

@app.route('/member_medebe_assignment', methods=['GET', 'POST'])
@login_required
def member_medebe_assignment():
    """Assign members to medebe (sub-sections)"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'assign':
            member_id = request.form.get('member_id')
            medebe_id = request.form.get('medebe_id')
            
            try:
                # Get member's section and medebe's section
                cursor.execute("SELECT section_name FROM member_registration WHERE id = %s", (member_id,))
                member_section = cursor.fetchone()[0]
                
                cursor.execute("SELECT section_name FROM medebe WHERE id = %s", (medebe_id,))
                medebe_section = cursor.fetchone()[0]
                
                # Validation: member and medebe must be in same section
                if member_section != medebe_section:
                    flash(f'Error: Member from {member_section} cannot be assigned to medebe in {medebe_section}!', 'danger')
                else:
                    # Check if already assigned
                    cursor.execute("SELECT id FROM member_medebe_assignment WHERE member_id = %s", (member_id,))
                    existing = cursor.fetchone()
                    
                    if existing:
                        # Update existing assignment
                        cursor.execute("""
                            UPDATE member_medebe_assignment 
                            SET medebe_id = %s, section_name = %s, assigned_date = %s, 
                                assigned_by = %s, assignment_method = 'Manual'
                            WHERE member_id = %s
                        """, (medebe_id, member_section, date.today(), 
                              session.get('payroll_number', 'SYSTEM'), member_id))
                        flash('Member reassigned successfully!', 'success')
                    else:
                        # Insert new assignment
                        cursor.execute("""
                            INSERT INTO member_medebe_assignment (
                                member_id, medebe_id, section_name, assigned_date, 
                                assigned_by, assignment_method
                            ) VALUES (%s, %s, %s, %s, %s, 'Manual')
                        """, (member_id, medebe_id, member_section, date.today(),
                              session.get('payroll_number', 'SYSTEM')))
                        flash('Member assigned successfully!', 'success')
                    
                    conn.commit()
                    
            except Exception as e:
                flash(f'Error assigning member: {str(e)}', 'danger')
                conn.rollback()
        
        elif action == 'auto_assign':
            section_name = request.form.get('section_name')
            
            try:
                # Get all unassigned members in this section
                cursor.execute("""
                    SELECT m.id FROM member_registration m
                    LEFT JOIN member_medebe_assignment mma ON m.id = mma.member_id
                    WHERE m.section_name = %s AND mma.id IS NULL
                """, (section_name,))
                unassigned_members = [row[0] for row in cursor.fetchall()]
                
                # Get all medebes for this section
                cursor.execute("SELECT id FROM medebe WHERE section_name = %s ORDER BY id", (section_name,))
                medebes = [row[0] for row in cursor.fetchall()]
                
                if not medebes:
                    flash(f'No medebe found for {section_name}!', 'danger')
                elif not unassigned_members:
                    flash(f'All members in {section_name} are already assigned!', 'info')
                else:
                    # Distribute members evenly across medebes
                    import random
                    random.shuffle(unassigned_members)
                    
                    assigned_count = 0
                    for i, member_id in enumerate(unassigned_members):
                        medebe_id = medebes[i % len(medebes)]
                        cursor.execute("""
                            INSERT INTO member_medebe_assignment (
                                member_id, medebe_id, section_name, assigned_date,
                                assigned_by, assignment_method
                            ) VALUES (%s, %s, %s, %s, %s, 'Auto')
                        """, (member_id, medebe_id, section_name, date.today(),
                              session.get('payroll_number', 'SYSTEM')))
                        assigned_count += 1
                    
                    conn.commit()
                    flash(f'Successfully auto-assigned {assigned_count} members to {len(medebes)} medebe groups!', 'success')
                    
            except Exception as e:
                flash(f'Error in auto-assignment: {str(e)}', 'danger')
                conn.rollback()
        
        elif action == 'remove':
            member_id = request.form.get('member_id')
            try:
                cursor.execute("DELETE FROM member_medebe_assignment WHERE member_id = %s", (member_id,))
                conn.commit()
                flash('Member removed from medebe!', 'success')
            except Exception as e:
                flash(f'Error removing member: {str(e)}', 'danger')
    
    # Get filters
    section_filter = request.args.get('section', '')
    medebe_filter = request.args.get('medebe', '')
    status_filter = request.args.get('status', '')
    
    # Get all sections
    sections = ['የሕፃናት ክፍል', 'ማህከላዊያን ክፍል', 'ወጣት ክፍል', 'ወላጅ ክፍል']
    
    # Get medebes for filter
    if section_filter:
        cursor.execute("SELECT id, medebe_name FROM medebe WHERE section_name = %s ORDER BY medebe_name", (section_filter,))
    else:
        cursor.execute("SELECT id, medebe_name FROM medebe ORDER BY section_name, medebe_name")
    medebes = cursor.fetchall()
    
    # Build member query
    query = """
        SELECT 
            m.id, m.full_name, m.phone, m.section_name,
            mma.medebe_id, meb.medebe_name, mma.assigned_date
        FROM member_registration m
        LEFT JOIN member_medebe_assignment mma ON m.id = mma.member_id
        LEFT JOIN medebe meb ON mma.medebe_id = meb.id
        WHERE 1=1
    """
    params = []
    
    if section_filter:
        query += " AND m.section_name = %s"
        params.append(section_filter)
    
    if medebe_filter:
        query += " AND mma.medebe_id = %s"
        params.append(medebe_filter)
    
    if status_filter == 'Assigned':
        query += " AND mma.id IS NOT NULL"
    elif status_filter == 'Not Assigned':
        query += " AND mma.id IS NULL"
    
    query += " ORDER BY m.section_name, m.full_name"
    
    cursor.execute(query, params)
    members = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(DISTINCT m.id) as total_members,
            COUNT(DISTINCT mma.id) as assigned_members
        FROM member_registration m
        LEFT JOIN member_medebe_assignment mma ON m.id = mma.member_id
    """)
    stats = cursor.fetchone()
    assignment_stats = {
        'total_members': stats[0] or 0,
        'assigned_members': stats[1] or 0,
        'unassigned_members': (stats[0] or 0) - (stats[1] or 0)
    }
    
    # Get medebes for assignment dropdown
    cursor.execute("SELECT id, medebe_name, section_name FROM medebe ORDER BY section_name, medebe_name")
    all_medebes = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('member_medebe_assignment.html',
                         members=members,
                         medebes=medebes,
                         all_medebes=all_medebes,
                         sections=sections,
                         assignment_stats=assignment_stats,
                         section_filter=section_filter,
                         medebe_filter=medebe_filter,
                         status_filter=status_filter)

@app.route('/medebe_report')
@login_required
def medebe_report():
    """Medebe statistics and reports"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filter
    section_filter = request.args.get('section', '')
    
    # Build query for medebe summary
    query = """
        SELECT 
            m.id, m.medebe_name, m.section_name, m.created_date, m.description,
            COUNT(mma.id) as member_count
        FROM medebe m
        LEFT JOIN member_medebe_assignment mma ON m.id = mma.medebe_id
    """
    params = []
    
    if section_filter:
        query += " WHERE m.section_name = %s"
        params.append(section_filter)
    
    query += " GROUP BY m.id, m.medebe_name, m.section_name, m.created_date, m.description"
    query += " ORDER BY m.section_name, m.medebe_name"
    
    cursor.execute(query, params)
    medebe_summary = cursor.fetchall()
    
    # Get overall statistics
    cursor.execute("""
        SELECT 
            COUNT(DISTINCT m.id) as total_medebes,
            COUNT(DISTINCT m.section_name) as total_sections,
            COUNT(DISTINCT mma.member_id) as total_assigned_members
        FROM medebe m
        LEFT JOIN member_medebe_assignment mma ON m.id = mma.medebe_id
    """)
    stats = cursor.fetchone()
    overall_stats = {
        'total_medebes': stats[0] or 0,
        'total_sections': stats[1] or 0,
        'total_assigned_members': stats[2] or 0
    }
    
    # Get chart data (members per medebe by section)
    cursor.execute("""
        SELECT m.section_name, m.medebe_name, COUNT(mma.id) as member_count
        FROM medebe m
        LEFT JOIN member_medebe_assignment mma ON m.id = mma.medebe_id
        GROUP BY m.section_name, m.medebe_name
        ORDER BY m.section_name, m.medebe_name
    """)
    chart_data = cursor.fetchall()
    
    # Organize chart data by section
    sections_chart = {}
    for row in chart_data:
        section = row[0]
        if section not in sections_chart:
            sections_chart[section] = {'labels': [], 'values': []}
        sections_chart[section]['labels'].append(row[1])
        sections_chart[section]['values'].append(row[2])
    
    sections = ['የሕፃናት ክፍል', 'ማህከላዊያን ክፍል', 'ወጣት ክፍል', 'ወላጅ ክፍል']
    
    cursor.close()
    conn.close()
    
    return render_template('medebe_report.html',
                         medebe_summary=medebe_summary,
                         overall_stats=overall_stats,
                         sections_chart=sections_chart,
                         sections=sections,
                         section_filter=section_filter)

@app.route('/member_medebe_report')
@login_required
def member_medebe_report():
    """Member to medebe assignment report"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filters
    section_filter = request.args.get('section', '')
    medebe_filter = request.args.get('medebe', '')
    
    # Build query
    query = """
        SELECT 
            m.full_name, m.phone, m.section_name,
            meb.medebe_name, mma.assigned_date, mma.assignment_method
        FROM member_registration m
        INNER JOIN member_medebe_assignment mma ON m.id = mma.member_id
        INNER JOIN medebe meb ON mma.medebe_id = meb.id
        WHERE 1=1
    """
    params = []
    
    if section_filter:
        query += " AND m.section_name = %s"
        params.append(section_filter)
    
    if medebe_filter:
        query += " AND mma.medebe_id = %s"
        params.append(medebe_filter)
    
    query += " ORDER BY m.section_name, meb.medebe_name, m.full_name"
    
    cursor.execute(query, params)
    member_assignments = cursor.fetchall()
    
    # Get medebes for filter
    cursor.execute("SELECT id, medebe_name, section_name FROM medebe ORDER BY section_name, medebe_name")
    medebes = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            m.section_name,
            COUNT(DISTINCT mma.member_id) as assigned_count,
            COUNT(DISTINCT meb.id) as medebe_count
        FROM member_registration m
        INNER JOIN member_medebe_assignment mma ON m.id = mma.member_id
        INNER JOIN medebe meb ON mma.medebe_id = meb.id
        GROUP BY m.section_name
    """)
    section_stats = cursor.fetchall()
    
    # Chart data for distribution
    cursor.execute("""
        SELECT meb.medebe_name, m.section_name, COUNT(mma.id) as count
        FROM medebe meb
        INNER JOIN member_medebe_assignment mma ON meb.id = mma.medebe_id
        INNER JOIN member_registration m ON mma.member_id = m.id
        GROUP BY meb.medebe_name, m.section_name
        ORDER BY m.section_name, meb.medebe_name
    """)
    chart_data = cursor.fetchall()
    
    sections = ['የሕፃናት ክፍል', 'ማህከላዊያን ክፍል', 'ወጣት ክፍል', 'ወላጅ ክፍል']
    
    cursor.close()
    conn.close()
    
    return render_template('member_medebe_report.html',
                         member_assignments=member_assignments,
                         medebes=medebes,
                         section_stats=section_stats,
                         chart_data=chart_data,
                         sections=sections,
                         section_filter=section_filter,
                         medebe_filter=medebe_filter)

@app.route('/member_accounts', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin')
def member_accounts():
    """Manage member portal accounts"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            try:
                member_id = request.form.get('member_id')
                username = request.form.get('username')
                password = request.form.get('password')
                email = request.form.get('email')
                phone = request.form.get('phone')
                
                # Check if account already exists
                cursor.execute("SELECT id FROM member_accounts WHERE member_id = %s", (member_id,))
                if cursor.fetchone():
                    flash('Account already exists for this member', 'danger')
                else:
                    # Hash password
                    import hashlib
                    password_hash = hashlib.sha256(password.encode()).hexdigest()
                    
                    cursor.execute("""
                        INSERT INTO member_accounts (member_id, username, password_hash, email, phone, created_by)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (member_id, username, password_hash, email, phone, session.get('username', 'Unknown')))
                    conn.commit()
                    flash('Member account created successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error creating account: {str(e)}', 'danger')
        
        elif action == 'generate_bulk':
            try:
                # Get all members without accounts
                cursor.execute("""
                    SELECT m.id, m.full_name, m.phone, m.email
                    FROM member_registration m
                    LEFT JOIN member_accounts ma ON m.id = ma.member_id
                    WHERE ma.id IS NULL
                """)
                members_without_accounts = cursor.fetchall()
                
                created_count = 0
                import hashlib
                
                # Default password for all members
                default_password = "12345678"
                password_hash = hashlib.sha256(default_password.encode()).hexdigest()
                
                for member in members_without_accounts:
                    member_id, full_name, phone, email = member
                    
                    # Generate username from name (remove spaces, lowercase)
                    username = full_name.replace(' ', '').lower()[:20]
                    
                    # Check if username exists, add number if needed
                    cursor.execute("SELECT COUNT(*) FROM member_accounts WHERE username LIKE %s", (f"{username}%",))
                    count = cursor.fetchone()[0]
                    if count > 0:
                        username = f"{username}{count + 1}"
                    
                    cursor.execute("""
                        INSERT INTO member_accounts (member_id, username, password_hash, email, phone, created_by)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (member_id, username, password_hash, email, phone, session.get('username', 'Unknown')))
                    created_count += 1
                
                conn.commit()
                flash(f'{created_count} member accounts created successfully with default password: 12345678', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error generating accounts: {str(e)}', 'danger')
        
        elif action == 'reset_password':
            try:
                account_id = request.form.get('account_id')
                new_password = request.form.get('new_password')
                
                import hashlib
                password_hash = hashlib.sha256(new_password.encode()).hexdigest()
                
                cursor.execute("""
                    UPDATE member_accounts 
                    SET password_hash = %s, login_attempts = 0, locked_until = NULL
                    WHERE id = %s
                """, (password_hash, account_id))
                conn.commit()
                flash('Password reset successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error resetting password: {str(e)}', 'danger')
        
        elif action == 'toggle_status':
            try:
                account_id = request.form.get('account_id')
                new_status = request.form.get('new_status')
                
                cursor.execute("UPDATE member_accounts SET account_status = %s WHERE id = %s", 
                             (new_status, account_id))
                conn.commit()
                flash(f'Account status updated to {new_status}', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error updating status: {str(e)}', 'danger')
        
        elif action == 'delete':
            try:
                account_id = request.form.get('account_id')
                cursor.execute("DELETE FROM member_accounts WHERE id = %s", (account_id,))
                conn.commit()
                flash('Account deleted successfully', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error deleting account: {str(e)}', 'danger')
        
        return redirect(url_for('member_accounts'))
    
    # Get all member accounts with member info
    cursor.execute("""
        SELECT ma.id, ma.member_id, m.full_name, m.section_name, m.phone, m.email,
               ma.username, ma.account_status, ma.is_verified, ma.last_login,
               ma.login_attempts, ma.mobile_platform, ma.created_at,
               (SELECT COUNT(*) FROM member_login_history mlh 
                WHERE mlh.member_account_id = ma.id) as login_count
        FROM member_accounts ma
        JOIN member_registration m ON ma.member_id = m.id
        ORDER BY ma.created_at DESC
    """)
    accounts = cursor.fetchall()
    
    # Get members without accounts
    cursor.execute("""
        SELECT m.id, m.full_name, m.section_name, m.phone, m.email
        FROM member_registration m
        LEFT JOIN member_accounts ma ON m.id = ma.member_id
        WHERE ma.id IS NULL
        ORDER BY m.full_name
    """)
    members_without_accounts = cursor.fetchall()
    
    # Get statistics
    cursor.execute("SELECT COUNT(*) FROM member_accounts")
    total_accounts = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM member_accounts WHERE account_status = 'Active'")
    active_accounts = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM member_accounts WHERE is_verified = 1")
    verified_accounts = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM member_accounts WHERE last_login IS NOT NULL")
    accounts_with_login = cursor.fetchone()[0]
    
    stats = [total_accounts, active_accounts, verified_accounts, accounts_with_login, len(members_without_accounts)]
    
    cursor.close()
    conn.close()
    
    return render_template('member_accounts.html',
                         accounts=accounts,
                         members_without_accounts=members_without_accounts,
                         stats=stats)

@app.route('/member_login_history/<int:account_id>')
@login_required
@role_required('Super Admin')
def member_login_history(account_id):
    """View login history for a member account"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get account info
    cursor.execute("""
        SELECT ma.id, ma.username, m.full_name, m.section_name
        FROM member_accounts ma
        JOIN member_registration m ON ma.member_id = m.id
        WHERE ma.id = %s
    """, (account_id,))
    account = cursor.fetchone()
    
    if not account:
        flash('Account not found', 'danger')
        return redirect(url_for('member_accounts'))
    
    # Get login history
    cursor.execute("""
        SELECT id, login_time, logout_time, ip_address, device_info,
               platform, app_version, location, status, failure_reason,
               session_duration
        FROM member_login_history
        WHERE member_account_id = %s
        ORDER BY login_time DESC
        LIMIT 100
    """, (account_id,))
    history = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_logins,
            COUNT(CASE WHEN status = 'Success' THEN 1 END) as successful_logins,
            COUNT(CASE WHEN status = 'Failed' THEN 1 END) as failed_logins,
            AVG(session_duration) as avg_session_duration
        FROM member_login_history
        WHERE member_account_id = %s
    """, (account_id,))
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('member_login_history.html',
                         account=account,
                         history=history,
                         stats=stats)

# ==================== POSTS / ANNOUNCEMENTS MANAGEMENT ====================

@app.route('/posts_management', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin', 'Communication Manager')
def posts_management():
    """Manage posts and announcements - CRUD operations"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get search and filter parameters
    search = request.args.get('search', '')
    post_type_filter = request.args.get('post_type', '')
    section_filter = request.args.get('section', '')
    status_filter = request.args.get('status', 'Active')
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            try:
                # Handle file upload
                attachment_path = None
                attachment_name = None
                attachment_type = None
                
                if 'attachment' in request.files:
                    file = request.files['attachment']
                    if file and file.filename:
                        from werkzeug.utils import secure_filename
                        import os
                        
                        filename = secure_filename(file.filename)
                        # Create uploads directory if it doesn't exist
                        upload_folder = os.path.join('static', 'uploads', 'posts')
                        os.makedirs(upload_folder, exist_ok=True)
                        
                        # Generate unique filename with timestamp
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        file_ext = os.path.splitext(filename)[1]
                        unique_filename = f"post_{timestamp}_{filename}"
                        file_path = os.path.join(upload_folder, unique_filename)
                        
                        file.save(file_path)
                        attachment_path = file_path.replace('\\', '/')
                        attachment_name = filename
                        
                        # Determine file type
                        if file_ext.lower() in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
                            attachment_type = 'image'
                        elif file_ext.lower() == '.pdf':
                            attachment_type = 'pdf'
                        else:
                            attachment_type = 'document'
                
                data = {
                    'post_title': request.form.get('post_title'),
                    'post_content': request.form.get('post_content'),
                    'post_type': request.form.get('post_type'),
                    'target_section': request.form.get('target_section'),
                    'target_medebe_id': request.form.get('target_medebe_id') or None,
                    'start_date': request.form.get('start_date') or None,
                    'end_date': request.form.get('end_date') or None,
                    'attachment_path': attachment_path,
                    'attachment_name': attachment_name,
                    'attachment_type': attachment_type,
                    'priority': request.form.get('priority', 'Normal'),
                    'status': 'Active',
                    'created_by': session.get('username', 'Unknown')
                }
                
                cursor.execute("""
                    INSERT INTO posts (
                        post_title, post_content, post_type, target_section, target_medebe_id,
                        start_date, end_date, attachment_path, attachment_name, attachment_type,
                        priority, status, created_by
                    ) VALUES (
                        %(post_title)s, %(post_content)s, %(post_type)s, %(target_section)s, 
                        %(target_medebe_id)s, %(start_date)s, %(end_date)s, %(attachment_path)s,
                        %(attachment_name)s, %(attachment_type)s, %(priority)s, %(status)s, %(created_by)s
                    )
                """, data)
                conn.commit()
                flash('Post created successfully! / ማስታወቂያ በተሳካ ሁኔታ ተፈጥሯል!', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error creating post: {str(e)}', 'danger')
        
        elif action == 'update':
            try:
                post_id = request.form.get('post_id')
                
                # Handle file upload for update
                attachment_path = request.form.get('existing_attachment_path')
                attachment_name = request.form.get('existing_attachment_name')
                attachment_type = request.form.get('existing_attachment_type')
                
                if 'attachment' in request.files:
                    file = request.files['attachment']
                    if file and file.filename:
                        from werkzeug.utils import secure_filename
                        import os
                        
                        filename = secure_filename(file.filename)
                        upload_folder = os.path.join('static', 'uploads', 'posts')
                        os.makedirs(upload_folder, exist_ok=True)
                        
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        file_ext = os.path.splitext(filename)[1]
                        unique_filename = f"post_{timestamp}_{filename}"
                        file_path = os.path.join(upload_folder, unique_filename)
                        
                        file.save(file_path)
                        attachment_path = file_path.replace('\\', '/')
                        attachment_name = filename
                        
                        if file_ext.lower() in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
                            attachment_type = 'image'
                        elif file_ext.lower() == '.pdf':
                            attachment_type = 'pdf'
                        else:
                            attachment_type = 'document'
                
                data = {
                    'id': post_id,
                    'post_title': request.form.get('post_title'),
                    'post_content': request.form.get('post_content'),
                    'post_type': request.form.get('post_type'),
                    'target_section': request.form.get('target_section'),
                    'target_medebe_id': request.form.get('target_medebe_id') or None,
                    'start_date': request.form.get('start_date') or None,
                    'end_date': request.form.get('end_date') or None,
                    'attachment_path': attachment_path,
                    'attachment_name': attachment_name,
                    'attachment_type': attachment_type,
                    'priority': request.form.get('priority', 'Normal'),
                    'status': request.form.get('status', 'Active')
                }
                
                cursor.execute("""
                    UPDATE posts SET
                        post_title = %(post_title)s,
                        post_content = %(post_content)s,
                        post_type = %(post_type)s,
                        target_section = %(target_section)s,
                        target_medebe_id = %(target_medebe_id)s,
                        start_date = %(start_date)s,
                        end_date = %(end_date)s,
                        attachment_path = %(attachment_path)s,
                        attachment_name = %(attachment_name)s,
                        attachment_type = %(attachment_type)s,
                        priority = %(priority)s,
                        status = %(status)s
                    WHERE id = %(id)s
                """, data)
                conn.commit()
                flash('Post updated successfully! / ማስታወቂያ በተሳካ ሁኔታ ተስተካክሏል!', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error updating post: {str(e)}', 'danger')
        
        elif action == 'delete':
            try:
                post_id = request.form.get('post_id')
                
                # Get attachment path to delete file
                cursor.execute("SELECT attachment_path FROM posts WHERE id = %s", (post_id,))
                result = cursor.fetchone()
                if result and result[0]:
                    import os
                    try:
                        if os.path.exists(result[0]):
                            os.remove(result[0])
                    except:
                        pass
                
                cursor.execute("DELETE FROM posts WHERE id = %s", (post_id,))
                conn.commit()
                flash('Post deleted successfully! / ማስታወቂያ በተሳካ ሁኔታ ተሰርዟል!', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error deleting post: {str(e)}', 'danger')
        
        return redirect(url_for('posts_management'))
    
    # Build query with filters
    query = """
        SELECT 
            p.id, p.post_title, p.post_content, p.post_type, p.target_section,
            p.target_medebe_id, m.medebe_name, p.start_date, p.end_date,
            p.attachment_path, p.attachment_name, p.attachment_type,
            p.priority, p.status, p.views_count, p.created_by, p.created_at
        FROM posts p
        LEFT JOIN medebe m ON p.target_medebe_id = m.id
        WHERE 1=1
    """
    params = []
    
    if search:
        query += " AND (p.post_title LIKE %s OR p.post_content LIKE %s)"
        search_param = f'%{search}%'
        params.extend([search_param, search_param])
    
    if post_type_filter:
        query += " AND p.post_type = %s"
        params.append(post_type_filter)
    
    if section_filter:
        query += " AND (p.target_section = %s OR p.target_section = 'All Sections')"
        params.append(section_filter)
    
    if status_filter:
        query += " AND p.status = %s"
        params.append(status_filter)
    
    query += " ORDER BY p.created_at DESC"
    
    cursor.execute(query, params)
    posts = cursor.fetchall()
    
    # Get distinct sections for filter dropdown
    cursor.execute("""
        SELECT DISTINCT section_name 
        FROM member_registration 
        WHERE section_name IS NOT NULL AND section_name != ''
        ORDER BY section_name
    """)
    sections = cursor.fetchall()
    
    # Get all medebe for dropdown
    cursor.execute("""
        SELECT id, medebe_name, section_name 
        FROM medebe 
        ORDER BY section_name, medebe_name
    """)
    medebes = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_posts,
            COUNT(CASE WHEN status = 'Active' THEN 1 END) as active_posts,
            COUNT(CASE WHEN status = 'Expired' THEN 1 END) as expired_posts,
            COUNT(CASE WHEN post_type = 'Event' THEN 1 END) as events,
            COUNT(CASE WHEN post_type = 'Announcement' THEN 1 END) as announcements,
            SUM(views_count) as total_views
        FROM posts
    """)
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('posts_management.html',
                         posts=posts,
                         sections=sections,
                         medebes=medebes,
                         stats=stats,
                         search=search,
                         post_type_filter=post_type_filter,
                         section_filter=section_filter,
                         status_filter=status_filter)


@app.route('/member_posts_view')
@login_required
def member_posts_view():
    """View posts relevant to logged-in member's section"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get member's section and medebe (if logged in as member)
    # For now, we'll show all active posts
    # In production, filter by member's section from session
    
    member_section = session.get('member_section', None)
    member_medebe_id = session.get('member_medebe_id', None)
    
    # Build query to get posts for member's section
    query = """
        SELECT 
            p.id, p.post_title, p.post_content, p.post_type, p.target_section,
            p.start_date, p.end_date, p.attachment_path, p.attachment_name, 
            p.attachment_type, p.priority, p.created_at, p.created_by
        FROM posts p
        WHERE p.status = 'Active'
        AND (
            p.target_section = 'All Sections'
            OR p.target_section = %s
            OR (p.target_medebe_id = %s AND p.target_medebe_id IS NOT NULL)
        )
        AND (
            p.start_date IS NULL 
            OR p.start_date <= CURDATE()
        )
        AND (
            p.end_date IS NULL 
            OR p.end_date >= CURDATE()
        )
        ORDER BY p.priority DESC, p.created_at DESC
    """
    
    cursor.execute(query, (member_section or 'All Sections', member_medebe_id))
    posts = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('member_posts_view.html', posts=posts)


@app.route('/mark_post_read/<int:post_id>')
@login_required
def mark_post_read(post_id):
    """Mark a post as read by the current member"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    try:
        member_id = session.get('member_id')
        if member_id:
            # Insert read status (ignore if already exists)
            cursor.execute("""
                INSERT IGNORE INTO post_read_status (post_id, member_id)
                VALUES (%s, %s)
            """, (post_id, member_id))
            
            # Increment view count
            cursor.execute("""
                UPDATE posts SET views_count = views_count + 1
                WHERE id = %s
            """, (post_id,))
            
            conn.commit()
            return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)})
    finally:
        cursor.close()
        conn.close()


@app.route('/posts_report')
@login_required
@role_required('Super Admin', 'Communication Manager', 'Report Viewer')
def posts_report():
    """Generate posts statistics and reports"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get filter parameters
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    post_type_filter = request.args.get('post_type', '')
    section_filter = request.args.get('section', '')
    
    # Overall statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total_posts,
            COUNT(CASE WHEN status = 'Active' THEN 1 END) as active_posts,
            COUNT(CASE WHEN status = 'Expired' THEN 1 END) as expired_posts,
            COUNT(CASE WHEN status = 'Draft' THEN 1 END) as draft_posts,
            COUNT(CASE WHEN post_type = 'Event' THEN 1 END) as events,
            COUNT(CASE WHEN post_type = 'Announcement' THEN 1 END) as announcements,
            COUNT(CASE WHEN post_type = 'General Info' THEN 1 END) as general_info,
            SUM(views_count) as total_views,
            AVG(views_count) as avg_views
        FROM posts
    """)
    overall_stats = cursor.fetchone()
    
    # Posts by type
    cursor.execute("""
        SELECT post_type, COUNT(*) as count, SUM(views_count) as total_views
        FROM posts
        GROUP BY post_type
        ORDER BY count DESC
    """)
    posts_by_type = cursor.fetchall()
    
    # Posts by section
    cursor.execute("""
        SELECT 
            COALESCE(target_section, 'Not Specified') as section,
            COUNT(*) as count,
            SUM(views_count) as total_views
        FROM posts
        GROUP BY target_section
        ORDER BY count DESC
    """)
    posts_by_section = cursor.fetchall()
    
    # Recent posts with details
    query = """
        SELECT 
            p.id, p.post_title, p.post_type, p.target_section,
            m.medebe_name, p.start_date, p.end_date, p.status,
            p.views_count, p.created_by, p.created_at
        FROM posts p
        LEFT JOIN medebe m ON p.target_medebe_id = m.id
        WHERE 1=1
    """
    params = []
    
    if date_from:
        query += " AND p.created_at >= %s"
        params.append(date_from)
    
    if date_to:
        query += " AND p.created_at <= %s"
        params.append(date_to + ' 23:59:59')
    
    if post_type_filter:
        query += " AND p.post_type = %s"
        params.append(post_type_filter)
    
    if section_filter:
        query += " AND p.target_section = %s"
        params.append(section_filter)
    
    query += " ORDER BY p.created_at DESC LIMIT 100"
    
    cursor.execute(query, params)
    recent_posts = cursor.fetchall()
    
    # Get most viewed posts
    cursor.execute("""
        SELECT 
            p.id, p.post_title, p.post_type, p.target_section,
            p.views_count, p.created_at
        FROM posts p
        WHERE p.views_count > 0
        ORDER BY p.views_count DESC
        LIMIT 10
    """)
    most_viewed = cursor.fetchall()
    
    # Get sections for filter
    cursor.execute("""
        SELECT DISTINCT target_section 
        FROM posts 
        WHERE target_section IS NOT NULL
        ORDER BY target_section
    """)
    sections = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('posts_report.html',
                         overall_stats=overall_stats,
                         posts_by_type=posts_by_type,
                         posts_by_section=posts_by_section,
                         recent_posts=recent_posts,
                         most_viewed=most_viewed,
                         sections=sections,
                         date_from=date_from,
                         date_to=date_to,
                         post_type_filter=post_type_filter,
                         section_filter=section_filter)


@app.route('/get_posts_for_dashboard')
@login_required
def get_posts_for_dashboard():
    """API endpoint to get posts for dashboard widget"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    # Get member's section from session (if available)
    member_section = session.get('member_section', 'All Sections')
    
    cursor.execute("""
        SELECT 
            p.id, p.post_title, p.post_type, p.post_content,
            p.priority, p.created_at, p.attachment_type
        FROM posts p
        WHERE p.status = 'Active'
        AND (
            p.target_section = 'All Sections'
            OR p.target_section = %s
        )
        AND (
            p.start_date IS NULL 
            OR p.start_date <= CURDATE()
        )
        AND (
            p.end_date IS NULL 
            OR p.end_date >= CURDATE()
        )
        ORDER BY p.priority DESC, p.created_at DESC
        LIMIT 5
    """, (member_section,))
    
    posts = []
    for row in cursor.fetchall():
        posts.append({
            'id': row[0],
            'title': row[1],
            'type': row[2],
            'content': row[3][:150] + '...' if len(row[3]) > 150 else row[3],
            'priority': row[4],
            'created_at': row[5].strftime('%Y-%m-%d %H:%M') if row[5] else '',
            'has_attachment': bool(row[6])
        })
    
    cursor.close()
    conn.close()
    
    return jsonify(posts)

# ==================== STUDY MANAGEMENT SYSTEM ====================

@app.route('/study_categories', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin', 'Study Coordinator')
def study_categories():
    """Manage study categories - CRUD operations"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    search = request.args.get('search', '')
    status_filter = request.args.get('status', 'Active')
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            try:
                cursor.execute("""
                    INSERT INTO study_categories (
                        category_name, description, status, display_order, created_by
                    ) VALUES (%s, %s, %s, %s, %s)
                """, (
                    request.form.get('category_name'),
                    request.form.get('description'),
                    'Active',
                    int(request.form.get('display_order', 0)),
                    session.get('username', 'Unknown')
                ))
                conn.commit()
                flash('Study category created successfully! / የትምህርት ምድብ በተሳካ ሁኔታ ተፈጥሯል!', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error creating category: {str(e)}', 'danger')
        
        elif action == 'update':
            try:
                cursor.execute("""
                    UPDATE study_categories SET
                        category_name = %s,
                        description = %s,
                        status = %s,
                        display_order = %s
                    WHERE id = %s
                """, (
                    request.form.get('category_name'),
                    request.form.get('description'),
                    request.form.get('status'),
                    int(request.form.get('display_order', 0)),
                    request.form.get('category_id')
                ))
                conn.commit()
                flash('Category updated successfully! / ምድብ በተሳካ ሁኔታ ተስተካክሏል!', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error updating category: {str(e)}', 'danger')
        
        elif action == 'delete':
            try:
                category_id = request.form.get('category_id')
                # Check if category has studies
                cursor.execute("SELECT COUNT(*) FROM studies WHERE category_id = %s", (category_id,))
                study_count = cursor.fetchone()[0]
                
                if study_count > 0:
                    flash(f'Cannot delete category: {study_count} studies are using this category', 'warning')
                else:
                    cursor.execute("DELETE FROM study_categories WHERE id = %s", (category_id,))
                    conn.commit()
                    flash('Category deleted successfully! / ምድብ በተሳካ ሁኔታ ተሰርዟል!', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error deleting category: {str(e)}', 'danger')
        
        return redirect(url_for('study_categories'))
    
    # Get categories
    query = "SELECT * FROM study_categories WHERE 1=1"
    params = []
    
    if search:
        query += " AND (category_name LIKE %s OR description LIKE %s)"
        search_param = f'%{search}%'
        params.extend([search_param, search_param])
    
    if status_filter:
        query += " AND status = %s"
        params.append(status_filter)
    
    query += " ORDER BY display_order, category_name"
    
    cursor.execute(query, params)
    categories = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'Active' THEN 1 ELSE 0 END) as active,
            (SELECT COUNT(*) FROM studies) as total_studies
        FROM study_categories
    """)
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('study_categories.html',
                         categories=categories,
                         stats=stats,
                         search=search,
                         status_filter=status_filter)


@app.route('/study_posting', methods=['GET', 'POST'])
@login_required
@role_required('Super Admin', 'Study Coordinator')
def study_posting():
    """Create and manage study materials with WYSIWYG editor"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    search = request.args.get('search', '')
    category_filter = request.args.get('category', '')
    audience_filter = request.args.get('audience', '')
    status_filter = request.args.get('status', 'Published')
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            try:
                # Handle file upload
                attachment_path = None
                attachment_name = None
                attachment_type = None
                
                if 'attachment' in request.files:
                    file = request.files['attachment']
                    if file and file.filename:
                        from werkzeug.utils import secure_filename
                        import os
                        
                        filename = secure_filename(file.filename)
                        upload_folder = os.path.join('static', 'uploads', 'studies')
                        os.makedirs(upload_folder, exist_ok=True)
                        
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        file_ext = os.path.splitext(filename)[1]
                        unique_filename = f"study_{timestamp}_{filename}"
                        file_path = os.path.join(upload_folder, unique_filename)
                        
                        file.save(file_path)
                        attachment_path = file_path.replace('\\', '/')
                        attachment_name = filename
                        
                        # Determine file type
                        if file_ext.lower() in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
                            attachment_type = 'image'
                        elif file_ext.lower() == '.pdf':
                            attachment_type = 'pdf'
                        elif file_ext.lower() in ['.mp3', '.wav', '.m4a']:
                            attachment_type = 'audio'
                        elif file_ext.lower() in ['.mp4', '.avi', '.mov']:
                            attachment_type = 'video'
                        else:
                            attachment_type = 'document'
                
                cursor.execute("""
                    INSERT INTO studies (
                        study_title, category_id, target_audience, content_body, summary,
                        attachment_path, attachment_name, attachment_type, publish_date,
                        author, status, priority, tags, created_by
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    request.form.get('study_title'),
                    request.form.get('category_id'),
                    request.form.get('target_audience'),
                    request.form.get('content_body'),
                    request.form.get('summary'),
                    attachment_path,
                    attachment_name,
                    attachment_type,
                    request.form.get('publish_date') or None,
                    request.form.get('author'),
                    request.form.get('status', 'Published'),
                    request.form.get('priority', 'Normal'),
                    request.form.get('tags'),
                    session.get('username', 'Unknown')
                ))
                conn.commit()
                flash('Study material created successfully! / የትምህርት ጽሑፍ በተሳካ ሁኔታ ተፈጥሯል!', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error creating study: {str(e)}', 'danger')
        
        elif action == 'update':
            try:
                study_id = request.form.get('study_id')
                
                # Handle file upload for update
                attachment_path = request.form.get('existing_attachment_path')
                attachment_name = request.form.get('existing_attachment_name')
                attachment_type = request.form.get('existing_attachment_type')
                
                if 'attachment' in request.files:
                    file = request.files['attachment']
                    if file and file.filename:
                        from werkzeug.utils import secure_filename
                        import os
                        
                        filename = secure_filename(file.filename)
                        upload_folder = os.path.join('static', 'uploads', 'studies')
                        os.makedirs(upload_folder, exist_ok=True)
                        
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        file_ext = os.path.splitext(filename)[1]
                        unique_filename = f"study_{timestamp}_{filename}"
                        file_path = os.path.join(upload_folder, unique_filename)
                        
                        file.save(file_path)
                        attachment_path = file_path.replace('\\', '/')
                        attachment_name = filename
                        
                        if file_ext.lower() in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
                            attachment_type = 'image'
                        elif file_ext.lower() == '.pdf':
                            attachment_type = 'pdf'
                        elif file_ext.lower() in ['.mp3', '.wav', '.m4a']:
                            attachment_type = 'audio'
                        elif file_ext.lower() in ['.mp4', '.avi', '.mov']:
                            attachment_type = 'video'
                        else:
                            attachment_type = 'document'
                
                cursor.execute("""
                    UPDATE studies SET
                        study_title = %s,
                        category_id = %s,
                        target_audience = %s,
                        content_body = %s,
                        summary = %s,
                        attachment_path = %s,
                        attachment_name = %s,
                        attachment_type = %s,
                        publish_date = %s,
                        author = %s,
                        status = %s,
                        priority = %s,
                        tags = %s
                    WHERE id = %s
                """, (
                    request.form.get('study_title'),
                    request.form.get('category_id'),
                    request.form.get('target_audience'),
                    request.form.get('content_body'),
                    request.form.get('summary'),
                    attachment_path,
                    attachment_name,
                    attachment_type,
                    request.form.get('publish_date') or None,
                    request.form.get('author'),
                    request.form.get('status'),
                    request.form.get('priority'),
                    request.form.get('tags'),
                    study_id
                ))
                conn.commit()
                flash('Study updated successfully! / ትምህርት በተሳካ ሁኔታ ተስተካክሏል!', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error updating study: {str(e)}', 'danger')
        
        elif action == 'delete':
            try:
                study_id = request.form.get('study_id')
                
                # Get attachment path to delete file
                cursor.execute("SELECT attachment_path FROM studies WHERE id = %s", (study_id,))
                result = cursor.fetchone()
                if result and result[0]:
                    import os
                    try:
                        if os.path.exists(result[0]):
                            os.remove(result[0])
                    except:
                        pass
                
                cursor.execute("DELETE FROM studies WHERE id = %s", (study_id,))
                conn.commit()
                flash('Study deleted successfully! / ትምህርት በተሳካ ሁኔታ ተሰርዟል!', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Error deleting study: {str(e)}', 'danger')
        
        return redirect(url_for('study_posting'))
    
    # Get studies
    query = """
        SELECT 
            s.id, s.study_title, s.category_id, sc.category_name, s.target_audience,
            s.content_body, s.summary, s.attachment_path, s.attachment_name, s.attachment_type,
            s.publish_date, s.author, s.status, s.priority, s.views_count, s.downloads_count,
            s.is_featured, s.tags, s.created_by, s.created_at
        FROM studies s
        JOIN study_categories sc ON s.category_id = sc.id
        WHERE 1=1
    """
    params = []
    
    if search:
        query += " AND (s.study_title LIKE %s OR s.content_body LIKE %s OR s.author LIKE %s)"
        search_param = f'%{search}%'
        params.extend([search_param, search_param, search_param])
    
    if category_filter:
        query += " AND s.category_id = %s"
        params.append(category_filter)
    
    if audience_filter:
        query += " AND s.target_audience = %s"
        params.append(audience_filter)
    
    if status_filter:
        query += " AND s.status = %s"
        params.append(status_filter)
    
    query += " ORDER BY s.is_featured DESC, s.created_at DESC"
    
    cursor.execute(query, params)
    studies = cursor.fetchall()
    
    # Get categories for dropdown
    cursor.execute("SELECT id, category_name FROM study_categories WHERE status = 'Active' ORDER BY display_order, category_name")
    categories = cursor.fetchall()
    
    # Get sections for audience filter
    cursor.execute("SELECT DISTINCT section_name FROM member_registration WHERE section_name IS NOT NULL ORDER BY section_name")
    sections = cursor.fetchall()
    
    # Get statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'Published' THEN 1 ELSE 0 END) as published,
            SUM(CASE WHEN status = 'Draft' THEN 1 ELSE 0 END) as drafts,
            SUM(views_count) as total_views,
            SUM(downloads_count) as total_downloads
        FROM studies
    """)
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    return render_template('study_posting.html',
                         studies=studies,
                         categories=categories,
                         sections=sections,
                         stats=stats,
                         search=search,
                         category_filter=category_filter,
                         audience_filter=audience_filter,
                         status_filter=status_filter)

@app.route('/get_study/<int:study_id>', methods=['GET'])
@login_required
@role_required('Super Admin', 'Study Coordinator')
def get_study(study_id):
    """Get single study details for editing"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    try:
        cursor.execute("""
            SELECT 
                s.id, s.study_title, s.category_id, sc.category_name, s.target_audience,
                s.content_body, s.summary, s.attachment_path, s.attachment_name, s.attachment_type,
                s.publish_date, s.author, s.status, s.priority, s.views_count, s.downloads_count,
                s.is_featured, s.tags
            FROM studies s
            JOIN study_categories sc ON s.category_id = sc.id
            WHERE s.id = %s
        """, (study_id,))
        
        study = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if study:
            return jsonify({
                'success': True,
                'study': {
                    'id': study[0],
                    'study_title': study[1],
                    'category_id': study[2],
                    'category_name': study[3],
                    'target_audience': study[4],
                    'content_body': study[5],
                    'summary': study[6],
                    'attachment_path': study[7],
                    'attachment_name': study[8],
                    'attachment_type': study[9],
                    'publish_date': str(study[10]) if study[10] else '',
                    'author': study[11],
                    'status': study[12],
                    'priority': study[13],
                    'views_count': study[14],
                    'downloads_count': study[15],
                    'is_featured': study[16],
                    'tags': study[17]
                }
            })
        else:
            return jsonify({'success': False, 'error': 'Study not found'}), 404
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/study_materials_view')
@login_required
def study_materials_view():
    """View published study materials (member-facing)"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    search = request.args.get('search', '')
    category_filter = request.args.get('category', '')
    member_section = session.get('member_section', 'All Members')
    
    query = """
        SELECT 
            s.id, s.study_title, s.category_id, sc.category_name, s.target_audience,
            s.content_body, s.summary, s.attachment_path, s.attachment_name, s.attachment_type,
            s.publish_date, s.author, s.views_count, s.downloads_count, s.is_featured,
            s.tags, s.created_at
        FROM studies s
        JOIN study_categories sc ON s.category_id = sc.id
        WHERE s.status = 'Published'
        AND (s.target_audience = 'All Members' OR s.target_audience = %s)
    """
    params = [member_section]
    
    if search:
        query += " AND (s.study_title LIKE %s OR s.content_body LIKE %s OR s.tags LIKE %s)"
        search_param = f'%{search}%'
        params.extend([search_param, search_param, search_param])
    
    if category_filter:
        query += " AND s.category_id = %s"
        params.append(category_filter)
    
    query += " ORDER BY s.is_featured DESC, s.publish_date DESC, s.created_at DESC"
    
    cursor.execute(query, params)
    studies = cursor.fetchall()
    
    # Get categories for filter
    cursor.execute("SELECT id, category_name FROM study_categories WHERE status = 'Active' ORDER BY display_order")
    categories = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('study_materials_view.html',
                         studies=studies,
                         categories=categories,
                         search=search,
                         category_filter=category_filter)


@app.route('/study_details/<int:study_id>')
@login_required
def study_details(study_id):
    """View full study details"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    try:
        cursor.execute("""
            SELECT 
                s.id, s.study_title, s.category_id, sc.category_name, s.target_audience,
                s.content_body, s.summary, s.attachment_path, s.attachment_name, s.attachment_type,
                s.publish_date, s.author, s.views_count, s.downloads_count, s.is_featured,
                s.tags, s.created_at
            FROM studies s
            JOIN study_categories sc ON s.category_id = sc.id
            WHERE s.id = %s
        """, (study_id,))
        
        study = cursor.fetchone()
        
        if not study:
            flash('Study not found', 'warning')
            return redirect(url_for('study_materials_view'))
        
        # Increment view count
        cursor.execute("UPDATE studies SET views_count = views_count + 1 WHERE id = %s", (study_id,))
        
        # Mark as read if member_id available
        member_id = session.get('member_id')
        if member_id:
            cursor.execute("""
                INSERT IGNORE INTO study_read_status (study_id, member_id)
                VALUES (%s, %s)
            """, (study_id, member_id))
        
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return render_template('study_details.html', study=study)
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'Error loading study: {str(e)}', 'danger')
        return redirect(url_for('study_materials_view'))


@app.route('/download_study_attachment/<int:study_id>')
@login_required
def download_study_attachment(study_id):
    """Download study attachment"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    try:
        cursor.execute("SELECT attachment_path, attachment_name FROM studies WHERE id = %s", (study_id,))
        result = cursor.fetchone()
        
        if result and result[0]:
            # Increment download count
            cursor.execute("UPDATE studies SET downloads_count = downloads_count + 1 WHERE id = %s", (study_id,))
            conn.commit()
            
            return send_file(result[0], as_attachment=True, download_name=result[1])
        else:
            flash('Attachment not found', 'warning')
            return redirect(url_for('study_materials_view'))
    except Exception as e:
        flash(f'Error downloading attachment: {str(e)}', 'danger')
        return redirect(url_for('study_materials_view'))
    finally:
        cursor.close()
        conn.close()


@app.route('/study_reports')
@login_required
@role_required('Super Admin', 'Study Coordinator', 'Report Viewer')
def study_reports():
    """Generate study statistics and reports"""
    conn = get_db_connection()
    cursor = conn.cursor(buffered=True)
    
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    category_filter = request.args.get('category', '')
    
    # Overall statistics
    cursor.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'Published' THEN 1 ELSE 0 END) as published,
            SUM(CASE WHEN status = 'Draft' THEN 1 ELSE 0 END) as drafts,
            SUM(views_count) as total_views,
            SUM(downloads_count) as total_downloads,
            AVG(views_count) as avg_views
        FROM studies
    """)
    overall_stats = cursor.fetchone()
    
    # Studies by category
    cursor.execute("""
        SELECT sc.category_name, COUNT(*) as count, SUM(s.views_count) as views
        FROM studies s
        JOIN study_categories sc ON s.category_id = sc.id
        GROUP BY sc.category_name
        ORDER BY count DESC
    """)
    by_category = cursor.fetchall()
    
    # Studies by audience
    cursor.execute("""
        SELECT target_audience, COUNT(*) as count
        FROM studies
        GROUP BY target_audience
        ORDER BY count DESC
    """)
    by_audience = cursor.fetchall()
    
    # Recent studies
    query = """
        SELECT 
            s.id, s.study_title, sc.category_name, s.target_audience, s.author,
            s.publish_date, s.status, s.views_count, s.downloads_count, s.created_at
        FROM studies s
        JOIN study_categories sc ON s.category_id = sc.id
        WHERE 1=1
    """
    params = []
    
    if date_from:
        query += " AND s.created_at >= %s"
        params.append(date_from)
    
    if date_to:
        query += " AND s.created_at <= %s"
        params.append(date_to + ' 23:59:59')
    
    if category_filter:
        query += " AND s.category_id = %s"
        params.append(category_filter)
    
    query += " ORDER BY s.created_at DESC LIMIT 100"
    
    cursor.execute(query, params)
    recent_studies = cursor.fetchall()
    
    # Get categories for filter
    cursor.execute("SELECT id, category_name FROM study_categories ORDER BY category_name")
    categories = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('study_reports.html',
                         overall_stats=overall_stats,
                         by_category=by_category,
                         by_audience=by_audience,
                         recent_studies=recent_studies,
                         categories=categories,
                         date_from=date_from,
                         date_to=date_to,
                         category_filter=category_filter)

if __name__ == '__main__':
    # Initialize database and tables
    initialize_app()
    
    # Run the Flask application
    app.run(
        host=Config.HOST,
        port=Config.PORT,
        debug=Config.DEBUG
    )

